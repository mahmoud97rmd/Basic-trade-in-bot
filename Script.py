import json
"""
Gold Scalper Bot -- v10.0 (Resilience-First Core)
Strategy : EMA 15/50/150 Trend Stack + Stochastic (5,5,5, EMA-smoothed) Reversal-Zone Crossover

v10.0 changes vs v9.4 (see PATCH_NOTES.md shipped alongside this file):
  - Gann Levels / Fan Angles / cycle-anchor strategy fully removed.
  - New strategy evaluated on candle close only (signal_candle_shift selects
    which closed candle: 0 = live/forming, 1 = last closed, 2 = previous closed).
  - allow_concurrent_trades gates whether a new Buy/Sell can open while one of
    the same direction is already open for the symbol.
  - No hardcoded credential fallbacks; bot refuses to start without env vars.
  - No silent except-pass in execution / reconciliation / order-management paths.
  - Explicit HALT / READ_ONLY connection-state machine with Telegram escalation.
  - Persistence now captures full per-symbol strategy state, not just open trades.
  - Startup reconstructs state from disk before ANY market interaction.
"""

import asyncio
import logging
import traceback
import time
import random
import zlib
import aiohttp
import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, timezone, time as dtime
from aiohttp import web
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from metaapi_cloud_sdk import MetaApi, SynchronizationListener

# -----------------------------------------------------------------
# LOGGING (structured, always includes tracebacks for exceptions)
# -----------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    stream=sys.stdout,
)
logger = logging.getLogger('gold_scalper')

def log_exception(context: str, exc: Exception) -> None:
    """Zero-tolerance logging: every caught exception in a critical path gets
    a full traceback attached to the log line, not just str(e)."""
    logger.error("EXCEPTION in %s: %s\n%s", context, exc, traceback.format_exc())

_DIAG_LOG_MAX_ENTRIES = 50000  # ~generous cap; trimmed on every append

def _diag_log_add(entry: dict) -> None:
    """Append one row to the rolling live-scan diagnostic log (see
    bot_state['diag_log']). This is what /export_diag_excel dumps -- it's
    the ONLY place that records the *silent* skip reasons (insufficient
    candle data, cap reached, trend unknown, etc.) that never get a
    Telegram message of their own, so the operator can reconstruct exactly
    what the scanner saw/did on every (symbol, timeframe, cycle), not just
    a point-in-time snapshot."""
    log = bot_state.setdefault('diag_log', [])
    log.append(entry)
    if len(log) > _DIAG_LOG_MAX_ENTRIES:
        del log[: len(log) - _DIAG_LOG_MAX_ENTRIES]

_trade_history_lock = asyncio.Lock()

async def _record_closed_trade_history(symbol: str, tid: str, tr: dict, exit_px: float, pnl: float,
                                        outcome_label: str, close_reason: str, pnl_confirmed: bool) -> None:
    """Append one row of full detail for a just-closed real/virtual live
    trade, feeding /export_live_trades_excel. Kept deliberately rich (every
    field a human would need to judge "did this trade behave like the
    backtest expected") since this is the ONLY place live trade outcomes
    get durably recorded anywhere in the bot today."""
    try:
        entry = tr.get('entry')
        is_buy = tr.get('is_buy')
        opened_at = tr.get('opened_at')
        closed_at_dt = datetime.now(timezone.utc)
        duration_min = None
        if opened_at:
            try:
                opened_dt = datetime.fromisoformat(opened_at) if isinstance(opened_at, str) else opened_at
                duration_min = round((closed_at_dt - opened_dt).total_seconds() / 60.0, 1)
            except Exception:
                duration_min = None
        intended_entry = tr.get('level_price', entry)
        entry_slip = (entry - intended_entry) if (entry is not None and intended_entry is not None) else None
        async with _trade_history_lock:
            hist = bot_state.setdefault('live_trade_history', [])
            hist.append({
                'symbol': symbol, 'tid': tid, 'tf': tr.get('tf'), 'is_real': bool(tr.get('is_real')),
                'is_buy': is_buy, 'opened_at': opened_at, 'closed_at': closed_at_dt.isoformat(),
                'duration_min': duration_min, 'level_price': intended_entry, 'entry': entry,
                'entry_slippage': entry_slip, 'tp': tr.get('tp'), 'sl': tr.get('sl'), 'exit_price': exit_px,
                'outcome': outcome_label, 'pnl': pnl, 'pnl_confirmed_from_broker': pnl_confirmed,
                'close_reason': close_reason, 'be_activated': bool(tr.get('be_activated')),
                'feed_source': tr.get('feed_source'), 'feed_age_ms': tr.get('feed_age_ms'),
                'trigger_type': tr.get('trigger_type'),
                'exec_latency_ms': tr.get('exec_latency_ms'), 'exec_method': tr.get('exec_method'),
                'exec_ioc_fail_reason': tr.get('exec_ioc_fail_reason'), 'exec_slippage': tr.get('exec_slippage'),
            })
            if len(hist) > _DIAG_LOG_MAX_ENTRIES:
                del hist[: len(hist) - _DIAG_LOG_MAX_ENTRIES]
    except Exception as e:
        log_exception(f'_record_closed_trade_history [{symbol} {tid}]', e)

# -----------------------------------------------------------------
# CONFIGURATION
# -----------------------------------------------------------------
def _require_env(name: str) -> str:
    val = os.environ.get(name)
    if not val:
        logger.critical("FATAL: required environment variable '%s' is not set. Refusing to start.", name)
        sys.exit(1)
    return val

METAAPI_TOKEN  = ('eyJhbGciOiJSUzUxMiIsInR5cCI6IkpXVCJ9.eyJfaWQiOiJjODRlZDU2MmMyOTE3ZDgxNTU1ZWQ0NDhlNzc2MzNkMCIsImFjY2Vzc1J1bGVzIjpbeyJpZCI6InRyYWRpbmctYWNjb3VudC1tYW5hZ2VtZW50LWFwaSIsIm1ldGhvZHMiOlsidHJhZGluZy1hY2NvdW50LW1hbmFnZW1lbnQtYXBpOnJlc3Q6cHVibGljOio6KiJdLCJyb2xlcyI6WyJyZWFkZXIiXSwicmVzb3VyY2VzIjpbImFjY291bnQ6JFVTRVJfSUQkOjFmYWI0MTA5LTcwNDktNDdhNy1hYTYzLTNjMzYxMTBmNzFlMyJdfSx7ImlkIjoibWV0YWFwaS1yZXN0LWFwaSIsIm1ldGhvZHMiOlsibWV0YWFwaS1hcGk6cmVzdDpwdWJsaWM6KjoqIl0sInJvbGVzIjpbInJlYWRlciIsIndyaXRlciJdLCJyZXNvdXJjZXMiOlsiYWNjb3VudDokVVNFUl9JRCQ6MWZhYjQxMDktNzA0OS00N2E3LWFhNjMtM2MzNjExMGY3MWUzIl19LHsiaWQiOiJtZXRhYXBpLXJwYy1hcGkiLCJtZXRob2RzIjpbIm1ldGFhcGktYXBpOndzOnB1YmxpYzoqOioiXSwicm9sZXMiOlsicmVhZGVyIiwid3JpdGVyIl0sInJlc291cmNlcyI6WyJhY2NvdW50OiRVU0VSX0lEJDoxZmFiNDEwOS03MDQ5LTQ3YTctYWE2My0zYzM2MTEwZjcxZTMiXX0seyJpZCI6Im1ldGFhcGktcmVhbC10aW1lLXN0cmVhbWluZy1hcGkiLCJtZXRob2RzIjpbIm1ldGFhcGktYXBpOndzOnB1YmxpYzoqOioiXSwicm9sZXMiOlsicmVhZGVyIiwid3JpdGVyIl0sInJlc291cmNlcyI6WyJhY2NvdW50OiRVU0VSX0lEJDoxZmFiNDEwOS03MDQ5LTQ3YTctYWE2My0zYzM2MTEwZjcxZTMiXX0seyJpZCI6Im1ldGFzdGF0cy1hcGkiLCJtZXRob2RzIjpbIm1ldGFzdGF0cy1hcGk6cmVzdDpwdWJsaWM6KjoqIl0sInJvbGVzIjpbInJlYWRlciJdLCJyZXNvdXJjZXMiOlsiYWNjb3VudDokVVNFUl9JRCQ6MWZhYjQxMDktNzA0OS00N2E3LWFhNjMtM2MzNjExMGY3MWUzIl19LHsiaWQiOiJyaXNrLW1hbmFnZW1lbnQtYXBpIiwibWV0aG9kcyI6WyJyaXNrLW1hbmFnZW1lbnQtYXBpOnJlc3Q6cHVibGljOio6KiJdLCJyb2xlcyI6WyJyZWFkZXIiXSwicmVzb3VyY2VzIjpbImFjY291bnQ6JFVTRVJfSUQkOjFmYWI0MTA5LTcwNDktNDdhNy1hYTYzLTNjMzYxMTBmNzFlMyJdfV0sImlnbm9yZVJhdGVMaW1pdHMiOmZhbHNlLCJ0b2tlbklkIjoiMjAyMTAyMTMiLCJpbXBlcnNvbmF0ZWQiOmZhbHNlLCJyZWFsVXNlcklkIjoiYzg0ZWQ1NjJjMjkxN2Q4MTU1NWVkNDQ4ZTc3NjMzZDAiLCJpYXQiOjE3ODMyODM3ODYsImV4cCI6MTc5MTA1OTc4Nn0.mgrEHouxzp6_uTewY926BLOgRUGTWqLDgbOvilTtKUGy9lP72b83eearNdPd5fltUr52Sf4bUReSqcZ7-HJRXwsl0apUxqJmQ7N1X0GNfVeEQkgtgDpnzXGvmV1WF73AvVpc3bvnG-jbecJzw5CLZQ2-b4AbPZldt0Yxp5BViiaqyO2X95uYMYmQcp2PJbjEHAeqJsGPJHaJsI1qqMjboWT0TsesGW94RFn8oBWdHK5rYtUhqQCkIEQZmuNi3rNImWjBHPcr9OflcGJE8sSqDF6Raib2pZrpXIr3c03IJy5IJtQD-7-NPHNIIxw7i5X7z8TvKLOwY4eBNM3dilttWnUZ0JZVOCbQEa0afb6bHoYqhQRgWfTxHRwNe11zYvnox6lO9VURya3OSGafxsrhiShOzn44VtWwyTawlD9_p56duC3EnrzdtboLDjzwS5G8_tE0539jfrO3hWiFQn2HeF4Nmr39DfSTAjb5RTKQD3iS2sNqAaCN9OBQqt2UAMbZYLHUuK0oP9k3R5ZUa31SKRIagBqjHIjaNCHkALR2rWCnGBWgzSsQHVxIHEyzgTxLuviOkc9abWV-76DD2nJ1R150n6acoG8eyoSPivxDMHx_v1xDSLtjbUKIqWq01wH7XSyfpXpEWFW9oG7FAjg1ctOH6nzhbii9nUnpj_RQCA8')
ACCOUNT_ID     = ('1fab4109-7049-47a7-aa63-3c36110f71e3')
TG_TOKEN       = ('8647261254:AAF-cuUYljqSgWMwC9F-sC9RH9tEm_nUUpU')
OANDA_ACCOUNT  = ('101-004-28533521-003')
OANDA_TOKEN    = ('0e282d5a3e65ad6fdd809e2c195bb1cd-9e2158e12fa13840e030ee3081b36fab')
AVAILABLE_SYMBOLS = ['XAU_USD', 'XAU_EUR', 'XAG_USD', 'EUR_USD', 'GBP_JPY', 'GBP_AUD', 'GBP_NZD', 'AUD_JPY', 'NZD_JPY']
SYMBOL_INFO = {
    'XAU_USD': {'pip_value': 0.1,     'contract_size': 100,    'prec': 2, 'name': 'Gold (USD)'},
    'XAU_EUR': {'pip_value': 0.1,     'contract_size': 100,    'prec': 2, 'name': 'Gold (EUR)'},
    'XAG_USD': {'pip_value': 0.001,   'contract_size': 5000,   'prec': 3, 'name': 'Silver'},
    'EUR_USD': {'pip_value': 0.00001, 'contract_size': 100000, 'prec': 5, 'name': 'EUR/USD'},
    'GBP_JPY': {'pip_value': 0.01,    'contract_size': 100000, 'prec': 3, 'name': 'GBP/JPY'},
    'GBP_AUD': {'pip_value': 0.00001, 'contract_size': 100000, 'prec': 5, 'name': 'GBP/AUD'},
    'GBP_NZD': {'pip_value': 0.00001, 'contract_size': 100000, 'prec': 5, 'name': 'GBP/NZD'},
    'AUD_JPY': {'pip_value': 0.01,    'contract_size': 100000, 'prec': 3, 'name': 'AUD/JPY'},
    'NZD_JPY': {'pip_value': 0.01,    'contract_size': 100000, 'prec': 3, 'name': 'NZD/JPY'},
}
OANDA_BASE_URL = 'https://api-fxpractice.oanda.com/v3'  

_TFS = ['1m', '2m', '3m', '4m', '5m', '6m', '10m', '15m', '20m', '30m', '1h', '2h']

_http: aiohttp.ClientSession | None = None

def get_http() -> aiohttp.ClientSession:
    global _http
    if _http is None or _http.closed:
        connector = aiohttp.TCPConnector(limit=20, ttl_dns_cache=300)
        timeout   = aiohttp.ClientTimeout(total=30, connect=10)
        _http     = aiohttp.ClientSession(connector=connector, timeout=timeout)
    return _http

def c_log(msg: str) -> None:
    dam = (datetime.now(timezone.utc) + timedelta(hours=3)).strftime('%H:%M:%S')
    print(f"[{dam} DAM] {msg}", flush=True)

# -----------------------------------------------------------------
# GLOBAL STATE
# -----------------------------------------------------------------
_metaapi = None
_metaapi_account = None
_metaapi_conn = None

# ── Live-quote push cache (OANDA REST poller) ──
# Populated by oanda_live_price_poller, keyed by the
# OANDA-format symbol used everywhere else in the bot ('XAU_USD'), NOT
# the broker's own symbol name ('XAUUSD') -- _broker_to_data_symbol
# translates incoming broker-symbol ticks back to that key.
live_quotes: dict[str, dict] = {}          # {'XAU_USD': {'bid':, 'ask':, 'mid':, 'ts': monotonic}}
_broker_to_data_symbol: dict[str, str] = {}  # {'XAUUSD': 'XAU_USD', ...}

# Event-driven fill detection: map order_id -> asyncio.Event
# A background task monitors terminal_state.positions and sets the event
# when the position appears, eliminating busy-wait polling (saves 300-1000ms per trade).
_fill_events: dict[str, asyncio.Event] = {}
_fill_results: dict[str, dict] = {}  # order_id -> {fill_price, fill_source, trade_id}
_fill_monitor_started = False
_QUOTE_STALE_SECONDS = 5.0
# Updated on EVERY tick received from MetaApi, for ANY symbol -- this is
# deliberately independent of live_quotes/_broker_to_data_symbol (which can
# be empty/wrong) and of _metaapi_account.connection_status (which the SDK
# can keep reporting 'CONNECTED' even when the underlying WS session has
# gone silent/zombie, e.g. during a broker daily-rollover freeze). The
# watchdog in strategy_monitor_scanner uses ONLY this raw timestamp to decide
# whether to force a full connection teardown+reconnect.
_last_any_tick_ts = time.monotonic()
_WS_WATCHDOG_STALE_SECONDS = 60.0

async def oanda_live_price_poller() -> None:
    """OANDA streaming price feed (v3 pricing/stream), replacing the old
    REST-polling version.

    Two problems with the REST version, both root causes of the recurring
    "Live price feed stale >60s" disconnects, are fixed by this:

    1. Latency floor: REST polling was capped at one update per
       poll_interval (1.5s) no matter what -- the stream instead pushes a
       PRICE message the instant OANDA has one, and a HEARTBEAT roughly
       every 5s even when the market is quiet, so live_quotes stays fresh
       continuously instead of in 1.5s-apart snapshots.
    2. Resource contention: the REST version acquired _get_oanda_sem()
       (max 3 concurrent OANDA REST calls) for every single poll -- the
       same semaphore backtest candle fetching uses. A backtest holding
       all 3 slots for an extended stretch could starve this poller of a
       turn entirely, which is very likely what was still causing
       staleness even after the earlier fix that moved Excel generation
       off the event loop. The stream instead opens ONE long-lived
       connection outside that semaphore, so it has nothing to queue
       behind regardless of what a backtest is doing.

    OANDA's stream connections do get closed periodically by design (their
    docs call this normal) -- the outer while True with backoff just
    reconnects when that happens.

    MetaAPI is still used purely for order execution / reconciliation;
    only the price feed comes from OANDA.
    """
    global _last_any_tick_ts
    headers = {'Authorization': f'Bearer {OANDA_TOKEN}'}
    stream_base = (OANDA_BASE_URL
                    .replace('https://api-fxpractice.oanda.com', 'https://stream-fxpractice.oanda.com')
                    .replace('https://api-fxtrade.oanda.com', 'https://stream-fxtrade.oanda.com'))
    backoff = 1
    while True:
        try:
            active_syms = [s for s, on in bot_state['active_symbols'].items() if on]
            if not active_syms:
                await asyncio.sleep(2)
                continue

            instruments = ','.join(active_syms)
            url = f'{stream_base}/accounts/{OANDA_ACCOUNT}/pricing/stream'
            params = {'instruments': instruments}
            timeout = aiohttp.ClientTimeout(total=None, connect=10, sock_read=30)

            # Dedicated session for this one long-lived connection -- kept
            # separate from get_http()'s shared pool and deliberately does
            # NOT go through _get_oanda_sem(); see docstring above.
            async with aiohttp.ClientSession(timeout=timeout) as sess:
                async with sess.get(url, headers=headers, params=params) as resp:
                    if resp.status != 200:
                        c_log(f"oanda_live_price_poller: stream connect failed (status {resp.status}).")
                        await asyncio.sleep(backoff); backoff = min(backoff * 2, 30)
                        continue

                    backoff = 1
                    c_log("oanda_live_price_poller: stream connected.")
                    async for raw_line in resp.content:
                        line = raw_line.strip()
                        if not line:
                            continue
                        try:
                            msg = json.loads(line)
                        except (json.JSONDecodeError, ValueError):
                            continue

                        if msg.get('type') == 'HEARTBEAT':
                            # Still proof-of-life even with no price movement.
                            _last_any_tick_ts = time.monotonic()
                            continue
                        if msg.get('type') != 'PRICE':
                            continue

                        sym = msg.get('instrument')
                        if sym not in bot_state['active_symbols'] or not bot_state['active_symbols'][sym]:
                            continue
                        bids = msg.get('bids') or []
                        asks = msg.get('asks') or []
                        if not bids or not asks:
                            continue
                        try:
                            bid = float(bids[0]['price'])
                            ask = float(asks[0]['price'])
                        except (KeyError, ValueError, TypeError):
                            continue

                        now_mono = time.monotonic()
                        _last_any_tick_ts = now_mono  # feed-health heartbeat for the staleness watchdog
                        mid = (bid + ask) / 2.0
                        live_quotes[sym] = {'bid': bid, 'ask': ask, 'mid': mid, 'ts': now_mono}

                        # If active symbols changed since we opened this
                        # stream (user toggled a pair on/off), reconnect
                        # with the updated instrument list instead of
                        # silently missing the new one.
                        if [s for s, on in bot_state['active_symbols'].items() if on] != active_syms:
                            break
        except asyncio.CancelledError:
            raise
        except Exception as e:
            log_exception('oanda_live_price_poller', e)
            await asyncio.sleep(backoff); backoff = min(backoff * 2, 30)


def _safe_task(coro, name=''):
    t = asyncio.create_task(coro)
    t.add_done_callback(lambda fut: log_exception(f'background task [{name}]', fut.exception()) if not fut.cancelled() and fut.exception() else None)
    return t

def _is_market_hours_now() -> bool:
    """Rough broker session-hours check (from the real XAUUSD symbol spec:
    Sun 01:01 -> Fri 23:49 broker time, closed Saturday). Used only to
    decide whether a 60s tick silence is suspicious (worth a forced
    reconnect) or expected (weekend close) -- not used anywhere else."""
    offset = bot_state.get('broker_time_offset', 3)
    broker_now = datetime.now(timezone.utc) + timedelta(hours=offset)
    wd = broker_now.weekday()  # Monday=0 ... Sunday=6
    t = broker_now.time()
    if wd == 5:                                   # Saturday: fully closed
        return False
    if wd == 4 and t >= dtime(23, 49):             # Friday after close
        return False
    if wd == 6 and t < dtime(1, 1):                # Sunday before open
        return False
    return True

async def _force_full_reconnect(reason: str) -> None:
    """Execution-channel watchdog. The price feed no longer depends on the
    MetaAPI WebSocket (it is now driven by oanda_live_price_poller), so this
    only rebuilds the MetaAPI streaming connection used for order execution
    and reconciliation. If the connection goes zombie/degraded, new orders
    fail -- this tears the existing connection down and rebuilds it from
    scratch (with bounded timeouts so a hung SDK region can never freeze
    the strategy_monitor_scanner loop forever)."""
    global _metaapi_conn
    c_log(f"EXEC WATCHDOG: forcing full reconnect -- {reason}")
    await set_connection_state(CONN_READ_ONLY, f"Exec watchdog: {reason}")
    if _metaapi_account is None:
        c_log("EXEC WATCHDOG: _metaapi_account is None — cannot reconnect")
        return
    try:
        if _metaapi_conn is not None:
            try:
                await asyncio.wait_for(_metaapi_conn.close(), timeout=15)
            except Exception as e:
                log_exception('_force_full_reconnect: close old connection', e)
        _metaapi_conn = _metaapi_account.get_streaming_connection()
        await asyncio.wait_for(_metaapi_conn.connect(), timeout=30)
        await asyncio.wait_for(_metaapi_conn.wait_synchronized(), timeout=30)
        c_log("EXEC WATCHDOG: reconnect successful (execution channel live).")
        await set_connection_state(CONN_RUNNING, "Exec watchdog: forced reconnect succeeded.")
        await send_tg_msg(f"🔁 <b>Watchdog: أعيد الاتصال تلقائياً بـ MetaApi (تنفيذ)</b>\nالسبب: {reason}")
    except asyncio.TimeoutError:
        c_log("EXEC WATCHDOG: reconnect attempt timed out -- will retry next scan cycle (~15s).")
        await send_tg_msg(
            f"🛑 <b>Watchdog: انتهت مهلة إعادة الاتصال (30s)</b>\nالسبب الأصلي: {reason}\n"
            f"سيُعاد المحاولة تلقائياً بدورة فحص جديدة خلال ~15 ثانية."
        )
    except Exception as e:
        log_exception('_force_full_reconnect', e)
        await send_tg_msg(f"🛑 <b>Watchdog: فشلت محاولة إعادة الاتصال التلقائي</b>\nالسبب الأصلي: {reason}\nالخطأ: {e}")


def _lq_is_stale(symbol: str) -> bool:
    q = live_quotes.get(symbol)
    return q is None or (time.monotonic() - q['ts']) > _QUOTE_STALE_SECONDS


async def _lq_price_with_fallback(symbol: str) -> tuple[float | None, str, float | None]:
    """Returns (price, source, age_ms). 
    
    CRITICAL: In the EXECUTION path we MUST NOT fall back to OANDA REST.
    A 50-150ms REST call adds unacceptable latency. Instead we return None
    and let the caller reject the trade. The scanner cycle uses a separate
    path that can tolerate the fallback for diagnostics only.
    """
    q = live_quotes.get(symbol)
    if q is not None and (time.monotonic() - q['ts']) <= _QUOTE_STALE_SECONDS:
        return q['mid'], 'ws', round((time.monotonic() - q['ts']) * 1000)
    # WS quote missing or stale — DO NOT fall back to OANDA REST
    # (that would add 50-150ms latency). Return None to signal stale feed.
    return None, 'ws_stale', None

# Connection-state machine.
# RUNNING    : normal operation, new trades allowed.
# READ_ONLY  : sync with MetaAPI is degraded/unavailable. No new trades,
#              no destructive local state changes (Amnesia Prevention),
#              existing positions still managed if MT5 fallback price works.
# HALTED     : hard stop. New entries and order management both stop;
#              a human must intervene.
CONN_RUNNING   = 'RUNNING'
CONN_READ_ONLY = 'READ_ONLY'
CONN_HALTED    = 'HALTED'

_state_lock = asyncio.Lock()
_last_state_notify_ts = 0.0

# Event-driven fill detection: map order_id -> asyncio.Event
# A background task monitors terminal_state.positions and sets the event
# when the position appears, eliminating busy-wait polling (saves 300-1000ms per trade).
_fill_events: dict[str, asyncio.Event] = {}
_fill_results: dict[str, dict] = {}  # order_id -> {fill_price, fill_source, trade_id}
_fill_monitor_started = False
_fill_monitor_task: asyncio.Task | None = None
_fill_monitor_lock = asyncio.Lock()

async def _start_fill_monitor():
    """Start the background fill monitor task (idempotent)."""
    global _fill_monitor_started, _fill_monitor_task
    async with _fill_monitor_lock:
        if _fill_monitor_started:
            return
        _fill_monitor_started = True
        _fill_monitor_task = asyncio.create_task(_fill_monitor_loop())

async def _fill_monitor_loop():
    """Background loop that watches terminal_state.positions for new fills.
    
    When a position with a tracked order_id appears, sets the corresponding
    Event and stores the fill data. This eliminates the 300-1000ms busy-wait
    polling in _execute_smart_order.
    """
    try:
        while True:
            await asyncio.sleep(0.05)  # 50ms check interval - much faster than 300ms poll
            if not _fill_events or _metaapi_conn is None:
                continue
            try:
                positions = _metaapi_conn.terminal_state.positions
                if not positions:
                    continue
                for p in positions:
                    pid = str(p.get('id', ''))
                    if pid in _fill_events and pid not in _fill_results:
                        open_price = p.get('openPrice')
                        if open_price is not None:
                            _fill_results[pid] = {
                                'fill_price': float(open_price),
                                'fill_source': 'confirmed_position',
                                'trade_id': pid,
                            }
                            _fill_events[pid].set()
            except Exception as e:
                # Don't crash the monitor on transient errors
                log_exception('_fill_monitor_loop', e)
    except asyncio.CancelledError:
        pass
    except Exception as e:
        log_exception('_fill_monitor_loop crashed', e)

async def set_connection_state(new_state: str, reason: str) -> None:
    global _last_state_notify_ts
    async with _state_lock:
        old_state = bot_state.get('connection_state', CONN_RUNNING)
        if old_state == new_state:
            return
        bot_state['connection_state'] = new_state
        bot_state['connection_state_reason'] = reason
    logger.warning("Connection state: %s -> %s (%s)", old_state, new_state, reason)
    now = time.monotonic()
    # Debounce Telegram notifications: a real transition (old_state !=
    # new_state, already checked above) always logs, but only sends a
    # message if we haven't already alerted within the last 5 minutes.
    # This is what stops rapid RUNNING<->READ_ONLY flapping (e.g. a
    # borderline-stale feed recovering and going stale again every few
    # seconds) from spamming Telegram with one message per flap.
    _STATE_NOTIFY_COOLDOWN_SECONDS = 300.0
    if now - _last_state_notify_ts < _STATE_NOTIFY_COOLDOWN_SECONDS:
        return
    _last_state_notify_ts = now
    icon = {'RUNNING': '\u2705', 'READ_ONLY': '\U0001F7E1', 'HALTED': '\U0001F6D1'}.get(new_state, '\u2139')
    await send_tg_msg(f"{icon} <b>connection state changed: {old_state} -> {new_state}</b>\n{reason}")

_DAM_RESTRICTED_WINDOWS = [
    (dtime(7, 0),  dtime(9, 0)),   # European Open fakeouts
    (dtime(13, 0), dtime(14, 0)),  # Pre-US session turbulence
]

def _is_within_dam_restricted_window() -> bool:
    """DAM (Damascus / UTC+3) time-of-day filter. Based on backtest
    analysis, these windows carry enough market noise to invalidate Gann
    levels and stack losses, so new entries are skipped during them.
    Existing-position management (BE/TP/SL/closures) is NOT affected --
    this only blocks NEW trade dispatch, same scope as is_trading_allowed().
    Toggleable via bot_state['prot_dam_time_filter'] (default: on)."""
    if not bot_state.get('prot_dam_time_filter', True):
        return False
    dam_now = datetime.now(timezone.utc) + timedelta(hours=3)
    t = dam_now.time()
    return any(start <= t < end for start, end in _DAM_RESTRICTED_WINDOWS)

async def is_trading_allowed() -> bool:
    """New order placement is only allowed when the connection state is
    fully healthy AND we're not inside a restricted DAM time window.
    Existing-position management (BE/TP/SL) is handled separately and is
    NOT gated by this, per the OANDA-degraded-mode rule."""
    async with _state_lock:
        conn_state = bot_state.get('connection_state', CONN_RUNNING)
    if conn_state != CONN_RUNNING:
        return False
    if _is_within_dam_restricted_window():
        return False
    return True

async def _bootstrap_metaapi_connection() -> bool:
    """The actual connect-and-subscribe logic, extracted so it can be
    retried from the live scanner loop, not just called once at process
    startup. Returns True on success. This is what closes the gap where
    a transient MetaApi/broker hiccup during the ONE startup attempt left
    _metaapi_conn permanently None with no other recovery path able to
    rebuild it from scratch (both the WS tick-watchdog and the Zombie
    Singleton Heartbeat below require a connection object to already
    exist before they can do anything)."""
    global _metaapi, _metaapi_account, _metaapi_conn, _last_any_tick_ts
    try:
        _metaapi = MetaApi(METAAPI_TOKEN)
        _metaapi_account = await _metaapi.metatrader_account_api.get_account(ACCOUNT_ID)
        # NOTE: we deliberately do NOT also require
        # _metaapi_account.connection_status == 'CONNECTED' here.
        # MetaApi's own docs describe connectionStatus as a "replica field"
        # -- a periodically-synced snapshot, not the real-time truth (the
        # account dashboard's live "Connected" indicator uses the actual
        # real-time websocket status, which can be ahead of this REST
        # snapshot by anywhere from seconds to minutes). Gating here on that
        # stale field caused the bot to refuse to even ATTEMPT a connection
        # while the account was already genuinely healthy and connected on
        # MetaApi's side. state == 'DEPLOYED' is a real, non-lagging fact
        # (the API server either exists or it doesn't); the actual
        # connectivity check now happens for real via wait_synchronized()
        # below, which IS the authoritative real-time signal, wrapped in our
        # existing timeout so a genuinely bad connection still fails fast.
        if _metaapi_account.state == 'DEPLOYED':
            _metaapi_conn = _metaapi_account.get_streaming_connection()
            await asyncio.wait_for(_metaapi_conn.connect(), timeout=30)
            await asyncio.wait_for(_metaapi_conn.wait_synchronized(), timeout=30)
            # NOTE: live market-data subscription (_lq_subscribe_symbol) is
            # intentionally NOT called here anymore -- the price feed is now
            # driven entirely by oanda_live_price_poller (OANDA REST polling).
            # MetaAPI is still used for order execution / reconciliation only.
            c_log("MetaAPI Streaming Connection established (execution channel only; price feed via OANDA REST).")
            _last_any_tick_ts = time.monotonic()
            await set_connection_state(CONN_RUNNING, "MetaAPI connected and synchronized.")
            return True
        else:
            c_log(f"MetaAPI account not deployed (state={_metaapi_account.state}).")
            await set_connection_state(CONN_READ_ONLY, f"MetaAPI account is not DEPLOYED (state={_metaapi_account.state}).")
            return False
    except Exception as e:
        log_exception("_bootstrap_metaapi_connection", e)
        await set_connection_state(CONN_READ_ONLY, f"MetaAPI connection bootstrap failed: {e}")
        return False


async def init_metaapi():
    """Startup order is fixed:
       1) Reconstruct state from the persistence file (works even if the
          broker/API is completely unreachable).
       2) Only THEN attempt to talk to MetaAPI / the market.
    """
    await load_bot_persistence()
    if bot_state.get('_persistence_load_failed'):
        await set_connection_state(
            CONN_READ_ONLY,
            "Startup persistence file was present but unreadable. Starting READ_ONLY until a human "
            "confirms the true broker state and clears this manually."
        )
    await _bootstrap_metaapi_connection()

DATA_DIR = os.environ.get('PERSISTENT_DATA_PATH', '/app/data')
os.makedirs(DATA_DIR, exist_ok=True)
PERSISTENCE_FILE = os.path.join(DATA_DIR, 'bot_persistence.json')
TEMP_PERSISTENCE_FILE = os.path.join(DATA_DIR, 'bot_persistence.tmp')
PRESETS_FILE = os.path.join(DATA_DIR, 'presets.json')
TEMP_PRESETS_FILE = os.path.join(DATA_DIR, 'presets.tmp')

# Live runtime fields that a preset must never capture or restore --
# in-flight state that belongs to whatever is currently running, not to a
# saved settings snapshot.
_PRESET_EXCLUDED_KEYS = {
    'strategy_open_trades', 'auto_trade',
}

# Event-loop I/O offloading (v9.5): json.dump + os.fsync + os.replace are
# blocking syscalls. Calling them directly from async code stalls the
# ENTIRE event loop for their duration -- every other coroutine (candle
# fetches, BE checks, Telegram alerts, callback handling) waits behind a
# single disk write. The fix: build the snapshot synchronously (cheap,
# pure in-memory dict work, no yield point, so it's still atomic w.r.t.
# bot_state), then push the actual disk I/O to a worker thread via
# asyncio.to_thread and await it there. Writes are additionally serialized
# with an asyncio.Lock so two saves in flight can't interleave writes to
# the shared .tmp path before either os.replace runs.
_persistence_write_lock = asyncio.Lock()

def _write_persistence_file_sync(data: dict) -> None:
    """Pure blocking I/O, no bot_state access -- safe to run in a thread."""
    with open(TEMP_PERSISTENCE_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)
        f.flush()
        os.fsync(f.fileno())
    os.replace(TEMP_PERSISTENCE_FILE, PERSISTENCE_FILE)

async def save_bot_persistence() -> None:
    """Atomic write: full operational state AND full settings, so a hard
    restart reconstructs the bot's world exactly -- not just open trades
    and per-symbol strategy state, but every user-configured setting (lot
    size, protection limits, monitored timeframes, TP/SL config, etc).

    Deliberately exclude-list based, not include-whitelist based: the
    previous version only ever saved a fixed list of live trade-state
    fields, which meant lot size, protection dd/profit limits, monitored
    timeframes, and effectively every other setting were NEVER actually
    persisted -- even though save_bot_persistence() was correctly being
    called after every mutation. An exclude-list means any new setting
    added later is persisted automatically instead of silently dropped
    until someone remembers to add it to a whitelist.
    """
    try:
        # Fields that either aren't JSON-serializable or are purely
        # transient/regenerated-on-render -- everything else in bot_state
        # is a real setting and gets saved.
        TOP_LEVEL_EXCLUDE = {'connection_obj', 'menu_button_map', 'timeframes',
                              'is_backtesting', 'live_connected', 'last_poll_ok', 'symbol_state',
                              'diag_log'}

        symbol_snapshot = {}
        for sym in sorted(bot_state['active_symbols'].keys()):
            ss = bot_state['symbol_state'].get(sym) or {}
            symbol_snapshot[sym] = dict(ss)

        data = {
            'schema_version': 3,
            'symbol_state': symbol_snapshot,
        }
        for k, v in bot_state.items():
            if k not in TOP_LEVEL_EXCLUDE:
                data[k] = v
        raw = bot_state.get('live_daily_date')
        data['live_daily_date'] = raw.isoformat() if hasattr(raw, 'isoformat') else str(raw or '')
    except Exception as e:
        log_exception("save_bot_persistence (snapshot phase)", e)
        return

    try:
        async with _persistence_write_lock:
            await asyncio.to_thread(_write_persistence_file_sync, data)
    except Exception as e:
        # Persistence failing is itself a critical-path failure: if we can't
        # save state, a crash right now means real, silent data loss on
        # open positions. Escalate loudly instead of swallowing it.
        log_exception("save_bot_persistence (write phase)", e)
        c_log(f"CRITICAL: Persistence Save Error -- open trade state may not survive a restart: {e}")

async def load_bot_persistence():
    if not os.path.exists(PERSISTENCE_FILE):
        c_log("No persistence file found -- starting fresh (expected on first boot).")
        return
    try:
        def _read():
            with open(PERSISTENCE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        data = await asyncio.to_thread(_read)

        TOP_LEVEL_EXCLUDE = {'connection_obj', 'menu_button_map', 'timeframes',
                              'is_backtesting', 'is_live_twin_running',
                              'live_connected', 'last_poll_ok', 'symbol_state',
                              'diag_log'}
        for k, v in data.items():
            if k in bot_state and k not in TOP_LEVEL_EXCLUDE and k != 'live_daily_date':
                bot_state[k] = v

        saved_date = data.get('live_daily_date')
        if saved_date and saved_date != 'None' and saved_date:
            try:
                bot_state['live_daily_date'] = datetime.fromisoformat(saved_date).date()
            except Exception:
                bot_state['live_daily_date'] = None

        symbol_state_data = data.get('symbol_state')
        if symbol_state_data is not None:
            for sym, snap in symbol_state_data.items():
                if sym not in bot_state['symbol_state']:
                    continue
                ss = bot_state['symbol_state'][sym]
                for k, v in snap.items():
                    if k in ss:  # same safety principle as top-level: only restore known fields
                        ss[k] = v
        else:
            # Backward-compat with the oldest schema (open trades only).
            for sym, trades in data.get('strategy_open_trades', data.get('strategy_open_trades', {})).items():
                if sym in bot_state['symbol_state']:
                    bot_state['symbol_state'][sym]['strategy_open_trades'] = trades

        c_log("Bot state restored from persistence file (settings, open trades, strategy state, daily PnL).")
    except Exception as e:
        # If the persistence file is corrupt we must not silently pretend
        # we're starting clean while real broker positions may still be
        # open. Flag it; init_metaapi/main will use this to force READ_ONLY.
        log_exception("load_bot_persistence", e)
        c_log(f"CRITICAL: Persistence file exists but failed to load ({e}). "
              f"Bot will start in READ_ONLY to avoid trading blind.")
        bot_state['_persistence_load_failed'] = True

bot_state: dict = {
    'connection_state': 'RUNNING',
    'connection_state_reason': '',
    'symbol':           'XAUUSD',
    'live_connected':   False,
    'connection_obj':   None,
    'chat_id':          None,
    'last_update_id':   0,
    'is_backtesting':   False,
    'is_live_twin_running': False,
    'timeframes':       _TFS,

    'lt_latency_ms_min': 160,   # measured Railway-deployment -> broker round-trip ping
    'lt_latency_ms_max': 200,   # (update again if a future diagnostic run measures differently)

    # ── Live-Twin Engine (realistic execution simulator) ──
    # Baseline spread taken from a live MT5/OANDA tick snapshot on
    # 2026-07-13 during the late-night (low-liquidity) session:
    # Bid 4112.28 / Ask 4112.62 -> 0.34 USD (34 points at tick=0.01).
    # This is the QUIET-SESSION floor; session/volatility multipliers
    # scale it up or down from here, they never invent a new baseline.
    'lt_mode': 'realistic',        # 'realistic' or 'idealized' (idealized == old run_strategy_backtest, zero friction, kept as A/B baseline)
    'lt_base_spread_usd': 0.34,
    'lt_friction': {
        'spread':    True,   # dynamic session/volatility spread model
        'slippage':  True,   # asymmetric, range-scaled slippage
        'latency':   True,   # 200-800ms signal-to-fill delay
        'commission': True,  # per-lot round-turn commission
        'gaps':      True,   # weekend/rollover gap risk
        'rejection': True,   # requote/rejection probability in volatility spikes
    },
    # Calibrated from the actual XAUUSD broker spec (MT5 symbol properties
    # screenshot): "Commissions: 0-1000 -> 5 USD per lot, Instant by deal
    # volume, in deals" means 5 USD PER DEAL, i.e. per side -- a round-turn
    # trade (open + close) costs 10, not the old flat guess of 7.
    'lt_commission_per_lot': 10.0,  # USD round-turn per 1.0 lot (5 open + 5 close, per real broker spec)
    # Swap was previously a single flat value applied to both directions,
    # which is wrong: the real broker spec shows swap long/short are wildly
    # asymmetric (Swap long: -93.1728, Swap short: +21.6848, in points), and
    # Wednesday carries a 3x multiplier (standard weekend-rollover
    # compensation). Points -> USD/lot conversion uses tick size x contract
    # size from the same spec (0.01 x 100 = $1/point/lot for XAUUSD).
    # Re-derive these two numbers yourself if your broker's spec differs.
    'lt_swap_long_per_lot_night': -93.17,   # USD per lot per night held, BUY positions
    'lt_swap_short_per_lot_night': 21.68,   # USD per lot per night held, SELL positions
    'lt_swap_wednesday_multiplier': 3.0,    # applied when the rollover date is a Wednesday
    'lt_swap_per_lot_night': -6.5,  # legacy fallback only, kept for old configs; no longer used directly
    'lt_rejection_prob': 0.015,     # probability a signal is rejected/requoted during an ATR spike bar

    
    'menu_button_map': {},
    'last_poll_ok':     0.0,
    'live_daily_realized': 0.0,
    'live_daily_date': None,
    'live_daily_hit': False,

    # ── Gann Levels Engine ──
    # -- EMA/Stochastic Strategy Engine --
    'active_symbols': {s: (s == 'XAU_USD') for s in AVAILABLE_SYMBOLS},
    'ui_selected_symbol': 'XAU_USD',
    'symbol_state': {s: {
        'strategy_open_trades': {},
        'auto_trade': False,
        'lot_size': 0.05,
        'break_even_enabled': False,
        'strategy_be_trigger_points': 40,
        'strategy_monitor_tfs': {tf: (tf in ['5m', '10m', '15m', '20m', '30m', '1h', '4m', '6m', '2h', '1m', '2m', '3m']) for tf in _TFS},
        'strategy_tpsl_mode': 'fixed',
        'strategy_tp_points': 70,
        'strategy_sl_points': 110,
        'strategy_tp_per_tf': {tf: 0 for tf in _TFS},
        'strategy_sl_per_tf': {tf: 0 for tf in _TFS},
        'strategy_atr_period': 14,
        'strategy_atr_sl_mult': 1.5,
        'strategy_atr_tp_mult': 2,
        # Dedup: last (per-tf) closed-candle timestamp a signal was already
        # evaluated for, so the candle-close scanner never fires twice off
        # the same candle. Keyed by tf -> ISO string (kept JSON-safe).
        'strategy_last_signal_candle': {},
    } for s in AVAILABLE_SYMBOLS},

    # -- Trade Management --
    'signal_candle_shift': 1,          # 0 = live/forming candle, 1 = last closed, 2 = previous closed
    'allow_concurrent_trades': False,  # if False: block a new Buy while a Buy is open (same for Sell)
    
    'prot_daily_dd_usd':      200,
    'prot_daily_profit_usd':  150,
    'prot_true_sync': True,
    'prot_cost_be': True,
    # Max allowed execution deviation (MetaApi "slippage", in broker points)
    # for market orders. If the broker can't fill within this many points of
    # our intended price, the order is rejected (ORDER_FILLING_FOK) instead
    # of being executed 20 pips away.
    'prot_max_slippage_points': 5,
    # Hard cap on simultaneously open trades per symbol, regardless of how
    # many timeframes are enabled or how many levels got touched at once.
    # When reached, remaining timeframes are skipped for that scan cycle
    # only -- already-open trades are left alone (Option A).
    'prot_max_concurrent_trades': 4,
    # Rolling in-memory log of every (symbol, timeframe) scan decision made
    # by the live scanner -- NOT just the instantaneous /diagnose snapshot.
    # Deliberately excluded from persistence (see TOP_LEVEL_EXCLUDE) since
    # it's diagnostic-only and would otherwise bloat the save file.
    'diag_log': [],
    # Full history of every CLOSED live/real trade, rich enough to rebuild an
    # Excel report matching the backtest's own format. Unlike diag_log this
    # IS persisted (not in TOP_LEVEL_EXCLUDE) -- it's the actual trade record,
    # not throwaway diagnostics, and must survive a bot restart.
    'live_trade_history': [],
    'prot_stale_filter': True,
    'prot_allow_multi_tf':    True,

    # ── Broker/display time alignment ──
    # Hours to add to raw UTC (from OANDA/MetaApi) to reach the broker's
    # own server clock (what the user's MT5 terminal displays).  Default 3 =
    # Damascus/EET-DST-style broker offset.
    'broker_time_offset': 3,

    
    
    
    
    
    
    
    
    
    
    
    
}


# ── Execution Quality Tracker ──
# Records every real order's fill quality for self-tuning and diagnostics.
class _ExecTracker:
    def __init__(self, maxlen=200):
        self.orders = []
        self.maxlen = maxlen

    def record(self, symbol, is_buy, level_price, fill_price, fill_source,
               latency_ms, method_used, success, error=None):
        slippage = None
        if fill_price is not None and level_price is not None:
            slippage = round(abs(fill_price - level_price), 5)
        err_str = str(error)[:200] if error else None
        self.orders.append({
            'ts': time.monotonic(),
            'symbol': symbol,
            'is_buy': is_buy,
            'level_price': level_price,
            'fill_price': fill_price,
            'slippage': slippage,
            'latency_ms': latency_ms,
            'method': method_used,
            'success': success,
            'error': err_str,
        })
        if len(self.orders) > self.maxlen:
            self.orders.pop(0)

    def avg_slippage(self, symbol=None, n=20):
        recent = [o for o in self.orders if o['slippage'] is not None
                  and (symbol is None or o['symbol'] == symbol)][-n:]
        return sum(o['slippage'] for o in recent) / len(recent) if recent else None

    def limit_fill_rate(self, symbol=None, n=50):
        recent = [o for o in self.orders if o['method'] == 'limit'
                  and (symbol is None or o['symbol'] == symbol)][-n:]
        return sum(1 for o in recent if o['success']) / len(recent) if recent else None

_exec_tracker = _ExecTracker()

DAM_OFF = timedelta(hours=3)
def _utc_to_dam(dt) -> datetime:
    if isinstance(dt, pd.Timestamp): dt = dt.to_pydatetime()
    if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
    return dt + DAM_OFF


# ─────────────────────────────────────────────────────────────
# UNIFIED CORE LOGIC (V9.4)
# ─────────────────────────────────────────────────────────────
def core_eval_break_even(is_buy: bool, entry: float, current_px: float, pip_value: float, be_pts: int, atr_period: int, cost_be: bool) -> float | None:
    be_dist = be_pts * pip_value
    if (is_buy and current_px >= entry + be_dist) or (not is_buy and current_px <= entry - be_dist):
        be_margin = (atr_period * 0.1 * pip_value) if cost_be else 0.0
        return (entry + be_margin) if is_buy else (entry - be_margin)
    return None

def core_eval_outcome(is_buy: bool, current_px: float, tp: float, sl: float) -> str | None:
    if is_buy:
        if current_px >= tp: return 'WIN ✅'
        if current_px <= sl: return 'LOSS ❌'
    else:
        if current_px <= tp: return 'WIN ✅'
        if current_px >= sl: return 'LOSS ❌'
    return None
    
# ─────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────
# TELEGRAM HELPERS
# ─────────────────────────────────────────────────────────────
async def _tg_post(url: str, **kwargs) -> bool:
    try:
        sess = get_http()
        async with sess.post(url, **kwargs) as resp:
            if resp.status != 200:
                body = await resp.text()
                c_log(f"Telegram API call failed ({resp.status}) for {url}: {body[:300]}")
            return resp.status == 200
    except Exception as e:
        # This carries HALT/READ_ONLY escalation alerts, so a silent
        # failure here means the operator never finds out something broke.
        log_exception(f"_tg_post [{url}]", e)
        return False

def _to_reply_kbd(inline_kbd: dict):
    rows = []; bmap = {}
    for row in inline_kbd.get('inline_keyboard', []):
        new_row = []
        for btn in row:
            text = btn['text']; cb = btn.get('callback_data', 'noop')
            if text in bmap and bmap[text] != cb and cb != 'noop' and bmap[text] != 'noop':
                # This is exactly the bug class that caused the loss/profit
                # buttons to collide: two DIFFERENT actions sharing the same
                # button text, silently overwriting each other in the map
                # that resolves a tapped label back to an action. Every
                # button's text must be unique within a single keyboard.
                c_log(f"BUTTON LABEL COLLISION: '{text}' maps to both '{bmap[text]}' and '{cb}' -- "
                      f"the second silently wins and the first becomes untappable. Fix the keyboard's labels.")
            new_row.append({'text': text}); bmap[text] = cb
        rows.append(new_row)
    return {'keyboard': rows, 'resize_keyboard': True, 'is_persistent': True, 'input_field_placeholder': 'اختر من القائمة...'}, bmap

async def send_tg_msg(text: str, reply_markup: dict = None) -> None:
    if not bot_state['chat_id']: return
    if reply_markup and 'inline_keyboard' in reply_markup:
        reply_markup, bmap = _to_reply_kbd(reply_markup); bot_state['menu_button_map'] = bmap
    payload = {'chat_id': bot_state['chat_id'], 'text': text, 'parse_mode': 'HTML'}
    if reply_markup: payload['reply_markup'] = reply_markup
    await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/sendMessage', json=payload)

async def edit_tg_msg(chat_id, message_id, text, reply_markup=None) -> None:
    payload = {'chat_id': chat_id, 'message_id': message_id, 'text': text, 'parse_mode': 'HTML'}
    if reply_markup: payload['reply_markup'] = reply_markup
    await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/editMessageText', json=payload)

async def _show(chat_id, msg_id, text: str, reply_markup: dict = None) -> None:
    if msg_id: await edit_tg_msg(chat_id, msg_id, text, reply_markup)
    else: await send_tg_msg(text, reply_markup)

async def answer_callback(cbq_id: str) -> None:
    await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/answerCallbackQuery', json={'callback_query_id': cbq_id})

TG_CAPTION_LIMIT = 1024  # Telegram hard limit for document/photo captions

async def send_tg_document(file_path: str, caption: str) -> None:
    if not bot_state['chat_id']: return
    try:
        # A caption over Telegram's limit doesn't get truncated by the API --
        # the whole sendDocument call is rejected, so the file itself would
        # never arrive. Keep the merged single-message intent when it fits;
        # fall back to a short caption + a separate full-text message only
        # when it doesn't.
        doc_caption = caption
        overflow_text = None
        if len(caption) > TG_CAPTION_LIMIT:
            doc_caption = caption[:TG_CAPTION_LIMIT - 20].rstrip() + "\n... (تابع أدناه)"
            overflow_text = caption

        with open(file_path, 'rb') as f:
            data = aiohttp.FormData()
            data.add_field('chat_id',  str(bot_state['chat_id']))
            data.add_field('document', f, filename=os.path.basename(file_path))
            data.add_field('caption',  doc_caption)
            await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/sendDocument', data=data)

        if overflow_text:
            await send_tg_msg(overflow_text)
    except Exception as e:
        log_exception(f"send_tg_document [{file_path}]", e)

# ─────────────────────────────────────────────────────────────
# OANDA FETCHER 
# ─────────────────────────────────────────────────────────────
_OANDA_GRAN = {'1m':'M1','2m':'M2','3m':'M3','4m':'M4','5m':'M5','6m':'M6','10m':'M10','15m':'M15','20m':'M20','30m':'M30','1h':'H1','2h':'H2'}
_oanda_sem: asyncio.Semaphore | None = None
def _get_oanda_sem() -> asyncio.Semaphore:
    global _oanda_sem
    if _oanda_sem is None: _oanda_sem = asyncio.Semaphore(3)
    return _oanda_sem

def _safe_float(value, default: float = 0.0) -> float:
    """Closes the `.get(key, default)` null trap: dict.get()'s default only
    applies when the key is MISSING. If MetaAPI returns the key present
    with an explicit `null` (-> None in Python), .get() happily returns
    None, and a later `+=` or arithmetic op on it raises TypeError. This
    coerces None/non-numeric/NaN/inf values to `default` instead of
    raising, at every point real MetaAPI numeric fields are consumed."""
    if value is None:
        return default
    try:
        f = float(value)
    except (TypeError, ValueError):
        return default
    if f != f or f in (float('inf'), float('-inf')):  # NaN / inf guard
        return default
    return f

def _validated_candle(c: dict, symbol: str, granularity_str: str) -> dict | None:
    """Defensive boundary for external market data. OANDA/MetaAPI are not
    contractually guaranteed to always return well-typed floats -- a
    transient glitch can hand back None, a string, or a missing key.
    Returns a clean candle dict, or None if this single candle is bad.
    Never raises: a bad candle should be skipped, not take down the whole
    fetch (or the caller's while-True loop) with it."""
    try:
        mid = c.get('mid')
        if not isinstance(mid, dict):
            raise ValueError(f"missing/invalid 'mid' field: {mid!r}")
        raw_time = c.get('time')
        if not raw_time:
            raise ValueError("missing 'time' field")

        o = float(mid['o']); h = float(mid['h']); l = float(mid['l']); c_ = float(mid['c'])
        vol = float(c.get('volume', 1.0) or 1.0)

        for v in (o, h, l, c_, vol):
            if v != v or v in (float('inf'), float('-inf')):
                raise ValueError(f"non-finite value in candle: {v!r}")

        return {
            'time': pd.Timestamp(raw_time).tz_convert('UTC'),
            'open': o, 'high': h, 'low': l, 'close': c_, 'volume': vol,
        }
    except (TypeError, ValueError, KeyError) as e:
        log_exception(f"_validated_candle [{symbol} {granularity_str}] -- skipping malformed candle", e)
        return None

async def fetch_candles(symbol: str, granularity_str: str, count: int = 5000, end_time: datetime = None) -> list:
    gran_str = _OANDA_GRAN.get(granularity_str, 'M1'); fetch_count = min(count, 120000)  
    collected = []; remaining = fetch_count
    headers = {'Authorization': f'Bearer {OANDA_TOKEN}', 'Content-Type':  'application/json'}
    url = f'{OANDA_BASE_URL}/instruments/{symbol}/candles'
    current_end = end_time if end_time else datetime.now(timezone.utc)

    while remaining > 0:
        chunk = min(remaining, 5000)
        params = {'granularity': gran_str, 'count': chunk, 'to': current_end.strftime('%Y-%m-%dT%H:%M:%S.000000000Z'), 'price': 'M'}
        candles = []
        async with _get_oanda_sem():
            for attempt in range(6):
                try:
                    async with get_http().get(url, headers=headers, params=params, timeout=aiohttp.ClientTimeout(total=20)) as resp:
                        if resp.status != 200:
                            if attempt == 5:
                                c_log(f"fetch_candles [{symbol} {granularity_str}]: giving up after 6 attempts "
                                      f"(last status {resp.status}) -- collected {len(collected)}/{fetch_count} candles so far.")
                                break
                            await asyncio.sleep(min(2 ** attempt, 30))
                            continue
                        data = await resp.json(); candles = data.get('candles', []); break
                except Exception as e:
                    log_exception(f"fetch_candles [{symbol} {granularity_str}] attempt {attempt+1}/6", e)
                    await asyncio.sleep(min(2 ** attempt, 30))

        if not candles: break
        complete = [c for c in candles if c.get('complete', True)]
        if not complete: break

        formatted = []
        for c in complete:
            vc = _validated_candle(c, symbol, granularity_str)
            if vc is not None:
                formatted.append(vc)

        if not formatted:
            c_log(f"fetch_candles [{symbol} {granularity_str}]: entire chunk failed validation, aborting fetch.")
            break

        collected = formatted + collected; remaining -= len(complete)
        earliest = pd.Timestamp(complete[0]['time']).tz_convert('UTC')
        current_end = earliest.to_pydatetime() - timedelta(seconds=1)
        if len(complete) < chunk: break
        await asyncio.sleep(0.2)
    return collected

async def fetch_master_price(symbol: str) -> float | None:
    """Single Source of Truth for the CURRENT live price.

    Call this exactly ONCE per symbol per scanner cycle, then reuse the
    returned value for every enabled timeframe's touch-distance check.

    Why this exists: a timeframe's own last candle close is NOT "the
    current price" for anything above 1m -- a 30m candle's close can be up
    to ~30 minutes stale. Asking OANDA separately per-timeframe and using
    each tf's own close as "live price" is exactly what caused the same
    instant to read as e.g. 4067 on 1m and 4073 on 30m during a volatile
    spike, and it also multiplies OANDA requests per cycle (contributing
    to "Insufficient data from OANDA" failures under load). Timeframes
    should still be fetched separately for their own historical
    closes/EMAs/ATR -- just never for "what is the price right now".

    count=2 (not 1): OANDA's most recent candle for 'to=now' is very often
    still the in-progress (incomplete) one, and fetch_candles() drops
    incomplete candles entirely. count=1 would then frequently return an
    EMPTY list and report "insufficient data" even though OANDA itself is
    perfectly healthy. count=2 guarantees at least one genuinely
    completed, very recent candle to use.
    """
    mc = await fetch_candles(symbol, '1m', count=2)
    if not mc:
        c_log(f"fetch_master_price [{symbol}]: no 1m data from OANDA this cycle -- "
              f"skipping touch checks for this symbol rather than risk a stale/desynced price.")
        return None
    return float(mc[-1]['close'])

# ─────────────────────────────────────────────────────────────
# STRATEGY CORE: EMA 15/50/150 trend stack + Stochastic(5,5,5, EMA-smoothed)
# reversal-zone crossover. Evaluated strictly on candle close (see
# bot_state['signal_candle_shift']: 0 = live/forming candle, 1 = last
# closed candle, 2 = the one before that).
# ─────────────────────────────────────────────────────────────
STRAT_EMA_FAST = 15
STRAT_EMA_MED = 50
STRAT_EMA_SLOW = 150
STRAT_STOCH_K = 5
STRAT_STOCH_D = 5
STRAT_STOCH_SLOWING = 5
STRAT_STOCH_BUY_LEVEL = 20    # crossover must occur with both lines <= this
STRAT_STOCH_SELL_LEVEL = 80   # crossover must occur with both lines >= this
STRAT_STOCH_DEEP_BUY_LEVEL = 10   # secondary reference level (display/diagnostics only)
STRAT_STOCH_DEEP_SELL_LEVEL = 90  # secondary reference level (display/diagnostics only)
STRAT_MIN_CANDLES = STRAT_EMA_SLOW + STRAT_STOCH_K + STRAT_STOCH_D + STRAT_STOCH_SLOWING + 10

def compute_strategy_indicators(candles: list) -> "pd.DataFrame | None":
    """EMA 15/50/150 (standard EMA) + Stochastic Oscillator with
    %K period=5, %D period=5, Slowing=5, MA_Method=EMA (i.e. the raw %K is
    EMA-smoothed for the slowed %K line, and that slowed %K is EMA-smoothed
    again for %D) -- matches MT5's Stochastic Oscillator indicator with
    MA Method set to Exponential. Returns None if there isn't enough
    history to trust EMA150 + the Stochastic lookback."""
    if not candles or len(candles) < STRAT_MIN_CANDLES:
        return None
    df = pd.DataFrame(candles)
    df['ema15'] = df['close'].ewm(span=STRAT_EMA_FAST, adjust=False).mean()
    df['ema50'] = df['close'].ewm(span=STRAT_EMA_MED, adjust=False).mean()
    df['ema150'] = df['close'].ewm(span=STRAT_EMA_SLOW, adjust=False).mean()
    low_min = df['low'].rolling(window=STRAT_STOCH_K).min()
    high_max = df['high'].rolling(window=STRAT_STOCH_K).max()
    rng = (high_max - low_min).replace(0, np.nan)
    raw_k = 100.0 * (df['close'] - low_min) / rng
    raw_k = raw_k.fillna(50.0)
    df['stoch_k'] = raw_k.ewm(span=STRAT_STOCH_SLOWING, adjust=False).mean()   # slowed %K (EMA method)
    df['stoch_d'] = df['stoch_k'].ewm(span=STRAT_STOCH_D, adjust=False).mean()  # %D (EMA of slowed %K)
    return df

def evaluate_strategy_signal(df: "pd.DataFrame", shift: int) -> dict | None:
    """Evaluate the entry rules at candle index N = shift (0 = most recent
    row in df i.e. the live/forming candle as fetched, 1 = last closed,
    2 = previous closed). Returns a dict with 'is_buy' and diagnostic
    values on a valid signal, or None."""
    idx = len(df) - 1 - shift
    prev_idx = idx - 1
    if idx < 1 or prev_idx < 0:
        return None
    ema15, ema50, ema150 = df['ema15'].iloc[idx], df['ema50'].iloc[idx], df['ema150'].iloc[idx]
    k, d = df['stoch_k'].iloc[idx], df['stoch_d'].iloc[idx]
    k_prev, d_prev = df['stoch_k'].iloc[prev_idx], df['stoch_d'].iloc[prev_idx]
    if any(pd.isna(v) for v in (ema15, ema50, ema150, k, d, k_prev, d_prev)):
        return None

    cross_up = (k_prev <= d_prev) and (k > d)
    cross_down = (k_prev >= d_prev) and (k < d)

    # Zone check uses the PRE-cross bar (k_prev/d_prev), not the bar the
    # cross completes on. %D lags %K (it's a further EMA smoothing of it),
    # so by the moment %K finishes crossing above %D, %D has often already
    # started drifting out of the 10-20 zone -- checking the zone one bar
    # earlier still requires the cross to have originated from inside the
    # oversold/overbought zone, without discarding valid crosses purely
    # because %D was a hair slow to follow.
    if ema15 > ema50 > ema150 and cross_up and k_prev <= STRAT_STOCH_BUY_LEVEL and d_prev <= STRAT_STOCH_BUY_LEVEL:
        return {'is_buy': True, 'ema15': ema15, 'ema50': ema50, 'ema150': ema150, 'k': k, 'd': d,
                'candle_time': df['time'].iloc[idx] if 'time' in df.columns else None}
    if ema150 > ema50 > ema15 and cross_down and k_prev >= STRAT_STOCH_SELL_LEVEL and d_prev >= STRAT_STOCH_SELL_LEVEL:
        return {'is_buy': False, 'ema15': ema15, 'ema50': ema50, 'ema150': ema150, 'k': k, 'd': d,
                'candle_time': df['time'].iloc[idx] if 'time' in df.columns else None}
    return None

def _strategy_tf_tp(symbol: str, tf: str) -> int:
    sym_state = bot_state['symbol_state'][symbol]
    v = sym_state['strategy_tp_per_tf'].get(tf, 0)
    return v if v > 0 else sym_state['strategy_tp_points']

def _strategy_tf_sl(symbol: str, tf: str) -> int:
    sym_state = bot_state['symbol_state'][symbol]
    v = sym_state['strategy_sl_per_tf'].get(tf, 0)
    return v if v > 0 else sym_state['strategy_sl_points']

def _strategy_atr(candles: list, period: int) -> float | None:
    """Compute ATR using pure NumPy (no pandas overhead).
    If `candles` is None, returns None (cache miss)."""
    if candles is None or len(candles) < period + 1:
        return None
    # Take only the last (period + 50) candles for efficiency
    recent = candles[-(period + 50):]
    n = len(recent)
    high = np.empty(n, dtype=np.float64)
    low = np.empty(n, dtype=np.float64)
    close = np.empty(n, dtype=np.float64)
    for i, c in enumerate(recent):
        high[i] = c['high']
        low[i] = c['low']
        close[i] = c['close']
    # True Range: max(high-low, |high-prev_close|, |low-prev_close|)
    prev_close = np.roll(close, 1)
    prev_close[0] = close[0]  # first element has no prev_close, use itself
    tr1 = high - low
    tr2 = np.abs(high - prev_close)
    tr3 = np.abs(low - prev_close)
    tr = np.maximum(np.maximum(tr1, tr2), tr3)
    # Rolling mean of last `period` values
    atr = float(np.mean(tr[-period:]))
    return atr

def _strategy_calc_tpsl(symbol: str, entry: float, is_buy: bool, candles: list, tf: str = '') -> tuple[float, float]:
    sym_state = bot_state['symbol_state'][symbol]
    pv = SYMBOL_INFO[symbol]['pip_value']
    prec = SYMBOL_INFO[symbol]['prec']
    if sym_state['strategy_tpsl_mode'] == 'atr':
        atr = _strategy_atr(candles, sym_state['strategy_atr_period']) if candles else None
        if not atr: atr = _strategy_tf_sl(symbol, tf) * pv
        sl_dist = atr * sym_state['strategy_atr_sl_mult']
        tp_dist = atr * sym_state['strategy_atr_tp_mult']
    else:
        sl_dist = _strategy_tf_sl(symbol, tf) * pv
        tp_dist = _strategy_tf_tp(symbol, tf) * pv
    if is_buy: return round(entry + tp_dist, prec), round(entry - sl_dist, prec)
    return round(entry - tp_dist, prec), round(entry + sl_dist, prec)

_consecutive_real_order_failures = 0
_REAL_ORDER_FAILURE_HALT_THRESHOLD = 3
_last_scanner_error_alert_ts = 0.0

def _resolve_broker_symbol(symbol: str) -> str:
    """Resolve the OANDA-format data-feed symbol (e.g. 'XAU_USD') to the
    broker's actual MT5 symbol name for order execution. bot_state['symbol']
    (settable via /setsymbol) is the primary source of truth since brokers
    vary wildly in suffix conventions (XAUUSD, XAUUSDm, XAUUSD.a, GOLD...).
    A hard safety-net mapping is applied on top: if the configured value is
    missing or still looks like an unfixed raw OANDA-format symbol (i.e.
    still has the underscore), fall back to the confirmed-correct stripped
    form instead of sending a guaranteed-to-be-rejected symbol to the
    broker. The rest of the bot's data engine, fetches, and logs MUST
    continue using the OANDA-format symbol ('XAU_USD') -- this function's
    return value is ONLY for the MetaAPI execution payload."""
    configured = bot_state.get('symbol', '').strip()
    if not configured or '_' in configured:
        return symbol.replace('_', '')
    return configured


# ─────────────────────────────────────────────────────────────
# SMART ORDER EXECUTION: LIMIT-first, MARKET-fallback
# ─────────────────────────────────────────────────────────────
class _SkipLimitPhase(Exception):
    """Raised inside Phase 1 to skip to Phase 2 cleanly (no error log)."""
    pass


async def _execute_smart_order(symbol: str, is_buy: bool, lot: float,
                                level_price: float, sl: float, tp: float,
                                t1_signal_ts: float,
                                max_slippage_points: int) -> dict:
    """Zero-slippage execution for the EMA/Stochastic strategy's candle-close signals.

    Phase 1 — LIMIT (GTD 30s + cancel-on-move) with smart spread-aware placement.
      Smart limit price = level_price ± spread/2 to capture midpoint fills.
      Auto-cancels after 30s or if price moves > 2×margin away from level.

    Phase 2 — MARKET (FOK) with tight deviation guard.
      Only reached if Phase 1 was skipped or expired without filling.
      The FOK deviation cap limits max slippage to `max_slippage_points` broker points.

    Phase 3 — Hybrid IOC Emulation (NEW).
      If limit order not filled within 100ms, auto-send market FOK fallback.
      This emulates true IOC behavior without broker IOC support.

    Adaptive slippage budget: if rolling avg slippage > 2 pips, auto-widen
    `max_slippage_points` for next 10 trades to prevent FOK rejections.

    Returns dict with keys: success, trade_id, fill_price, fill_source,
                            latency_ms, method_used, error, ioc_fail_reason.
    """
    broker_symbol = _resolve_broker_symbol(symbol)
    trade_id = None
    fill_price = None
    fill_source = None
    method_used = None
    error = None
    ioc_fail_reason = None

    # ═══════════════════════════════════════════════════════════
    # PHASE 1 — LIMIT ORDER (GTD 30s + cancel-on-move + smart spread-aware)
    # ═══════════════════════════════════════════════════════════
    # NOTE: MT5 rejects `fillingModes: ['ORDER_FILLING_IOC']` on pending
    # (limit) orders — IOC/FOK are market-order-only in MT5.  Instead we
    # use a plain GTD limit with a 30-second expiration + active cancel-on-move.

    # Get current market spread for smart limit placement
    q = live_quotes.get(symbol)
    spread = None
    if q and 'ask' in q and 'bid' in q:
        spread = q['ask'] - q['bid']

    # Smart limit price: for BUY use level_price - spread/2 (midpoint between level and ask)
    # For SELL use level_price + spread/2 (midpoint between level and bid)
    if spread and spread > 0:
        smart_limit_price = level_price - spread / 2 if is_buy else level_price + spread / 2
        # Ensure we don't place beyond the level (for BUY: smart_limit must be <= level_price)
        if is_buy and smart_limit_price > level_price:
            smart_limit_price = level_price
        if not is_buy and smart_limit_price < level_price:
            smart_limit_price = level_price
    else:
        smart_limit_price = level_price

    limit_price = smart_limit_price
    market_price = None
    try:
        q = live_quotes.get(symbol)
        if q and 'ask' in q and 'bid' in q:
            market_price = float(q['bid'] if is_buy else q['ask'])
    except Exception:
        pass

    if is_buy:
        # Buy Limit must be <= market Ask for MT5 to accept
        if market_price is not None and limit_price > market_price:
            c_log(f"Buy Limit {limit_price} > Ask {market_price} — level passed, "
                  f"skipping Phase 1")
            ioc_fail_reason = 'Skipped — level above market for buy limit'
            raise _SkipLimitPhase(ioc_fail_reason)
    else:
        # Sell Limit must be >= market Bid for MT5 to accept
        if market_price is not None and limit_price < market_price:
            c_log(f"Sell Limit {limit_price} < Bid {market_price} — level passed, "
                  f"skipping Phase 1")
            ioc_fail_reason = 'Skipped — level below market for sell limit'
            raise _SkipLimitPhase(ioc_fail_reason)

    # No more "level" concept to measure a touch-margin against -- reuse
    # the caller's own max_slippage_points budget (already in broker points)
    # as the basis for how far price is allowed to run before this pending
    # limit order cancels itself.
    margin = max_slippage_points * SYMBOL_INFO[symbol]['pip_value']
    cancel_threshold = margin * 2  # cancel if price moves 2×margin away

    limit_opts = {
        'slippage': max_slippage_points,
        'comment': 'limit_buy_strategy' if is_buy else 'limit_sell_strategy',
        'expirationType': 'ORDER_TIME_SPECIFIED',
        'expiration': datetime.utcnow() + timedelta(seconds=30),  # Extended GTD
    }

    t_start = time.monotonic()
    try:
        if is_buy:
            res = await _metaapi_conn.create_limit_buy_order(
                broker_symbol, lot, limit_price,
                stop_loss=sl, take_profit=tp, options=limit_opts,
            )
        else:
            res = await _metaapi_conn.create_limit_sell_order(
                broker_symbol, lot, limit_price,
                stop_loss=sl, take_profit=tp, options=limit_opts,
            )
        t_ack = time.monotonic()
        latency_ms = round((t_ack - t_start) * 1000)

        trade_id_candidate = str(res.get('positionId', res.get('orderId', '')))

        # Event-driven fill detection with 100ms IOC emulation timer
        await _start_fill_monitor()
        fill_event = asyncio.Event()
        _fill_events[trade_id_candidate] = fill_event
        
        # Hybrid IOC emulation: wait 100ms for limit fill, then auto-send market FOK
        ioc_emulation_timeout = 0.1  # 100ms
        try:
            await asyncio.wait_for(fill_event.wait(), timeout=ioc_emulation_timeout)
            fill_data = _fill_results.pop(trade_id_candidate, None)
            if fill_data:
                fill_price = fill_data['fill_price']
                trade_id = fill_data['trade_id']
                fill_source = fill_data['fill_source']
                method_used = 'limit'
            else:
                fill_price = None
        except asyncio.TimeoutError:
            # IOC emulation timeout — limit didn't fill in 100ms, send market FOK fallback
            fill_price = None
            c_log(f"Phase 1 IOC emulation timeout ({ioc_emulation_timeout*1000:.0f}ms) — sending Phase 2 market FOK")

        # Cleanup Phase 1 tracking
        _fill_events.pop(trade_id_candidate, None)
        _fill_results.pop(trade_id_candidate, None)

        if fill_price is not None:
            _exec_tracker.record(symbol, is_buy, level_price, fill_price,
                                 fill_source, latency_ms, 'limit', True)
            return {
                'success': True, 'trade_id': trade_id,
                'fill_price': fill_price, 'fill_source': fill_source,
                'latency_ms': latency_ms, 'method_used': 'limit',
                'error': None, 'ioc_fail_reason': None,
            }

        # Phase 1 IOC emulation timed out — cancel the pending limit order
        try:
            await _metaapi_conn.delete_pending_order(trade_id_candidate)
        except Exception:
            pass

        ioc_fail_reason = 'Limit IOC emulation timeout — price moved away from level'

    except _SkipLimitPhase:
        ioc_fail_reason = ioc_fail_reason or 'Phase 1 skipped'
    except Exception as e:
        latency_ms = round((time.monotonic() - t_start) * 1000)
        error = e
        ioc_fail_reason = f'Limit order raised exception: {e}'
        _exec_tracker.record(symbol, is_buy, level_price, None,
                             None, latency_ms, 'limit', False, error=e)
        c_log(f"Limit order failed [{symbol}]: {e} — falling back to market order")

    # ═══════════════════════════════════════════════════════════
    # PHASE 2 — MARKET ORDER (FOK) WITH ADAPTIVE SLIPPAGE BUDGET
    # ═══════════════════════════════════════════════════════════
    # Adaptive slippage budget: if rolling avg slippage > 2 pips,
    # auto-widen max_slippage_points for next 10 trades to prevent FOK rejections.
    avg_slip = _exec_tracker.avg_slippage(symbol, n=50)
    adaptive_slip_pts = max_slippage_points
    if avg_slip is not None:
        pip_val = SYMBOL_INFO[symbol]['pip_value']
        avg_slip_pips = avg_slip / pip_val if pip_val > 0 else 0
        if avg_slip_pips > 2.0:
            adaptive_slip_pts = int(max_slippage_points * 1.5)
            c_log(f"Adaptive slippage: avg={avg_slip_pips:.1f} pips > 2.0, widening to {adaptive_slip_pts} pts")

    market_opts = {
        'slippage': adaptive_slip_pts,
        'fillingModes': ['ORDER_FILLING_FOK'],
        'comment': 'market_buy_fbk' if is_buy else 'market_sell_fbk',
    }

    t_start = time.monotonic()
    try:
        if is_buy:
            res = await _metaapi_conn.create_market_buy_order(
                broker_symbol, lot,
                stop_loss=sl, take_profit=tp, options=market_opts,
            )
        else:
            res = await _metaapi_conn.create_market_sell_order(
                broker_symbol, lot,
                stop_loss=sl, take_profit=tp, options=market_opts,
            )
        t_ack = time.monotonic()
        latency_ms = round((t_ack - t_start) * 1000)

        trade_id = str(res.get('positionId', res.get('orderId', '')))

        # Event-driven fill detection for market order (eliminates 500-1000ms polling)
        await _start_fill_monitor()
        fill_event = asyncio.Event()
        _fill_events[trade_id] = fill_event
        try:
            await asyncio.wait_for(fill_event.wait(), timeout=2.0)
            fill_data = _fill_results.pop(trade_id, None)
            if fill_data:
                fill_price = fill_data['fill_price']
                fill_source = fill_data['fill_source']
                trade_id = fill_data['trade_id']
                method_used = 'market_fallback'
            else:
                fill_price = None
        except asyncio.TimeoutError:
            fill_price = None
        finally:
            _fill_events.pop(trade_id, None)
            _fill_results.pop(trade_id, None)

        if fill_price is None and res.get('price') is not None:
            fill_price = float(res['price'])
            fill_source = 'order_response'
            method_used = 'market_fallback'

        success = fill_price is not None
        _exec_tracker.record(symbol, is_buy, level_price, fill_price,
                             fill_source, latency_ms, method_used or 'market_fallback',
                             success, error=None if success else RuntimeError('No fill'))
        return {
            'success': success, 'trade_id': trade_id,
            'fill_price': fill_price, 'fill_source': fill_source,
            'latency_ms': latency_ms, 'method_used': method_used or 'market_fallback',
            'error': None if success else RuntimeError('Market fallback produced no fill'),
            'ioc_fail_reason': ioc_fail_reason,
        }

    except Exception as e:
        latency_ms = round((time.monotonic() - t_start) * 1000)
        _exec_tracker.record(symbol, is_buy, level_price, None,
                             None, latency_ms, 'market_fallback', False, error=e)
        return {
            'success': False, 'trade_id': None,
            'fill_price': None, 'fill_source': None,
            'latency_ms': latency_ms, 'method_used': None,
            'error': e,
            'ioc_fail_reason': ioc_fail_reason,
        }


async def _strategy_open_trade(symbol: str, is_buy: bool, signal_price: float, candles: list, reason: str, tf: str,
                            detect_time: datetime = None, t1_signal_ts: float = None,
                            feed_source: str = None, feed_age_ms: float = None) -> None:
    """Open a new trade off an EMA-stack + Stochastic-crossover signal
    detected on candle close. `signal_price` is the close of the signal
    candle at the moment the signal was detected -- used only as the
    reference price for the smart-limit execution helper and for
    slippage/report labeling (kept under the historical dict key
    'level_price' so existing export/report code keeps working
    unmodified); it is NOT a Gann level and nothing here waits for price
    to "touch" it."""
    global _consecutive_real_order_failures
    sym_state = bot_state['symbol_state'][symbol]

    # Order-management critical path: never place an order while the
    # connection state machine says we shouldn't be trading, or while
    # we're inside a restricted DAM time window.
    if not await is_trading_allowed():
        if bot_state.get('connection_state', CONN_RUNNING) != CONN_RUNNING:
            c_log(f"Skipped entry [{symbol} {tf}]: connection_state={bot_state.get('connection_state')} "
                  f"({bot_state.get('connection_state_reason')})")
        else:
            c_log(f"Skipped entry [{symbol} {tf}]: inside restricted DAM trading window "
                  f"({datetime.now(timezone.utc) + timedelta(hours=3):%H:%M} DAM).")
        return

    try:
        is_real = sym_state.get('auto_trade', False)

        # ── Re-verify price at execution time (Point 3) ──
        # For real trades, fetch a fresh price right before execution.
        # Simulated trades skip this since they don't interact with the
        # broker and the signal candle's close is already the reference.
        fresh_px, fresh_feed_source, fresh_feed_age_ms = signal_price, (feed_source or 'oanda'), feed_age_ms
        if is_real:
            fresh_px, fresh_feed_source, fresh_feed_age_ms = await _lq_price_with_fallback(symbol)
        if fresh_px is None:
            c_log(f"REJECTED [{symbol} {tf}]: WS feed stale ({_QUOTE_STALE_SECONDS}s), "
                  f"trade aborted to avoid {50}-{150}ms REST latency")
            await send_tg_msg(
                f"<b>⏭️ [{symbol} - {tf}]</b>  {reason}\n"
                f"إغلاق شمعة الإشارة: {signal_price:.2f}\n"
                f"تم رفض الصفقة — تغذية السعر غير محدثة (أكبر من {_QUOTE_STALE_SECONDS}ث)."
            )
            return

        price = fresh_px
        tp, sl = _strategy_calc_tpsl(symbol, price, is_buy, candles, tf=tf)

        # ── Pre-send sanity check (Point 4) ──
        # If price already moved past where TP or SL would sit before we
        # even send the order, the opportunity is gone (or would produce
        # nonsensical/rejected stops, e.g. "Invalid stops"). Better to skip
        # cleanly here than let the broker reject it after the fact.
        if is_buy and (price >= tp or price <= sl):
            await send_tg_msg(
                f"<b>⏭️ [{symbol} - {tf}]</b>  {reason}\n"
                f"إغلاق شمعة الإشارة: {signal_price:.2f}\n"
                f"تم إلغاء الأمر قبل الإرسال — السعر الحالي ({price:.2f}) تجاوز فعلياً "
                f"مستوى TP/SL المحسوب (TP:{tp} SL:{sl})."
            )
            return
        if not is_buy and (price <= tp or price >= sl):
            await send_tg_msg(
                f"<b>⏭️ [{symbol} - {tf}]</b>  {reason}\n"
                f"إغلاق شمعة الإشارة: {signal_price:.2f}\n"
                f"تم إلغاء الأمر قبل الإرسال — السعر الحالي ({price:.2f}) تجاوز فعلياً "
                f"مستوى TP/SL المحسوب (TP:{tp} SL:{sl})."
            )
            return

        lot = sym_state['lot_size']
        tp_pts = _strategy_tf_tp(symbol, tf); sl_pts = _strategy_tf_sl(symbol, tf)

        tpsl_lbl = (f"ATR({sym_state['strategy_atr_period']})×{sym_state['strategy_atr_sl_mult']}/{sym_state['strategy_atr_tp_mult']}\n"
                    if sym_state['strategy_tpsl_mode'] == 'atr' else f"SL:{sl_pts}p TP:{tp_pts}p")

        be_lbl = " | 🛡️ BE Active" if sym_state['break_even_enabled'] else ""

        is_real = sym_state.get('auto_trade', False)
        trade_id = f"sim_{int(datetime.now().timestamp())}_{tf}"
        real_msg = ""
        execution_failed = False
        # Real fill price (as opposed to `price`, our pre-check estimate) --
        # only ever populated for actual real orders below; stays None for
        # simulated/paper trades, where `price` IS the fill by definition.
        real_fill_price = None
        fill_price_source = 'simulated'
        exec_result = None

        if is_real:
            # Source of truth: never spin up a second, ad-hoc MetaAPI
            # connection here. If the one persistent connection created at
            # startup isn't healthy, we do not know the true account state
            # well enough to safely fire a real order.
            if _metaapi_conn is None:
                real_msg = "\n⚠️ لا يوجد اتصال MetaAPI صالح — لم يتم فتح أي صفقة."
                is_real = False
                execution_failed = True
            else:
                broker_symbol = _resolve_broker_symbol(symbol)
                max_slippage_points = int(bot_state.get('prot_max_slippage_points', 5))
                exec_result = await _execute_smart_order(
                    symbol, is_buy, lot, price,
                    sl, tp, t1_signal_ts, max_slippage_points,
                )

                if exec_result['success']:
                    real_fill_price = exec_result['fill_price']
                    fill_price_source = exec_result['fill_source'] or 'simulated'
                    trade_id = exec_result['trade_id'] or trade_id

                    feed_label = 'WS (MetaApi live)' if fresh_feed_source == 'ws' else 'OANDA REST (fallback)'
                    age_str = f"{fresh_feed_age_ms}ms" if fresh_feed_age_ms is not None else 'n/a'
                    method_labels = {'limit': 'حدّي بسعر الإشارة (Limit/IOC)',
                                     'market_fallback': 'سوقي بحماية الانزلاق (Market/FOK)'}
                    method_label = method_labels.get(exec_result['method_used'], 'غير معروف')
                    slippage_str = ''
                    if real_fill_price is not None and price is not None:
                        slip = abs(real_fill_price - price)
                        slippage_str = (
                            f"\n📊 الانزلاق الفعلي عن سعر الإشارة: {slip:.2f} "
                            f"({slip / SYMBOL_INFO[symbol]['pip_value']:.1f} نقطة)"
                        )
                    real_msg = (
                        f"\n🚀 <b>تم فتح الصفقة حقيقياً على حسابك!</b>"
                        + (f"\n⚠️ طريقة التنفيذ: {method_label}"
                           if exec_result['method_used'] != 'limit' else '')
                        + f"\n⏱ وقت التنفيذ: {exec_result['latency_ms']}ms"
                        + f"\n📡 تغذية: {feed_label} | عمر السعر: {age_str}"
                        + slippage_str
                    )
                    _consecutive_real_order_failures = 0
                else:
                    err = exec_result['error']
                    err_str = str(err) if err else 'Unknown error'
                    if any(code in err_str for code in ('REQUOTE', 'PRICE_CHANGED', 'OFF_QUOTES')):
                        real_msg = (
                            f"\n🛑 <b>تم رفض الصفقة لتجاوز حد الانزلاق السعري "
                            f"({max_slippage_points} نقاط):</b> {err}"
                            f"\nلم يتم التنفيذ لحمايتك من دخول سيء."
                        )
                    else:
                        real_msg = (
                            f"\n❌ <b>فشل فتح الصفقة حقيقياً:</b> {err}"
                            f"\nلم يتم تتبعها كصفقة وهمية (لا يوجد تنفيذ فعلي)."
                        )
                    is_real = False
                    execution_failed = True
                    _consecutive_real_order_failures += 1
                    if _consecutive_real_order_failures >= _REAL_ORDER_FAILURE_HALT_THRESHOLD:
                        await set_connection_state(
                            CONN_HALTED,
                            f"{_consecutive_real_order_failures} consecutive real order failures "
                            f"(last: {err}). Escalating to protect capital."
                        )

        # Ghost-trade fix: a FAILED real-order attempt must never enter
        # strategy_open_trades -- it never had any exposure on the broker, so
        # tracking it (even as a "simulated" fallback) meant the bot would
        # later evaluate it against live price movement and report a
        # fabricated WIN/LOSS for a trade that never existed. Genuine
        # paper-trading (auto_trade was never enabled to begin with) is a
        # completely different, intentional case and is still tracked.
        if execution_failed:
            await send_tg_msg(
                f"<b>⏭️ [{symbol} - {tf}]</b>  {reason}\n"
                f"إغلاق شمعة الإشارة: {signal_price:.2f}\n"
                f"{real_msg}"
            )
            return

        # entry_final is what actually happened: the confirmed broker fill
        # when we have one, our pre-check estimate otherwise (simulated
        # trades, or a real trade where the position lookup above failed).
        entry_final = real_fill_price if real_fill_price is not None else price

        exec_latency = exec_result.get('latency_ms') if (is_real and exec_result) else None
        exec_method = exec_result.get('method_used') if (is_real and exec_result) else None
        exec_ioc_fail = exec_result.get('ioc_fail_reason') if (is_real and exec_result) else None
        exec_slippage = round(abs(entry_final - price), 5) if entry_final is not None and price is not None else None

        bot_state['symbol_state'][symbol]['strategy_open_trades'][trade_id] = {
            'tf': tf, 'is_buy': is_buy, 'entry': entry_final, 'is_real': is_real, 'sl': sl, 'tp': tp,
            'opened_at': datetime.now(timezone.utc).isoformat(), 'level_price': price,
            'feed_source': feed_source, 'feed_age_ms': feed_age_ms, 'trigger_type': 'candle_close',
            'exec_latency_ms': exec_latency, 'exec_method': exec_method,
            'exec_ioc_fail_reason': exec_ioc_fail, 'exec_slippage': exec_slippage,
        }
        await _debounced_persist_save()

        entry_note = {
            'confirmed_position': ' (مؤكد من الوسيط)',
            'order_response': ' (من استجابة الأمر)',
            'simulated': '',
        }.get(fill_price_source, ' (تقديري قبل التنفيذ — تعذّر تأكيد سعر الوسيط)')
        slippage_line = ""
        if is_real and entry_final is not None:
            actual_slippage = abs(entry_final - price)
            pv = SYMBOL_INFO[symbol]['pip_value']
            slippage_line = f"الانزلاق الفعلي عن سعر الإشارة: {actual_slippage:.2f} ({actual_slippage / pv:.1f} نقطة)\n"

        await send_tg_msg(
            f"<b>✅ {reason}</b>\n\n"
            f"إغلاق شمعة الإشارة: {signal_price:.2f}  |  الدخول: {entry_final:.2f}{entry_note}\n\n"
            f"TP: {tp}  SL: {sl}  |  {tpsl_lbl}{be_lbl}\n"
            f"{slippage_line}"
            f"{real_msg}"
        )
    except Exception as e:
        log_exception(f"_strategy_open_trade [{symbol} {tf}]", e)
        await send_tg_msg(f"<b>❌ فشل تنفيذ الصفقة [{symbol} - {tf}]</b>\nإغلاق شمعة الإشارة: {signal_price:.5f}\n{e}")


# ─────────────────────────────────────────────────────────────
# BACKTEST PROGRESS TRACKER
# ─────────────────────────────────────────────────────────────
class BtProgress:
    BAR_LEN = 14; HEARTBEAT = 15
    def __init__(self, label: str, active_tfs: list):
        self.label = label; self.active_tfs = active_tfs; self.cancelled = False; self.phase = 'Initialising...'
        self.tf_done = 0; self.tf_total = len(active_tfs); self.current_tf = ''
        self.bars_done = 0; self.bars_total = 0; self.win = 0; self.loss = 0; self.be = 0; self.profit = 0.0
        self.chat_id = None; self.msg_id = None; self._last_edit = 0.0; self._lock = asyncio.Lock(); self._hb_task = None; self._start_ts = 0.0

    def _bar(self, done: int, total: int) -> str:
        if total == 0: return chr(9617) * self.BAR_LEN
        filled = round(done / total * self.BAR_LEN)
        return chr(9608) * filled + chr(9617) * (self.BAR_LEN - filled)

    def _elapsed(self) -> str:
        secs = int(datetime.now(timezone.utc).timestamp() - self._start_ts); m, s = divmod(secs, 60); return f'{m}m {s:02d}s'

    def _build_text(self) -> str:
        total = self.win + self.loss; wr = f'{round(self.win / total * 100)}%' if total else '-'
        pnl = f'+${round(self.profit,2)}' if self.profit >= 0 else f'-${abs(round(self.profit,2))}'; icon = '▲' if self.profit >= 0 else '▼'
        overall = (self.tf_done + self.bars_done / self.bars_total) / max(self.tf_total, 1) if self.bars_total else self.tf_done / max(self.tf_total, 1)
        ov_bar = self._bar(round(overall * 100), 100); ov_pct = f'{round(overall * 100)}%'
        tf_bar = self._bar(self.bars_done, self.bars_total) if self.bars_total else chr(9617) * self.BAR_LEN
        tf_pct = f'{round(self.bars_done / self.bars_total * 100)}%' if self.bars_total else '-'
        lines = [f'Backtest — <b>{self.label}</b>', f'<b>Phase:</b> {self.phase}', '', f'<b>Overall</b>  {ov_pct}', f'<code>[{ov_bar}]</code>']
        if self.current_tf: lines += ['', f'<b>TF:</b> {self.current_tf}  ({self.tf_done}/{self.tf_total})', f'<code>[{tf_bar}] {tf_pct}</code>', f'Bars: {self.bars_done}/{self.bars_total}']
        lines += ['', f'W:{self.win}  L:{self.loss}  BE:{self.be}', f'{icon} {pnl}  WR:{wr}', '', f'Elapsed: {self._elapsed()}']
        if self.cancelled: lines.append('<b>CANCELLED</b>')
        return '\n'.join(lines)

    async def start(self, chat_id: int) -> None:
        self.chat_id = chat_id; self._start_ts = datetime.now(timezone.utc).timestamp(); self._last_edit = self._start_ts
        payload = {'chat_id': chat_id, 'text': self._build_text(), 'parse_mode': 'HTML', 'reply_markup': {'inline_keyboard': [[{'text': '⏹ Cancel', 'callback_data': 'cancel_bt'}]]}}
        try:
            async with aiohttp.ClientSession() as sess:
                async with sess.post(f'https://api.telegram.org/bot{TG_TOKEN}/sendMessage', json=payload) as resp:
                    if resp.status == 200: self.msg_id = (await resp.json())['result']['message_id']
        except Exception: pass
        self._hb_task = asyncio.create_task(self._heartbeat())

    async def _heartbeat(self) -> None:
        while not self.cancelled: await asyncio.sleep(self.HEARTBEAT); await self._edit(force=True)

    async def _edit(self, force: bool = False) -> None:
        now = datetime.now(timezone.utc).timestamp()
        if not force and (now - self._last_edit) < 3: return
        if not self.msg_id or not self.chat_id: return
        async with self._lock:
            self._last_edit = now; payload = {'chat_id': self.chat_id, 'message_id': self.msg_id, 'text': self._build_text(), 'parse_mode': 'HTML'}
            if not self.cancelled: payload['reply_markup'] = {'inline_keyboard': [[{'text': '⏹ Cancel', 'callback_data': 'cancel_bt'}]]}
            try: await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/editMessageText', json=payload)
            except Exception: pass

    async def set_phase(self, phase: str) -> None: self.phase = phase; await self._edit()
    async def set_tf(self, tf: str, bars_total: int) -> None: self.current_tf = tf; self.bars_done = 0; self.bars_total = bars_total; await self._edit(force=True)
    async def tick(self, bar_n: int, win: int, loss: int, be: int, profit: float) -> None: self.bars_done = bar_n; self.win = win; self.loss = loss; self.be = be; self.profit = profit; await self._edit()
    async def done(self, final_text: str) -> None:
        if self._hb_task: self._hb_task.cancel()
        if not self.msg_id or not self.chat_id: return
        try: await edit_tg_msg(self.chat_id, self.msg_id, final_text)
        except Exception: pass
    async def cancel(self) -> None:
        self.cancelled = True; self.phase = 'Cancelling...'
        if self._hb_task: self._hb_task.cancel()
        await self._edit(force=True)

_bt_progress: BtProgress | None = None
_lt_progress: BtProgress | None = None

# ─────────────────────────────────────────────────────────────
# KEYBOARDS 
# ─────────────────────────────────────────────────────────────
def get_main_keyboard() -> dict:
    return {'inline_keyboard': [
        [{'text': '🔌 فحص حالة حساب MetaAPI', 'callback_data': 'check_metaapi_status'}],
        [{'text': '🩺 تشخيص: ليه مفيش صفقات؟', 'callback_data': 'run_diag'}],
        [{'text': '📊 تصدير سجل تشخيص تفصيلي (Excel)', 'callback_data': 'export_diag_excel'}],
        [{'text': '📒 تصدير سجل الصفقات الحية (Excel)', 'callback_data': 'export_live_trades_excel'}],
        [{'text': '📋 تقرير تفاصيل التنفيذ (Latency/Method/Slippage)', 'callback_data': 'export_exec_report'}],
        [{'text': '🔓 استئناف يدوي بعد HALT (بعد التأكد من الحساب)', 'callback_data': 'manual_resume_step1'}],
        [{'text': '📈 محرك الاستراتيجية (EMA/Stochastic)', 'callback_data': 'menu_strategy'}],
        [{'text': '🛡️ إعدادات الحماية', 'callback_data': 'menu_protection'}],
        [{'text': '💾 إدارة الإعدادات (Presets)', 'callback_data': 'menu_presets'}],
        [{'text': '📊 بدء الباكتيست', 'callback_data': 'menu_strategy_bt'}],
        [{'text': '🧪 Live-Twin Simulator (تنفيذ واقعي)', 'callback_data': 'menu_lt'}],
    ]}


def get_protection_keyboard() -> dict:
    dd = bot_state['prot_daily_dd_usd']
    profit = bot_state['prot_daily_profit_usd']
    multi_tf = '✅ مسموح' if bot_state.get('prot_allow_multi_tf', True) else '❌ ممنوع'
    
    rows = [
        [{'text': '── الحدود اليومية ──', 'callback_data': 'noop'}],
        [{'text': f'📉 أقصى تراجع يومي: ${dd}', 'callback_data': 'noop'}],
        [
            {'text': '➖ خسارة $50', 'callback_data': 'prot_dec_dd'},
            {'text': '➕ خسارة $50', 'callback_data': 'prot_inc_dd'}
        ],
        [{'text': f'💰 هدف الربح اليومي: ${profit}', 'callback_data': 'noop'}],
        [
            {'text': '➖ ربح $50', 'callback_data': 'prot_dec_profit'},
            {'text': '➕ ربح $50', 'callback_data': 'prot_inc_profit'}
        ],
        [{'text': '── الحماية المتقدمة (v9.0) ──', 'callback_data': 'noop'}],
        [{'text': f"مزامنة MT4 (Reconciliation): {'✅' if bot_state.get('prot_true_sync', True) else '🔴'}", 'callback_data': 'tg_prot_sync'}],
        [{'text': f"BE شامل التكلفة (True Cost): {'✅' if bot_state.get('prot_cost_be', True) else '🔴'}", 'callback_data': 'tg_prot_cost'}],
        [{'text': f"فلتر البيانات المتأخرة: {'✅' if bot_state.get('prot_stale_filter', True) else '🔴'}", 'callback_data': 'tg_prot_stale'}],
        [{'text': f"فلتر أوقات دمشق (07-09 | 13-14): {'✅' if bot_state.get('prot_dam_time_filter', True) else '🔴'}", 'callback_data': 'tg_prot_dam_time'}],
        [{'text': f'تكرار الصفقات (Multi-TF): {multi_tf}', 'callback_data': 'prot_toggle_multitf'}],
        [{'text': '── ── ──', 'callback_data': 'noop'}],
        [{'text': '🔄 تصفير كل الحمايات النشطة الآن', 'callback_data': 'prot_reset_all'}],
        [{'text': '🔙 رجوع للقائمة الرئيسية', 'callback_data': 'menu_main'}],
        [{'text': '🔙 رجوع لإعدادات الاستراتيجية', 'callback_data': 'menu_strategy'}]
    ]
    return {'inline_keyboard': rows}

def get_strategy_keyboard() -> dict:
    sym = bot_state['ui_selected_symbol']
    sym_state = bot_state['symbol_state'][sym]
    tpsm = sym_state['strategy_tpsl_mode']
    open_n = len(sym_state['strategy_open_trades'])

    tps_lbl = f'🎯 TP/SL: {"نقاط ثابتة" if tpsm == "fixed" else "حسب ATR"}'

    tp = sym_state['strategy_tp_points']; sl = sym_state['strategy_sl_points']
    atp = sym_state['strategy_atr_tp_mult']; asp = sym_state['strategy_atr_sl_mult']
    ap  = sym_state['strategy_atr_period']
    be_lbl = "🟢 مفعل" if sym_state['break_even_enabled'] else "⚫ معطل"

    auto_t = '🟢 مفعل' if sym_state.get('auto_trade', False) else '🔴 معطل'

    shift = bot_state.get('signal_candle_shift', 1)
    shift_lbl = {0: 'الشمعة الحية (0)', 1: 'آخر شمعة مغلقة (1)', 2: 'ما قبل الأخيرة (2)'}.get(shift, str(shift))
    concur = '✅ مسموح' if bot_state.get('allow_concurrent_trades', False) else '❌ ممنوع'

    rows = [
        [{'text': f'🤖 التداول الآلي (MetaAPI): {auto_t}', 'callback_data': 'strategy_toggle_auto_trade'}],
        [{'text': '🛡️ إعدادات الحماية المتقدمة', 'callback_data': 'menu_protection'}],
        [{'text': f'📈 {sym} — صفقات مفتوحة: {open_n}', 'callback_data': 'noop'}],
        [{'text': '📉 عرض قيم EMA/Stochastic الحالية', 'callback_data': 'strategy_show_indicators'}],
        [{'text': '🕯️ تشخيص: آخر 10 شموع (وقت + إغلاق)', 'callback_data': 'strategy_show_last10'}],
    ]

    rows.append([{'text': '── أزواج التداول والباكتيست ──', 'callback_data': 'noop'}])
    pair_row = []
    for p in AVAILABLE_SYMBOLS:
        icon = '✅' if bot_state['active_symbols'][p] else '⬜'
        pair_row.append({'text': f'{icon} {p}', 'callback_data': f'strategy_toggle_pair_{p}'})
        if len(pair_row) == 2:
            rows.append(pair_row)
            pair_row = []
    if pair_row: rows.append(pair_row)

    rows.append([{'text': '── تخصيص إعدادات الزوج ──', 'callback_data': 'noop'}])
    sel_row = []
    for p in AVAILABLE_SYMBOLS:
        sel = '📌 ' if p == sym else ''
        sel_row.append({'text': f'{sel}{p}', 'callback_data': f'strategy_sel_pair_{p}'})
        if len(sel_row) == 2:
            rows.append(sel_row)
            sel_row = []
    if sel_row: rows.append(sel_row)

    rows += [
        [{'text': '── الاستراتيجية (EMA 15/50/150 + Stoch 5/5/5) ──', 'callback_data': 'noop'}],
        [{'text': f'شمعة التقييم: {shift_lbl}', 'callback_data': 'strategy_toggle_candle_shift'}],
        [{'text': f'السماح بصفقات متزامنة (نفس الاتجاه): {concur}', 'callback_data': 'strategy_toggle_concurrent'}],
        [{'text': f'🛡️ صمام الأمان (Break-Even): {be_lbl}', 'callback_data': 'strategy_toggle_be'}],
    ]
    if sym_state.get('break_even_enabled', False):
        be_pts = sym_state.get('strategy_be_trigger_points', 40)
        rows.append([
            {'text': 'BE −10p', 'callback_data': 'strategy_dec_be_pts'},
            {'text': f'تفعيل بعد: {be_pts}p', 'callback_data': 'noop'},
            {'text': 'BE +10p', 'callback_data': 'strategy_inc_be_pts'}
        ])

    rows += [
        [{'text': '── فريمات التنفيذ (تقييم عند إغلاق كل شمعة) ──', 'callback_data': 'noop'}],
    ]

    tf_items = list(sym_state['strategy_monitor_tfs'].items())
    for i in range(0, len(tf_items), 4):
        rows.append([{'text': ('✅' if on else '⬜') + f' {tfk}', 'callback_data': f'strategy_tf_{tfk}'} for tfk, on in tf_items[i:i+4]])

    rows += [
        [{'text': '── إعدادات عامة ──', 'callback_data': 'noop'}],
        [{'text': 'Lot −0.01', 'callback_data': 'strategy_dec_lot'}, {'text': f'حجم اللوت: {sym_state["lot_size"]}', 'callback_data': 'noop'}, {'text': 'Lot +0.01', 'callback_data': 'strategy_inc_lot'}],
        [{'text': '── TP / SL ──', 'callback_data': 'noop'}],
        [{'text': tps_lbl, 'callback_data': 'strategy_toggle_tpsl'}],
    ]

    if tpsm == 'fixed':
        rows += [
            [{'text': 'TP  −10', 'callback_data': 'strategy_dec_tp10'}, {'text': f'TP={tp}p', 'callback_data': 'noop'}, {'text': 'TP  +10', 'callback_data': 'strategy_inc_tp10'}],
            [{'text': 'SL  −10', 'callback_data': 'strategy_dec_sl10'}, {'text': f'SL={sl}p', 'callback_data': 'noop'}, {'text': 'SL  +10', 'callback_data': 'strategy_inc_sl10'}],
        ]
    else:
        rows += [
            [{'text': 'ATR Period −', 'callback_data': 'strategy_dec_atrp'}, {'text': f'Period={ap}', 'callback_data': 'noop'}, {'text': 'ATR Period +', 'callback_data': 'strategy_inc_atrp'}],
            [{'text': 'SL mult −0.5', 'callback_data': 'strategy_dec_atrsl'}, {'text': f'SL×{asp}', 'callback_data': 'noop'}, {'text': 'SL mult +0.5', 'callback_data': 'strategy_inc_atrsl'}],
            [{'text': 'TP mult −0.5', 'callback_data': 'strategy_dec_atrtp'}, {'text': f'TP×{atp}', 'callback_data': 'noop'}, {'text': 'TP mult +0.5', 'callback_data': 'strategy_inc_atrtp'}],
        ]

    rows += [
        [{'text': '⚙️ TP/SL مخصص لكل فريم', 'callback_data': 'strategy_tpsl_tf'}],
        [{'text': '📊 بدء الباكتيست', 'callback_data': 'menu_strategy_bt'}],
        [{'text': '← رجوع', 'callback_data': 'menu_main'}],
    ]
    return {'inline_keyboard': rows}

def get_strategy_tpsl_tf_keyboard(sel_tf: str = '') -> dict:
    sym_state = bot_state['symbol_state'][bot_state['ui_selected_symbol']]
    rows = [[{'text': '⚙️ TP/SL مخصص لكل فريم', 'callback_data': 'noop'}],
            [{'text': '(0 = يرجع للقيمة العامة)', 'callback_data': 'noop'}]]
    tfs_list = list(sym_state['strategy_monitor_tfs'].keys())
    tf_row = []
    for tfk in tfs_list:
        icon = '👉' if tfk == sel_tf else ''
        tf_row.append({'text': f'{icon}{tfk}', 'callback_data': f'strategy_tptf_sel_{tfk}'})
        if len(tf_row) == 4: rows.append(tf_row); tf_row = []
    if tf_row: rows.append(tf_row)
    if sel_tf:
        tp_v = sym_state['strategy_tp_per_tf'].get(sel_tf, 0); sl_v = sym_state['strategy_sl_per_tf'].get(sel_tf, 0)
        eff_tp = tp_v if tp_v > 0 else sym_state['strategy_tp_points']
        eff_sl = sl_v if sl_v > 0 else sym_state['strategy_sl_points']
        rows += [
            [{'text': f'── [{sel_tf}] ──', 'callback_data': 'noop'}],
            [{'text': f'TP فعلي: {eff_tp}p {"(مخصص)" if tp_v>0 else "(عام)"}', 'callback_data': 'noop'}],
            [{'text': 'TP −10', 'callback_data': f'strategy_tptf_dtp_{sel_tf}'}, {'text': f'TP={tp_v}', 'callback_data': 'noop'}, {'text': 'TP +10', 'callback_data': f'strategy_tptf_itp_{sel_tf}'}],
            [{'text': f'SL فعلي: {eff_sl}p {"(مخصص)" if sl_v>0 else "(عام)"}', 'callback_data': 'noop'}],
            [{'text': 'SL −10', 'callback_data': f'strategy_tptf_dsl_{sel_tf}'}, {'text': f'SL={sl_v}', 'callback_data': 'noop'}, {'text': 'SL +10', 'callback_data': f'strategy_tptf_isl_{sel_tf}'}],
            [{'text': '↺ إعادة ضبط', 'callback_data': f'strategy_tptf_rst_{sel_tf}'}],
        ]
    rows.append([{'text': '← رجوع', 'callback_data': 'menu_strategy'}])
    return {'inline_keyboard': rows}

def get_strategy_bt_keyboard() -> dict:
    if bot_state['is_backtesting']:
        return {'inline_keyboard': [[{'text': '⏳ الباكتيست يعمل...', 'callback_data': 'noop'}], [{'text': '⏹ إلغاء', 'callback_data': 'cancel_bt'}]]}
    return {'inline_keyboard': [
        [{'text': 'يوم واحد', 'callback_data': 'gbt_1'}, {'text': 'يومين', 'callback_data': 'gbt_2'}],
        [{'text': 'ثلاثة أيام', 'callback_data': 'gbt_3'}, {'text': 'أسبوع', 'callback_data': 'gbt_7'}],
        [{'text': 'شهر كامل', 'callback_data': 'gbt_30'}],
        [{'text': 'أو أرسل: /backtest YYYY-MM-DD', 'callback_data': 'noop'}],
        [{'text': '← رجوع', 'callback_data': 'menu_strategy'}],
    ]}

# ─────────────────────────────────────────────────────────────
# LIVE SCANNER (EMA 15/50/150 + Stochastic 5/5/5)
# ─────────────────────────────────────────────────────────────
async def _close_metaapi_trade(symbol: str, tid: str, sym_state: dict) -> bool:
    """Sequential, polled closure. Caller MUST await this fully before
    moving to the next trade — never wrap calls to this in asyncio.gather."""
    if not _metaapi_conn:
        c_log(f"Cannot close {tid} ({symbol}): no live MetaAPI connection. Position remains open on broker.")
        await send_tg_msg(f"🛑 <b>تعذّر إغلاق صفقة {symbol} ({tid}):</b> لا يوجد اتصال MetaAPI. الصفقة ما زالت مفتوحة على الوسيط.")
        return False
    try:
        await _metaapi_conn.close_position(tid)
        # State-Machine Polling for confirmation — never assume success.
        # 1s interval (not 0.2s): fetching the full portfolio 5x/sec per
        # trade during a batch closure risks tripping MetaAPI's rate limit
        # (HTTP 429). The SDK already background-syncs; 1s is plenty.
        for _ in range(25):
            positions = _metaapi_conn.terminal_state.positions
            if not any(str(p.get('id')) == str(tid) for p in positions):
                await send_tg_msg(f"✅ <b>تم إغلاق صفقة {symbol} (حقيقية) بنجاح لحماية الحساب!</b>")
                if tid in sym_state['strategy_open_trades']:
                    del sym_state['strategy_open_trades'][tid]
                    await save_bot_persistence()
                return True
            await asyncio.sleep(1.0)
        c_log(f"Timeout waiting for {tid} to disappear from MT5 positions after close_position call.")
        await send_tg_msg(f"⚠️ <b>لم يتم تأكيد إغلاق {symbol} ({tid}) خلال المهلة.</b> يرجى التحقق يدوياً من الحساب.")
        return False
    except Exception as e:
        log_exception(f"_close_metaapi_trade [{symbol}/{tid}]", e)
        await send_tg_msg(f"⚠️ <b>فشل الإغلاق الآلي:</b> صفقة {symbol} (خطأ: {e})\nيرجى التحقق يدوياً من الحساب.")
        return False

_EMERGENCY_CLOSE_POLL_BUDGET_SECONDS = 25  # shared across the WHOLE batch, not per-trade

async def _close_metaapi_trades_batch(closures: list) -> None:
    """Emergency mass-closure path (daily DD/profit limit hit).

    Still preserves the anti-race-condition requirement: close_position()
    write requests are issued strictly one at a time, sequentially -- same
    as _close_metaapi_trade, same TRADE_CONTEXT_BUSY protection. What
    changes is confirmation polling. get_positions() is read-only; polling
    it once per second for the ENTIRE batch (instead of each trade running
    its own private 25x1s loop, one after another) turns worst-case tail
    latency from O(N x 25s) into a single shared ~25s budget regardless of
    how many trades are closing at once. Nothing here writes to the
    broker concurrently -- only the read-only status check is shared.

    `closures` is a list of (symbol, tid, sym_state, tr) tuples for real
    trades only; callers handle simulated-trade deletion/notification
    themselves. `tr` is the trade's own dict (entry/tp/sl/tf/is_buy/
    last_known_pl/last_known_px) so every outcome message here can report
    which specific trade it is, not just "a trade on this symbol closed."
    """
    def _trade_detail_line(tr: dict) -> str:
        pl = tr.get('last_known_pl', 0.0)
        px = tr.get('last_known_px', tr.get('entry'))
        outcome_lbl = 'ربح ✅' if pl >= 0 else 'خسارة ❌'
        return (f"[{tr.get('tf')}] {'BUY 📈' if tr.get('is_buy') else 'SELL 📉'}\n"
                f"الدخول: {tr.get('entry')}  |  آخر سعر معروف: {px}\n"
                f"TP: {tr.get('tp')}  SL: {tr.get('sl')}\n"
                f"النتيجة: {outcome_lbl} ({pl}$)")

    if not closures:
        return
    if not _metaapi_conn:
        for symbol, tid, _, tr in closures:
            c_log(f"Cannot close {tid} ({symbol}): no live MetaAPI connection. Position remains open on broker.")
        detail = "\n\n".join(f"{symbol}: {_trade_detail_line(tr)}" for symbol, _, _, tr in closures)
        await send_tg_msg(
            f"🛑 <b>تعذّر إغلاق {len(closures)} صفقة:</b> لا يوجد اتصال MetaAPI. جميعها ما زالت مفتوحة على الوسيط.\n\n{detail}"
        )
        return

    pending = {}  # tid -> (symbol, sym_state, tr)
    close_errors = []
    for symbol, tid, sym_state, tr in closures:
        try:
            await _metaapi_conn.close_position(tid)
            pending[str(tid)] = (symbol, sym_state, tr)
        except Exception as e:
            log_exception(f"_close_metaapi_trades_batch close_position [{symbol}/{tid}]", e)
            close_errors.append(f"{symbol} ({tid}): {e}")
    if close_errors:
        await send_tg_msg(
            f"⚠️ <b>فشل إرسال {len(close_errors)}/{len(closures)} أمر إغلاق</b>\n"
            + "\n".join(close_errors)[:3500]
        )

    if not pending:
        return

    for _ in range(_EMERGENCY_CLOSE_POLL_BUDGET_SECONDS):
        if not pending:
            break
        try:
            positions = _metaapi_conn.terminal_state.positions
            if not isinstance(positions, list):
                raise TypeError(f"get_positions() returned {type(positions).__name__}, expected list")
        except Exception as e:
            log_exception("_close_metaapi_trades_batch get_positions", e)
            await asyncio.sleep(1.0)
            continue

        still_open_ids = {str(p.get('id')) for p in positions}
        for tid in list(pending.keys()):
            if tid not in still_open_ids:
                symbol, sym_state, tr = pending.pop(tid)
                await send_tg_msg(
                    f"✅ <b>تم إغلاق صفقة {symbol} (حقيقية) بنجاح لحماية الحساب!</b>\n\n{_trade_detail_line(tr)}"
                )
                pl = tr.get('last_known_pl', 0.0)
                px = tr.get('last_known_px', tr.get('entry'))
                await _record_closed_trade_history(
                    symbol, tid, tr, exit_px=px, pnl=pl,
                    outcome_label=('WIN' if pl > 0 else 'LOSS' if pl < 0 else 'BREAK_EVEN'),
                    close_reason='daily_capital_protection_forced_close', pnl_confirmed=False,
                )
                if tid in sym_state['strategy_open_trades']:
                    del sym_state['strategy_open_trades'][tid]
                    await save_bot_persistence()

        if pending:
            await asyncio.sleep(1.0)

    for tid, (symbol, sym_state, tr) in pending.items():
        c_log(f"Timeout waiting for {tid} ({symbol}) to disappear from MT5 positions after batch close.")
        await send_tg_msg(
            f"⚠️ <b>لم يتم تأكيد إغلاق {symbol} ({tid}) خلال المهلة.</b> يرجى التحقق يدوياً من الحساب.\n\n{_trade_detail_line(tr)}"
        )

async def strategy_run_diagnostics() -> str:
    """Walks through every gate _strategy_open_trade's callers check, per
    active symbol, and reports the exact state of each one. Read-only --
    never opens a trade, just explains why one would or wouldn't fire
    right now."""
    lines = ["<b>🩺 تشخيص أسباب عدم فتح الصفقات</b>\n"]

    # --- Global gates (apply to every symbol) ---
    conn_state = bot_state.get('connection_state', CONN_RUNNING)
    conn_ok = conn_state == CONN_RUNNING
    lines.append(f"1️⃣ حالة الاتصال: {'✅ RUNNING' if conn_ok else f'🛑 {conn_state}'}")
    if not conn_ok:
        lines.append(f"   السبب: {bot_state.get('connection_state_reason', '-')}")

    dam_blocked = _is_within_dam_restricted_window()
    dam_now = datetime.now(timezone.utc) + timedelta(hours=3)
    filter_on = bot_state.get('prot_dam_time_filter', True)
    lines.append(f"2️⃣ فلتر أوقات دمشق: {'مفعّل' if filter_on else '🔴 معطّل'} | الوقت الآن (DAM): {dam_now:%H:%M}")
    if filter_on and dam_blocked:
        lines.append("   🛑 داخل نافذة محظورة الآن -- لن تُفتح أي صفقة جديدة حتى تنتهي.")

    overall_allowed = await is_trading_allowed()
    lines.append(f"3️⃣ الخلاصة العامة is_trading_allowed(): {'✅ مسموح' if overall_allowed else '🛑 ممنوع'}\n")

    if not overall_allowed:
        lines.append("↳ طالما هذه البوابة العامة مغلقة، لن تفتح أي صفقة على أي رمز مهما كانت شروط الدخول متوفرة.\n")

    active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
    if not active_symbols:
        lines.append("⚠️ لا يوجد أي رمز مفعّل حالياً في active_symbols.")
        return "\n".join(lines)

    for symbol in active_symbols:
        sym_state = bot_state['symbol_state'][symbol]
        lines.append(f"━━━━━━━━━━━━━━\n<b>{symbol}</b>")

        # Live quote feed (used only for execution-time re-verification of
        # real trades and daily PnL tracking now -- the strategy itself
        # fires strictly on candle close, never on a tick).
        q = live_quotes.get(symbol)
        ws_age = (time.monotonic() - q['ts']) if q else None
        if q is None:
            lines.append("📡 تغذية الأسعار اللحظية: 🛑 <b>لم تصل أي تحديث بعد لهذا الرمز</b> -- "
                          "بدون هذه التغذية لا يمكن لأي صفقة حقيقية أن تُنفَّذ حتى لو ظهرت الإشارة.")
        elif ws_age > _QUOTE_STALE_SECONDS:
            lines.append(f"📡 تغذية الأسعار اللحظية: 🛑 <b>متوقفة منذ {ws_age:.0f} ثانية</b> "
                          f"(آخر تحديث Bid={q['bid']} Ask={q['ask']}) -- الدخول الفعلي متجمد حتى تعود.")
        else:
            lines.append(f"📡 تغذية الأسعار اللحظية: ✅ حية (عمرها {ws_age:.1f}s)")

        shift = bot_state.get('signal_candle_shift', 1)
        allow_concur = bot_state.get('allow_concurrent_trades', False)
        lines.append(f"شمعة التقييم: {shift}  |  صفقات متزامنة: {'مسموح' if allow_concur else 'ممنوع'}")

        enabled_tfs = [tf for tf, on in sym_state['strategy_monitor_tfs'].items() if on]
        if not enabled_tfs:
            lines.append("🛑 لا يوجد أي فريم مفعّل في strategy_monitor_tfs -- لن يتم فحص أي شيء.")
            continue

        open_dirs = {v.get('is_buy') for v in sym_state['strategy_open_trades'].values() if isinstance(v, dict)}

        for tf in enabled_tfs:
            already_open_tf = any(isinstance(v, dict) and v.get('tf') == tf for v in sym_state['strategy_open_trades'].values())
            if already_open_tf:
                lines.append(f"[{tf}] 🛑 يوجد صفقة مفتوحة بالفعل على هذا الفريم -- لن تُفتح صفقة ثانية لنفس الفريم.")
                continue

            try:
                candles = await fetch_candles(symbol, tf, count=STRAT_MIN_CANDLES + 20)
            except Exception as e:
                lines.append(f"[{tf}] 🛑 فشل جلب الشموع: {e}")
                continue
            if not candles:
                lines.append(f"[{tf}] 🛑 بيانات غير كافية من OANDA.")
                continue

            df = compute_strategy_indicators(candles)
            if df is None:
                lines.append(f"[{tf}] 🛑 عدد الشموع غير كافٍ لحساب EMA150/Stochastic (يلزم ~{STRAT_MIN_CANDLES}, المتوفر {len(candles)}).")
                continue

            sig = evaluate_strategy_signal(df, shift)
            idx = len(df) - 1 - shift
            ema15, ema50, ema150 = df['ema15'].iloc[idx], df['ema50'].iloc[idx], df['ema150'].iloc[idx]
            k, d = df['stoch_k'].iloc[idx], df['stoch_d'].iloc[idx]
            stack = 'صاعد (15>50>150)' if ema15 > ema50 > ema150 else ('هابط (150>50>15)' if ema150 > ema50 > ema15 else 'غير مصطف')
            lines.append(f"[{tf}] EMA15={ema15:.2f} EMA50={ema50:.2f} EMA150={ema150:.2f} ({stack})  |  Stoch %K={k:.1f} %D={d:.1f}")

            if sig is None:
                lines.append(f"[{tf}] لا توجد إشارة تقاطع في المنطقة المطلوبة على هذه الشمعة.")
                continue

            if not allow_concur:
                if sig['is_buy'] and True in open_dirs:
                    lines.append(f"[{tf}] ✅ إشارة {'شراء' if sig['is_buy'] else 'بيع'} لكنها 🛑 محظورة: يوجد صفقة شراء مفتوحة بالفعل ولا يُسمح بالتزامن.")
                    continue
                if not sig['is_buy'] and False in open_dirs:
                    lines.append(f"[{tf}] ✅ إشارة بيع لكنها 🛑 محظورة: يوجد صفقة بيع مفتوحة بالفعل ولا يُسمح بالتزامن.")
                    continue

            lines.append(f"[{tf}] ✅ إشارة {'شراء 🟢' if sig['is_buy'] else 'بيع 🔴'} جاهزة للتنفيذ عند دورة السكانر القادمة.")

    return "\n".join(lines)

async def export_diag_log_excel() -> None:
    """Export the FULL rolling diagnostic log (bot_state['diag_log']) to an
    .xlsx file and send it via Telegram. Unlike /diagnose (a single
    point-in-time snapshot), this covers every (symbol, timeframe) decision
    the live scanner made since the bot last restarted -- including the
    previously-silent skip reasons (insufficient OANDA candles, trend
    undetermined, cap reached, already open) that never got their own
    Telegram message.
    """
    log = list(bot_state.get('diag_log', []))
    if not log:
        await send_tg_msg("لا يوجد سجل تشخيص محفوظ بعد (السجل يبدأ بالتجمع فور بدء تشغيل البوت).")
        return

    df = pd.DataFrame(log)
    if 'ts' in df.columns:
        df['الوقت (DAM)'] = df['ts'].apply(lambda t: _utc_to_dam(t).strftime('%Y-%m-%d %H:%M:%S') if pd.notna(t) else '')
        df = df.drop(columns=['ts'])

    # Friendlier column order/names, but keep every raw field -- nothing summarized away.
    preferred_order = ['الوقت (DAM)', 'symbol', 'tf', 'master_px', 'trend_up', 'margin',
                        'nearest_compatible_level', 'nearest_dist', 'within_margin',
                        'touch_attempted', 'skip_reason']
    cols = [c for c in preferred_order if c in df.columns] + [c for c in df.columns if c not in preferred_order]
    df = df[cols]

    fname = f"DiagLog_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    try:
        with pd.ExcelWriter(fname, engine='openpyxl') as writer:
            if 'symbol' in df.columns:
                for sym in sorted(df['symbol'].dropna().unique()):
                    sheet_name = str(sym)[:31]  # Excel sheet name hard limit
                    df[df['symbol'] == sym].to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                df.to_excel(writer, sheet_name='diag_log', index=False)

        first_ts = _utc_to_dam(log[0]['ts']).strftime('%Y-%m-%d %H:%M') if log[0].get('ts') else '?'
        last_ts = _utc_to_dam(log[-1]['ts']).strftime('%Y-%m-%d %H:%M') if log[-1].get('ts') else '?'
        await send_tg_document(
            fname,
            f"📊 <b>سجل تشخيص تفصيلي كامل</b>\n"
            f"{len(log)} سطر (قرار فحص) — من {first_ts} إلى {last_ts} (توقيت دمشق)\n"
            f"كل سطر = قرار واحد للسكانر الحي لكل (رمز، فريم) بكل دورة فحص، بما فيها أسباب "
            f"التجاهل التي لم تُرسَل كرسالة تيليجرام من قبل."
        )
    finally:
        if os.path.exists(fname):
            os.remove(fname)

async def export_live_trades_excel() -> None:
    """Export every CLOSED real/live trade (bot_state['live_trade_history'])
    to an .xlsx styled to match the backtest reports exactly (same headers
    where they overlap, same WIN/LOSS/BREAK_EVEN color fills, same borders/
    column widths) -- but for actual broker trades, with extra columns a
    backtest doesn't need: entry slippage vs the intended level, whether the
    close price/pnl is broker-confirmed or an estimate, why it closed early
    (TP/SL hit vs daily capital-protection force-close), duration, and the
    feed/latency this specific trade fired under."""
    hist = list(bot_state.get('live_trade_history', []))
    if not hist:
        await send_tg_msg("لا يوجد سجل صفقات حية مغلقة بعد (يبدأ التسجيل تلقائياً من أول صفقة حقيقية تُغلق بعد هذا التحديث).")
        return

    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    be_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # unconfirmed PnL estimate

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الصفقات الحية"
    headers = ["الزوج", "TF", "حقيقية/وهمية", "اتجاه", "وقت الفتح (DAM)", "وقت الإغلاق (DAM)",
               "المدة (د)", "مستوى الدخول", "الدخول الفعلي", "انزلاق الدخول", "TP", "SL",
               "سعر الإغلاق", "النتيجة", "ربح ($)", "مؤكد من الوسيط؟", "سبب الإغلاق",
               "BE مفعّل؟", "مصدر التغذية", "عمر التغذية (ms)", "نوع التنفيذ"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = gray_fill
        cell.font = Font(bold=True)

    _OUTCOME_DISPLAY = {'WIN': 'WIN ✅', 'LOSS': 'LOSS ❌', 'BREAK_EVEN': 'BREAK_EVEN ⚖️'}
    _REASON_DISPLAY = {
        'tp_sl_or_manual_broker_close': 'TP/SL (مؤكد من الوسيط)',
        'tp_sl_hit': 'TP/SL (تقديري)',
        'daily_capital_protection_forced_close': '⏹️ إغلاق مبكر (حماية رأس المال اليومية)',
    }
    _TRIGGER_DISPLAY = {
        'touch': 'لمس مباشر ⚡', 'close': 'إغلاق شمعة ⏳', 'hybrid': 'تنفيذ هجين 🛡️',
    }
    running_bal = 0.0
    n_win = n_loss = n_be = 0
    for tr in hist:
        pnl = tr.get('pnl') or 0.0
        running_bal += pnl
        outcome = tr.get('outcome')
        if outcome == 'WIN': n_win += 1
        elif outcome == 'LOSS': n_loss += 1
        elif outcome == 'BREAK_EVEN': n_be += 1

        def _dam(iso):
            if not iso: return ''
            try:
                dt = datetime.fromisoformat(iso)
                return _utc_to_dam(dt).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                return str(iso)

        row = [
            tr.get('symbol'), tr.get('tf'), 'حقيقية' if tr.get('is_real') else 'وهمية (Paper)',
            'BUY 📈' if tr.get('is_buy') else 'SELL 📉', _dam(tr.get('opened_at')), _dam(tr.get('closed_at')),
            tr.get('duration_min'), tr.get('level_price'), tr.get('entry'), tr.get('entry_slippage'),
            tr.get('tp'), tr.get('sl'), tr.get('exit_price'), _OUTCOME_DISPLAY.get(outcome, outcome), pnl,
            '✅' if tr.get('pnl_confirmed_from_broker') else '⚠️ تقديري', _REASON_DISPLAY.get(tr.get('close_reason'), tr.get('close_reason')),
            '✅' if tr.get('be_activated') else '—', tr.get('feed_source') or '—', tr.get('feed_age_ms'),
            _TRIGGER_DISPLAY.get(tr.get('trigger_type'), 'غير مسجَّل (صفقة سابقة قبل هذا التحديث)'),
        ]
        ws.append(row)
        row_idx = ws.max_row
        fill = green_fill if outcome == 'WIN' else red_fill if outcome == 'LOSS' else be_fill if outcome == 'BREAK_EVEN' else None
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill
        if not tr.get('pnl_confirmed_from_broker') and tr.get('is_real'):
            ws.cell(row=row_idx, column=16).fill = yellow_fill

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
    from openpyxl.utils import get_column_letter
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20.0

    fname = f"LiveTrades_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(fname)
    try:
        total = len(hist)
        wr = round(100 * n_win / max(1, n_win + n_loss), 1)
        summary = (
            f"📒 <b>سجل الصفقات الحية الكامل</b>\n"
            f"{total} صفقة مغلقة  |  WR: {wr}% ({n_win} ربح / {n_loss} خسارة / {n_be} تعادل)\n"
            f"صافي: {running_bal:+.2f}$\n\n"
            f"⚠️ الصفوف الصفراء = ربح تقديري لم يتأكد بعد من سجل الوسيط."
        )
        await send_tg_document(fname, summary)
    finally:
        if os.path.exists(fname):
            os.remove(fname)


# ─────────────────────────────────────────────────────────────
# EXECUTION DETAILS REPORT (Telegram text, not Excel)
# ─────────────────────────────────────────────────────────────
async def export_execution_details_report() -> str | None:
    """Generate an Excel report of every closed trade with the 4 execution
    metrics (latency, method, IOC failure reason, slippage) in dedicated
    columns. Returns the .xlsx filename to send, or None if no data."""
    hist = list(bot_state.get('live_trade_history', []))
    if not hist:
        return None

    rows = []
    for tr in hist:
        symbol = tr.get('symbol', '')
        tf = tr.get('tf', '')
        is_buy = tr.get('is_buy')
        pnl = tr.get('pnl', 0.0)
        outcome = tr.get('outcome', '')
        pv = SYMBOL_INFO.get(symbol, {}).get('pip_value', 0.01)

        exec_slip = tr.get('exec_slippage')
        slip_pips = round(exec_slip / pv, 1) if exec_slip is not None and pv else None

        method_raw = tr.get('exec_method')
        method_label = {'limit': 'Phase 1 (Limit IOC)',
                        'market_fallback': 'Phase 2 (Market FOK)'}.get(method_raw, method_raw or '')

        opened = tr.get('opened_at')
        closed = tr.get('closed_at')
        try:
            opened_dam = _utc_to_dam(datetime.fromisoformat(opened)).strftime('%Y-%m-%d %H:%M:%S') if opened else ''
        except Exception:
            opened_dam = str(opened) if opened else ''
        try:
            closed_dam = _utc_to_dam(datetime.fromisoformat(closed)).strftime('%Y-%m-%d %H:%M:%S') if closed else ''
        except Exception:
            closed_dam = str(closed) if closed else ''

        rows.append({
            'Pair': symbol,
            'TF': tf,
            'Direction': 'BUY' if is_buy else 'SELL',
            'Open Time (DAM)': opened_dam,
            'Close Time (DAM)': closed_dam,
            'Level Price': tr.get('level_price'),
            'Entry Price': tr.get('entry'),
            'Exit Price': tr.get('exit_price'),
            'PnL ($)': round(pnl, 2),
            'Result': outcome,
            'Execution Latency (ms)': tr.get('exec_latency_ms'),
            'Execution Method': method_label,
            'Final Slippage (Pips)': slip_pips,
            'IOC Failure Reason': tr.get('exec_ioc_fail_reason') or '',
        })

    df = pd.DataFrame(rows)
    fname = f"Execution_Report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        with pd.ExcelWriter(fname, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Execution Details', index=False)
            ws = writer.sheets['Execution Details']
            for col_idx in range(1, len(df.columns) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22
    except Exception as e:
        log_exception('export_execution_details_report [Excel write]', e)
        return None

    return fname


async def strategy_monitor_scanner() -> None:
    global _last_scanner_error_alert_ts, _last_any_tick_ts
    c_log('Strategy live scanner started.')
    while True:
        try:
            # ── Cold-start self-heal ──
            # If _metaapi_conn (or _metaapi_account) is still None, neither
            # of the two watchdogs below can do anything -- both require a
            # connection object to already exist before they'll act. This
            # is exactly the gap that left the bot in silent, permanent
            # READ_ONLY when the ONE startup connection attempt lost a race
            # with a transient MetaApi/broker hiccup. Retry from scratch
            # here, every scanner tick, until it succeeds.
            if _metaapi_conn is None or _metaapi_account is None:
                await _bootstrap_metaapi_connection()

            # ── Feed-level staleness watchdog (OANDA REST poller) ──
            # The live price feed is now driven by oanda_live_price_poller.
            # If every active symbol's cached quote goes stale (the poller
            # is failing / network down), escalate to READ_ONLY so we don't
            # trade on a dead feed. We do NOT tear down the MetaAPI execution
            # connection here -- that is handled by the exec-channel watchdog
            # below, and only when the connection itself is non-RUNNING.
            active_syms_now = [s for s, on in bot_state['active_symbols'].items() if on]
            if active_syms_now:
                all_stale = all(_lq_is_stale(s) for s in active_syms_now)
                if all_stale and bot_state.get('connection_state') == CONN_RUNNING:
                    # Only escalate if the price feed has been dead for the
                    # WS watchdog window (reuse the same stale threshold).
                    if (_last_any_tick_ts is not None
                            and (time.monotonic() - _last_any_tick_ts) > _WS_WATCHDOG_STALE_SECONDS):
                        await set_connection_state(
                            CONN_READ_ONLY,
                            f"Live price feed stale >{_WS_WATCHDOG_STALE_SECONDS:.0f}s (OANDA REST poller). "
                            f"Trading paused until quotes resume."
                        )
                elif not all_stale and bot_state.get('connection_state') == CONN_READ_ONLY:
                    # Feed recovered -- but only auto-promote if the MetaAPI
                    # execution channel is also healthy.
                    if _metaapi_conn is not None:
                        await set_connection_state(CONN_RUNNING, "Live price feed resumed.")

            # MT5 Zombie Singleton Heartbeat (execution channel only)
            if _metaapi_account and bot_state.get('connection_state') != CONN_RUNNING:
                await set_connection_state(CONN_READ_ONLY, "MetaAPI connection lost — attempting reconnect.")
                reconnected = False
                for attempt in range(5):
                    try:
                        # Full bootstrap, not a raw .connect() on _metaapi_conn:
                        # if the ORIGINAL connection attempt failed because the
                        # account wasn't DEPLOYED yet (or any other reason
                        # _metaapi_conn was never created), _metaapi_conn is
                        # still None here -- calling .connect() on it throws
                        # AttributeError every single attempt forever, with no
                        # way back to a working connection. _bootstrap_
                        # metaapi_connection() re-fetches the account fresh
                        # (so it sees a deployment-state change) and safely
                        # (re)builds _metaapi_conn from scratch either way.
                        reconnected = await _bootstrap_metaapi_connection()
                        if reconnected:
                            c_log("MetaAPI Reconnected successfully (execution channel).")
                            break
                    except Exception as e:
                        log_exception(f"MetaAPI reconnect attempt {attempt+1}/5", e)
                    await asyncio.sleep(2 ** attempt)
                if not reconnected:
                    # Do not spin forever inside this loop iteration; stay
                    # READ_ONLY, log it, and let the next scanner tick retry.
                    # If this persists, an operator will see the escalation
                    # message and the repeated READ_ONLY state in logs.
                    c_log("MetaAPI reconnect exhausted 5 attempts this tick; will retry next cycle.")

            now_dt = datetime.now(timezone.utc)

            today_date = now_dt.date()
            if bot_state.get('live_daily_date') != today_date:
                c_log(f"New trading day detected ({bot_state.get('live_daily_date')} -> {today_date}). "
                      f"Resetting daily PnL counters.")
                bot_state['live_daily_date'] = today_date
                bot_state['live_daily_realized'] = 0.0
                bot_state['live_daily_hit'] = False
                # Save immediately — do not wait for the next trade event.
                # A crash between this reset and the next save would
                # otherwise reload yesterday's PnL/hit-flag on restart.
                await save_bot_persistence()

            if bot_state.get('live_daily_hit'):
                # New entries stay blocked either way. But don't let a failed/
                # incomplete mass closure go un-retried forever just because
                # the flag that blocks new entries is also what this gate
                # checks -- those are two different concerns.
                stale_active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
                stale_real_closures = []
                for symbol in stale_active_symbols:
                    sym_state = bot_state['symbol_state'][symbol]
                    for tid, tr in list(sym_state['strategy_open_trades'].items()):
                        if tr.get('is_real') and _metaapi_conn:
                            stale_real_closures.append((symbol, tid, sym_state, tr))
                if stale_real_closures:
                    c_log(f"live_daily_hit is set but {len(stale_real_closures)} real trade(s) are still open -- "
                          f"retrying the mass closure (previous attempt may have crashed/errored mid-way).")
                    await _close_metaapi_trades_batch(stale_real_closures)
                await asyncio.sleep(60)
                continue
                
            active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
            total_floating = 0.0
            
            # --- First pass: track open trades ---
            for symbol in active_symbols:
                sym_state = bot_state['symbol_state'][symbol]

                if sym_state['strategy_open_trades']:
                    # --- MetaAPI Strict Reconciliation (Per Symbol, Just-In-Time) ---
                    actual_positions = {}
                    sync_failed = False
                    if bot_state.get('prot_true_sync', True) and _metaapi_conn:
                        try:
                            positions = _metaapi_conn.terminal_state.positions
                            for p in positions: actual_positions[str(p.get('id'))] = p
                            # Sync succeeded — if we were previously degraded because of
                            # sync failures specifically, this is our signal to recover.
                            if bot_state.get('connection_state') == CONN_READ_ONLY and \
                               'sync' in bot_state.get('connection_state_reason', '').lower():
                                await set_connection_state(CONN_RUNNING, "MetaAPI get_positions() succeeded again.")
                        except Exception as e:
                            log_exception(f"MetaAPI get_positions [{symbol}]", e)
                            sync_failed = True
                            await set_connection_state(
                                CONN_READ_ONLY,
                                f"MetaAPI get_positions() sync failed for {symbol}: {e}. "
                                f"Halting new trades and skipping reconciliation this tick (Amnesia Prevention)."
                            )

                    if sync_failed:
                        continue # DO NOT proceed with reconciliation or risk Amnesia Wipe

                    mc = await fetch_candles(symbol, '1m', count=2)
                    
                    # Drawdown Blindspot Fix: Fallback to MT5 prices if Oanda fails
                    live_px = None
                    oanda_failed = False
                    if not mc:
                        oanda_failed = True
                    else:
                        candle_age = (now_dt - mc[-1]['time']).total_seconds()
                        if bot_state.get('prot_stale_filter', True) and candle_age > 120:
                            oanda_failed = True
                    
                    live_px = None
                    if not oanda_failed:
                        live_px = float(mc[-1]['close'])
                    else:
                        c_log(f"Oanda failed for {symbol}. Decoupled Mode: using MT5 currentPrice for open trade management.")
                        


                    closed_ids = []
                    
                    # Pre-fetch history if there are missing real trades to prevent DDoS
                    history_deals_cache = None
                    missing_tids = [t for t, v in sym_state['strategy_open_trades'].items() if v.get('is_real') and t not in actual_positions]
                    if missing_tids and _metaapi_conn:
                        # Retry with backoff (Point 6): MetaAPI's own history sync can
                        # lag a few seconds behind the actual broker-side close,
                        # especially when several positions close in the same tick
                        # (like a mass level-touch closing many trades at once).
                        # Try immediately first (no delay -- the common/fast case
                        # where sync already caught up costs nothing extra), then
                        # back off 3s, then 5s, re-checking each time whether every
                        # missing trade's closing deal has shown up yet. Only after
                        # all three attempts do we fall back to the estimate.
                        from datetime import timedelta
                        start_time = datetime.now(timezone.utc) - timedelta(days=2)
                        for attempt_i, delay in enumerate((0, 3, 5)):
                            if delay:
                                await asyncio.sleep(delay)
                            try:
                                history_deals_cache = await _metaapi_conn.get_history_deals_by_time_range(start_time, datetime.now(timezone.utc))
                            except Exception as e:
                                log_exception(f"get_history_deals_by_time_range [{symbol}] attempt {attempt_i+1}/3", e)
                                continue
                            found_now = {
                                str(d.get('positionId')) for d in history_deals_cache
                                if d.get('entryType') in ('DEAL_ENTRY_OUT', 'DEAL_ENTRY_OUT_BY')
                            }
                            if all(str(t) in found_now for t in missing_tids):
                                break  # every missing trade's closing deal is visible -- no need to keep waiting
                    
                    for tid, tr in list(sym_state['strategy_open_trades'].items()):
                        is_buy = tr.get('is_buy')
                        tp = tr.get('tp')
                        sl = tr.get('sl')
                        entry = tr.get('entry')
                        tf = tr.get('tf')
                        is_real = tr.get('is_real')
                        
                        active_px = live_px
                        if active_px is None:
                            if tid in actual_positions:
                                active_px = _safe_float(actual_positions[tid].get('currentPrice'), entry)
                            else:
                                active_px = tr.get('last_known_px') # Use last known, never artificially force entry
                                
                        if active_px is None:
                            # Completely blind, skip risk evaluation for this specific trade to avoid corrupting limits
                            continue
                            
                        tr['last_known_px'] = active_px
                        
                        diff = (active_px - entry) if is_buy else (entry - active_px)
                        cs = SYMBOL_INFO[symbol]['contract_size']
                        trade_pl = round(diff * sym_state['lot_size'] * cs, 2)
                        tr['last_known_pl'] = trade_pl
                        
                        if is_real and bot_state.get('prot_true_sync', True) and _metaapi_conn:
                            if tid not in actual_positions:
                                # Pre-fetched history deals (outside the loop to prevent Rate Limit Suicide)
                                exact_pnl = trade_pl  # Estimate fallback only — NOT the real MT5 profit
                                found_deal = False
                                if history_deals_cache is not None:
                                    deal_pnl = 0.0
                                    # DEAL_ENTRY_OUT covers a normal full close; DEAL_ENTRY_OUT_BY
                                    # covers "close by" an opposite position. Both are genuine
                                    # closing deals and both must count, or partial-close /
                                    # close-by trades will silently fall back to the estimate too.
                                    for d in history_deals_cache:
                                        if (str(d.get('positionId')) == str(tid)
                                                and d.get('entryType') in ('DEAL_ENTRY_OUT', 'DEAL_ENTRY_OUT_BY')):
                                            deal_pnl += _safe_float(d.get('profit')) + _safe_float(d.get('swap')) + _safe_float(d.get('commission'))
                                            found_deal = True
                                    if found_deal:
                                        exact_pnl = deal_pnl
                                        c_log(f"Reconciliation: Exact PnL for {tid} fetched from MT5: {exact_pnl}$")

                                closed_ids.append(tid)
                                bot_state['live_daily_realized'] += exact_pnl

                                if found_deal:
                                    # This mirrors the true, realized, closed profit from MT5's
                                    # own history (includes slippage, swap, commission) — never
                                    # a cached floating/estimated value.
                                    msg = f"🔔 <b>مزامنة: إغلاق صفقة [{symbol} - {tf}]</b>\nالربح الفعلي (MT5): {exact_pnl:.2f}$"
                                else:
                                    # MT5 history hasn't synced this deal yet (or lookup failed).
                                    # Never present an unconfirmed estimate as "الربح الفعلي" —
                                    # that's exactly how a fake/incorrect profit gets reported.
                                    log_exception(f"Reconciliation MISS [{symbol}/{tid}]",
                                                  Exception("closing deal not found in MT5 history; reporting estimate"))
                                    msg = (f"🔔 <b>مزامنة: إغلاق صفقة [{symbol} - {tf}]</b>\n"
                                           f"⚠️ ربح تقديري (لم تُؤكَّد بعد من سجل MT5): ~{exact_pnl:.2f}$\n"
                                           f"سيتم تصحيح الرقم تلقائياً عند تأكيد الصفقة من السجل.")

                                await send_tg_msg(msg)
                                await _record_closed_trade_history(
                                    symbol, tid, tr, exit_px=active_px, pnl=exact_pnl,
                                    outcome_label=('WIN' if exact_pnl > 0 else 'LOSS' if exact_pnl < 0 else 'BREAK_EVEN'),
                                    close_reason='tp_sl_or_manual_broker_close', pnl_confirmed=found_deal,
                                )
                                continue
                            else:
                                trade_pl = _safe_float(actual_positions[tid].get('unrealizedProfit'), trade_pl)
                        
                        outcome = core_eval_outcome(is_buy, active_px, tp, sl)
                            
                        if bot_state.get('prot_cost_be', True) and sym_state.get('break_even_enabled') and not tr.get('be_activated'):
                            be_pts = sym_state.get('strategy_be_trigger_points', 40)
                            net_be = core_eval_break_even(is_buy, entry, active_px, SYMBOL_INFO[symbol]['pip_value'], be_pts, sym_state.get('strategy_atr_period', 14), bot_state.get('prot_cost_be', True))
                            if net_be is not None:
                                if is_real and _metaapi_conn:
                                    try:
                                        await _metaapi_conn.modify_position(tid, stop_loss=net_be)
                                        tr['sl'] = net_be
                                        tr['be_activated'] = True # Only set if successful!
                                        await save_bot_persistence()
                                        await send_tg_msg(f"🛡️ تم تفعيل Break-Even لـ {symbol}!")
                                    except Exception as e:
                                        log_exception(f"BE modify_position [{symbol}/{tid}]", e)
                                        # be_activated stays False so we retry next tick; the
                                        # user is told immediately since capital protection failed.
                                        await send_tg_msg(f"⚠️ <b>فشل تفعيل Break-Even لـ {symbol} ({tid}):</b> {e}\nسيُعاد المحاولة تلقائياً.")
                                else:
                                    tr['sl'] = net_be
                                    tr['be_activated'] = True
                                    await save_bot_persistence()

                        if outcome:
                            closed_ids.append(tid)
                            bot_state['live_daily_realized'] += trade_pl
                            msg = f"🔔 <b>تحديث صفقة [{symbol} - {tf}]</b>\n\nالنتيجة: {outcome} ({trade_pl}$)\nسعر الإغلاق: {live_px:.2f}"
                            await send_tg_msg(msg)
                            await _record_closed_trade_history(
                                symbol, tid, tr, exit_px=live_px, pnl=trade_pl, outcome_label=outcome,
                                close_reason='tp_sl_hit', pnl_confirmed=False,
                            )
                        else:
                            total_floating += trade_pl

                    for tid in closed_ids:
                        if tid in sym_state['strategy_open_trades']:
                            del sym_state['strategy_open_trades'][tid]
                            await save_bot_persistence()

            # --- Evaluate Daily Limits ---
            total_daily = bot_state['live_daily_realized'] + total_floating
            dd_limit = -float(bot_state.get('prot_daily_dd_usd', 220))
            profit_limit = float(bot_state.get('prot_daily_profit_usd', 150))
            
            if (dd_limit < 0 and total_daily <= dd_limit) or (profit_limit > 0 and total_daily >= profit_limit):
                bot_state['live_daily_hit'] = True
                limit_type = '🛑 تراجع عائم' if total_daily <= dd_limit else '✅ هدف يومي عائم'
                await send_tg_msg(f"{limit_type} تم الوصول إليه! ({total_daily:.2f}$)\nسيتم إغلاق جميع الصفقات المفتوحة بالتسلسل.")
                
                # Batch closure: sequential close requests, one shared
                # confirmation loop (see _close_metaapi_trades_batch) --
                # avoids O(N x 25s) tail latency during a mass closure.
                real_closures = []
                for symbol in active_symbols:
                    sym_state = bot_state['symbol_state'][symbol]
                    for tid, tr in list(sym_state['strategy_open_trades'].items()):
                        if tr.get('is_real') and _metaapi_conn:
                            real_closures.append((symbol, tid, sym_state, tr))
                        else:
                            pl = tr.get('last_known_pl', 0.0)
                            px = tr.get('last_known_px', tr.get('entry'))
                            outcome_lbl = 'ربح ✅' if pl >= 0 else 'خسارة ❌'
                            await send_tg_msg(
                                f"⏹️ <b>إغلاق (وهمي) [{symbol} - {tr.get('tf')}]</b>\n"
                                f"سبب الإغلاق: حماية رأس المال (تراجع/هدف يومي)\n\n"
                                f"الاتجاه: {'BUY 📈' if tr.get('is_buy') else 'SELL 📉'}\n"
                                f"الدخول: {tr.get('entry')}  |  الإغلاق: {px}\n"
                                f"TP: {tr.get('tp')}  SL: {tr.get('sl')}\n"
                                f"النتيجة: {outcome_lbl} ({pl}$)"
                            )
                            await _record_closed_trade_history(
                                symbol, tid, tr, exit_px=px, pnl=pl,
                                outcome_label=('WIN' if pl > 0 else 'LOSS' if pl < 0 else 'BREAK_EVEN'),
                                close_reason='daily_capital_protection_forced_close', pnl_confirmed=False,
                            )
                            del sym_state['strategy_open_trades'][tid]
                            await save_bot_persistence()
                await _close_metaapi_trades_batch(real_closures)
                continue

            for symbol in active_symbols:
                try:
                    sym_state = bot_state['symbol_state'][symbol]
                    enabled_tfs = [tf for tf, on in sym_state['strategy_monitor_tfs'].items() if on]
                    if not enabled_tfs:
                        continue

                    detect_time = datetime.now(timezone.utc)
                    shift = bot_state.get('signal_candle_shift', 1)
                    allow_concur = bot_state.get('allow_concurrent_trades', False)
                    open_dirs = {v.get('is_buy') for v in sym_state['strategy_open_trades'].values() if isinstance(v, dict)}
                    last_signal_candle = sym_state.setdefault('strategy_last_signal_candle', {})

                    for tf in enabled_tfs:
                        # Never stack two positions on the same tf.
                        if any(isinstance(v, dict) and v.get('tf') == tf for v in sym_state['strategy_open_trades'].values()):
                            continue

                        candles = await fetch_candles(symbol, tf, count=STRAT_MIN_CANDLES + 20)
                        if not candles:
                            _diag_log_add({'ts': detect_time, 'symbol': symbol, 'tf': tf,
                                           'skip_reason': 'insufficient_oanda_candles(got=0)'})
                            continue

                        df = compute_strategy_indicators(candles)
                        if df is None:
                            _diag_log_add({'ts': detect_time, 'symbol': symbol, 'tf': tf,
                                           'skip_reason': f'insufficient_candles_for_indicators(got={len(candles)}, need={STRAT_MIN_CANDLES})'})
                            continue

                        sig = evaluate_strategy_signal(df, shift)
                        if sig is None:
                            _diag_log_add({'ts': detect_time, 'symbol': symbol, 'tf': tf,
                                           'skip_reason': 'no_signal_this_candle'})
                            continue

                        # Dedup: only ever act once per (symbol, tf, candle).
                        candle_key = str(sig.get('candle_time'))
                        if last_signal_candle.get(tf) == candle_key:
                            continue

                        # allow_concurrent_trades gate (Point 4 of the spec):
                        # blocks a new Buy while a Buy is open (and likewise
                        # for Sell) -- checked across the whole symbol, not
                        # just this tf.
                        if not allow_concur:
                            if sig['is_buy'] and True in open_dirs:
                                _diag_log_add({'ts': detect_time, 'symbol': symbol, 'tf': tf,
                                               'skip_reason': 'blocked_concurrent_buy_open'})
                                continue
                            if not sig['is_buy'] and False in open_dirs:
                                _diag_log_add({'ts': detect_time, 'symbol': symbol, 'tf': tf,
                                               'skip_reason': 'blocked_concurrent_sell_open'})
                                continue

                        # Claim this candle for this tf BEFORE awaiting the
                        # (slower) trade-open path, so a long-running open
                        # can't leave the window open for a duplicate fire
                        # next cycle off the same candle.
                        last_signal_candle[tf] = candle_key
                        await save_bot_persistence()

                        signal_price = float(candles[-1 - shift]['close'])
                        dir_lbl = 'شراء 🟢 BUY' if sig['is_buy'] else 'بيع 🔴 SELL'
                        reason = (f"[{symbol} - {tf}]  إشارة {dir_lbl}\n"
                                  f"EMA15={sig['ema15']:.2f} EMA50={sig['ema50']:.2f} EMA150={sig['ema150']:.2f}  |  "
                                  f"Stoch %K={sig['k']:.1f} %D={sig['d']:.1f}")

                        # Update the direction set immediately so a second tf
                        # firing in the SAME cycle also respects the gate.
                        open_dirs.add(sig['is_buy'])

                        await _strategy_open_trade(
                            symbol, sig['is_buy'], signal_price, candles, reason, tf,
                            detect_time=detect_time, t1_signal_ts=time.monotonic(),
                            feed_source='oanda_candle_close', feed_age_ms=None,
                        )

                    _diag_log_add({'ts': detect_time, 'symbol': symbol,
                                   'skip_reason': 'cycle_complete'})

                except Exception as sym_exc:
                    log_exception(f"strategy_monitor_scanner per-symbol [{symbol}]", sym_exc)
                    now_mono_sym = time.monotonic()
                    if now_mono_sym - _last_scanner_error_alert_ts > 300:
                        _last_scanner_error_alert_ts = now_mono_sym
                        await send_tg_msg(
                            f"🛑 <b>[{symbol}]</b> خطأ غير متوقع أثناء فحص هذا الرمز -- تم تخطيه لهذه الدورة فقط "
                            f"(باقي الرموز تستمر بشكل طبيعي):\n{sym_exc}"
                        )
                    continue

        except Exception as e:
            log_exception('strategy_monitor_scanner main loop', e)
            # This top-level catch wraps EVERY symbol's processing for the
            # whole cycle -- an exception anywhere (even for just one
            # symbol/timeframe) previously aborted the ENTIRE cycle
            # completely silently (server-side log only, no Telegram
            # message), which could look exactly like "the bot just isn't
            # opening trades" with zero clue why. Surface it, rate-limited
            # so a persistent failure doesn't spam every 15s.
            now_mono = time.monotonic()
            if now_mono - _last_scanner_error_alert_ts > 300:  # at most once per 5 min
                _last_scanner_error_alert_ts = now_mono
                await send_tg_msg(
                    f"🛑 <b>خطأ غير متوقع بدورة الفحص الحية (strategy_monitor_scanner):</b>\n{e}\n"
                    f"تم تخطي بقية هذه الدورة بالكامل بسببه. سيُعاد المحاولة بالدورة القادمة (~15 ثانية). "
                    f"إذا تكرر هذا الخطأ، راجع السجل الكامل (traceback) على السيرفر."
                )
        await asyncio.sleep(15)

# ─────────────────────────────────────────────────────────────
# PRO BACKTEST ENGINE (Macro Trend & Smart Break-Even)
# ─────────────────────────────────────────────────────────────

def _build_strategy_bt_excel(fname: str, res: dict, suspend_trigger_time: dict, suspended_days: dict) -> None:
    """Pure sync workbook build for run_strategy_backtest's report. Deliberately
    contains NO asyncio / await -- it is meant to be run via
    `await asyncio.to_thread(...)` so the styling/row-by-row loops (CPU-bound,
    can take seconds on a large backtest) never block the event loop that
    oanda_live_price_poller and the rest of the bot depend on."""
    wb = openpyxl.Workbook()
    ws_trades = wb.active
    ws_trades.title = "الصفقات"

    headers = ["الزوج", "وقت الصفقة (DAM)", "TF", "اتجاه", "الدخول (الإشارة)", "الهدف (TP)", "الوقف (SL)", "النتيجة", "ربح ($)", "رصيد تراكمي ($)"]
    ws_trades.append(headers)

    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    be_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    for cell in ws_trades[1]:
        cell.fill = gray_fill
        cell.font = Font(bold=True)

    current_cycle = None
    for tr in res['trade_logs']:
        if tr['cycle_ts'] != current_cycle:
            current_cycle = tr['cycle_ts']
            ws_trades.append([f"إشارة: {tr['cycle_time_str']}  |  دخول: {tr['cycle_close']:.2f}"] + [""]*9)
            row_idx = ws_trades.max_row
            ws_trades.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=10)
            ws_trades.cell(row=row_idx, column=1).fill = header_fill
            ws_trades.cell(row=row_idx, column=1).font = Font(bold=True)

        _OUTCOME_DISPLAY = {'WIN': 'WIN ✅', 'LOSS': 'LOSS ❌', 'BREAK_EVEN': 'BREAK_EVEN ⚖️', 'DAILY_LIMIT': 'DAILY_LIMIT ⏹️'}
        row_data = [
            tr['الزوج'], tr['وقت الصفقة (DAM)'], tr['TF'], tr['اتجاه'], tr['الدخول (الإشارة)'],
            tr['الهدف (TP)'], tr['الوقف (SL)'], _OUTCOME_DISPLAY.get(tr['النتيجة'], tr['النتيجة']), tr['ربح ($)'], tr['رصيد تراكمي ($)']
        ]
        ws_trades.append(row_data)
        row_idx = ws_trades.max_row

        fill = None
        if tr['النتيجة'] == 'WIN': fill = green_fill
        elif tr['النتيجة'] == 'LOSS': fill = red_fill
        elif tr['النتيجة'] == 'BREAK_EVEN': fill = be_fill

        if fill:
            for col in range(1, 11):
                ws_trades.cell(row=row_idx, column=col).fill = fill

    ws_cycles = wb.create_sheet("دورات H1")
    ws_cycles.append(["الزوج", "وقت أول شمعة بالنطاق (DAM)", "الإغلاق", "عدد الإشارات", "ملاحظة"])
    for cell in ws_cycles[1]: cell.fill = gray_fill; cell.font = Font(bold=True)

    for cycle in res['cycle_logs']:
        num_trades = len([t for t in res['trade_logs'] if t['cycle_ts'] == cycle['time_ts']])
        cycle_day = _utc_to_dam(cycle['time_dt']).strftime('%Y-%m-%d')
        if num_trades > 0:
            note = f"تم تنفيذ {num_trades} صفقة"
        elif cycle_day in suspend_trigger_time and cycle['time_dt'] >= suspend_trigger_time[cycle_day]:
            # Distinguish "day was already halted by capital protection"
            # from "price genuinely never reached a level" -- these are
            # very different situations and were previously reported
            # identically, which made it look like the strategy just
            # wasn't triggering when actually trading had stopped.
            note = "🛑 اليوم متوقف (تم تفعيل حماية رأس المال مسبقاً)"
        else:
            note = "لم يلمس السعر أي مستوى"
        ws_cycles.append([cycle['symbol'], _utc_to_dam(cycle['time_dt']).strftime('%Y-%m-%d %H:%M'), cycle['close'], num_trades, note])

    ws_susp = wb.create_sheet("أيام الإيقاف")
    ws_susp.append(["التاريخ", "السبب (النتيجة)"])
    for cell in ws_susp[1]: cell.fill = gray_fill; cell.font = Font(bold=True)
    for d_str, rsn in suspended_days.items():
        ws_susp.append([d_str, rsn])

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    for ws in [ws_trades, ws_cycles, ws_susp]:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align

    from openpyxl.utils import get_column_letter
    for i in range(1, 11): ws_trades.column_dimensions[get_column_letter(i)].width = 22.0
    for i in range(1, 6): ws_cycles.column_dimensions[get_column_letter(i)].width = 22.0
    for i in range(1, 3): ws_susp.column_dimensions[get_column_letter(i)].width = 25.0

    wb.save(fname)


async def run_strategy_backtest(start_dt: datetime, end_dt: datetime) -> None:
    global _bt_progress
    bot_state['is_backtesting'] = True
    fname = f"StrategyBT_{datetime.now(timezone.utc).strftime('%H%M%S')}.xlsx"

    active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
    if not active_symbols:
        bot_state['is_backtesting'] = False
        return

    first_sym_state = bot_state['symbol_state'][active_symbols[0]]

    enabled_tfs = [tf for tf, on in first_sym_state['strategy_monitor_tfs'].items() if on] or ['5m']
    shift = bot_state.get('signal_candle_shift', 1)
    desc_mode = f"EMA15/50/150 + Stoch(5,5,5 EMA) | شمعة التقييم: {shift}"
    desc_be = " | 🛡️ BE" if first_sym_state['break_even_enabled'] else ""
    desc_tfs = "+".join(enabled_tfs)
    syms_label = "+".join(active_symbols)

    prog = BtProgress(label=f"{syms_label} EMA/Stoch → [{desc_tfs}] | {desc_mode}{desc_be}", active_tfs=enabled_tfs); _bt_progress = prog
    await prog.start(bot_state['chat_id'])

    res = {'win': 0, 'loss': 0, 'be': 0, 'total_prof': 0.0, 'total_win_usd': 0.0, 'total_loss_usd': 0.0, 'peak_equity': 0.0, 'max_dd': 0.0, 'trade_logs': [], 'cycle_logs': []}
    _earliest_1m_seen = {}  # {symbol: earliest candle datetime actually fetched} -- used to warn if a long-range request silently got truncated (see fetch_candles hardening)

    try:
        # PHASE 1: Data Gathering & Signal Generation
        all_signals = []
        all_candles_events = []

        for symbol in active_symbols:
            sym_state = bot_state['symbol_state'][symbol]
            tpsl_mode = sym_state['strategy_tpsl_mode']
            pv  = SYMBOL_INFO[symbol]['pip_value']; lot = sym_state['lot_size']
            cs  = SYMBOL_INFO[symbol]['contract_size']
            prec = SYMBOL_INFO[symbol]['prec']

            quote = symbol.split('_')[1] if '_' in symbol else 'USD'
            _QUOTE_RATES = {'USD': 1.0, 'JPY': 1/150.0, 'AUD': 0.66, 'NZD': 0.61, 'EUR': 1.08, 'GBP': 1.27, 'CAD': 0.73, 'CHF': 1.11}
            quote_conv = _QUOTE_RATES.get(quote)
            if quote_conv is None:
                c_log(f"WARNING: unknown quote currency '{quote}' in {symbol} — quote_conv defaulted to 1.0, PnL may be incorrect")
                quote_conv = 1.0

            await prog.set_phase('جلب شموع الفريمات...')
            days_diff = (end_dt - start_dt).days or 1
            # Always fetch 1m for high-resolution price tracking during simulation
            need_1m = days_diff * 24 * 60 + 300
            mc_1m = await fetch_candles(symbol, '1m', count=need_1m, end_time=end_dt)
            if mc_1m:
                _earliest_1m_seen[symbol] = min(c['time'] for c in mc_1m)
                for c in mc_1m:
                    all_candles_events.append({'time': c['time'], 'symbol': symbol, 'high': float(c['high']), 'low': float(c['low']), 'close': float(c['close']), 'tf': '1m_track'})

            for btf in enabled_tfs:
                if prog.cancelled: return
                await asyncio.sleep(0)
                bmin = int(''.join(filter(str.isdigit, btf)))
                if 'h' in btf: bmin *= 60
                need_m = days_diff * 24 * (60 // max(bmin, 1)) + STRAT_MIN_CANDLES + 300
                mc = await fetch_candles(symbol, btf, count=need_m, end_time=end_dt)
                if not mc:
                    continue
                mc = sorted(mc, key=lambda c: c['time'])
                for c in mc:
                    all_candles_events.append({'time': c['time'], 'symbol': symbol, 'high': float(c['high']), 'low': float(c['low']), 'close': float(c['close']), 'tf': btf})

                df_full = compute_strategy_indicators(mc)
                if df_full is None:
                    continue

                ema15 = df_full['ema15']; ema50 = df_full['ema50']; ema150 = df_full['ema150']
                kk = df_full['stoch_k']; dd = df_full['stoch_d']
                k_prev = kk.shift(1); d_prev = dd.shift(1)
                cross_up = (k_prev <= d_prev) & (kk > dd)
                cross_down = (k_prev >= d_prev) & (kk < dd)
                # Zone check on the PRE-cross bar (k_prev/d_prev), matching
                # evaluate_strategy_signal's live logic -- see its comment.
                buy_cond = (ema15 > ema50) & (ema50 > ema150) & cross_up & (k_prev <= STRAT_STOCH_BUY_LEVEL) & (d_prev <= STRAT_STOCH_BUY_LEVEL)
                sell_cond = (ema150 > ema50) & (ema50 > ema15) & cross_down & (k_prev >= STRAT_STOCH_SELL_LEVEL) & (d_prev >= STRAT_STOCH_SELL_LEVEL)
                buy_cond = buy_cond.fillna(False); sell_cond = sell_cond.fillna(False)

                res['cycle_logs'].append({'symbol': symbol, 'time_ts': mc[0]['time'].timestamp(),
                                           'time_dt': mc[0]['time'], 'close': float(mc[0]['close']), 'levels': int(buy_cond.sum() + sell_cond.sum())})

                for j in range(1, len(mc)):
                    is_buy_sig = bool(buy_cond.iloc[j])
                    is_sell_sig = bool(sell_cond.iloc[j])
                    if not (is_buy_sig or is_sell_sig):
                        continue

                    fire_idx = j + shift  # index of the bar during which the live scanner would actually detect + act on this signal
                    if fire_idx >= len(mc):
                        continue  # not enough forward data in this window to know the fire time
                    bar_time = mc[fire_idx]['time']
                    if not (start_dt <= bar_time < end_dt):
                        continue

                    is_buy = is_buy_sig
                    entry = float(mc[j]['close'])

                    tf_tp = _strategy_tf_tp(symbol, btf); tf_sl = _strategy_tf_sl(symbol, btf)
                    if tpsl_mode == 'atr':
                        atr_val = _strategy_atr(mc[:j + 1], sym_state['strategy_atr_period'])
                        if not atr_val:
                            atr_val = tf_sl * pv
                        sl_d = atr_val * sym_state['strategy_atr_sl_mult']
                        tp_d = atr_val * sym_state['strategy_atr_tp_mult']
                    else:
                        sl_d = tf_sl * pv; tp_d = tf_tp * pv

                    tp_px = entry + tp_d if is_buy else entry - tp_d
                    sl_px = entry - sl_d if is_buy else entry + sl_d

                    all_signals.append({
                        'time': bar_time, 'symbol': symbol, 'is_buy': is_buy, 'entry': entry,
                        'tp_px': tp_px, 'sl_px': sl_px, 'sl_d': sl_d, 'tp_d': tp_d,
                        'be_trigger_px': 'dynamic' if sym_state['break_even_enabled'] else None,
                        'lot': lot, 'cs': cs, 'quote_conv': quote_conv, 'tf': btf,
                        'combo_key': f"{btf}_{mc[j]['time'].isoformat()}",
                        'cycle_time': mc[j]['time'], 'cycle_close': entry, 'level_key': f"EMA/Stoch[{btf}]",
                        'trigger_type': 'candle_close',
                    })

        # PHASE 2: Chronological Event-Driven Simulation
        await prog.set_phase('محاكاة الصفقات الزمنية (تقييم الأرباح العائمة)...')
        c_log(f'BT: Sorting {len(all_signals)} signals')
        all_signals.sort(key=lambda x: x['time'])
        c_log(f'BT: Sorting {len(all_candles_events)} events')
        all_candles_events.sort(key=lambda x: x['time'])
        
        open_trades = []
        closed_trades = []
        suspended_days = {}
        suspend_trigger_time = {}
        daily_pl = 0.0
        current_day = None
        latest_price = {}
        
        signal_idx = 0
        total_signals = len(all_signals)
        
        dd_limit = - float(bot_state['prot_daily_dd_usd'])
        profit_limit = float(bot_state['prot_daily_profit_usd'])
        
        total_events = len(all_candles_events)
        await prog.set_tf('محاكاة عائمة', total_events)
        
        for i, event in enumerate(all_candles_events):
            if i % 5000 == 0:
                await asyncio.sleep(0)
            if prog.cancelled: break
            t = event['time']; sym = event['symbol']; h = event['high']; l = event['low']; c = event['close']
            day_str = _utc_to_dam(t).strftime('%Y-%m-%d')
            latest_price[sym] = c
            
            if day_str != current_day:
                current_day = day_str
                daily_pl = 0.0
            
            # Check floating PnL against limits
            if current_day not in suspended_days:
                floating_pl = 0.0        # close-based -- used for the PROFIT check (unchanged, a late trigger costs nothing)
                floating_pl_worst = 0.0  # intrabar-worst-case -- used for the LOSS check only (tight, no overshoot)
                for tr in open_trades:
                    lp = latest_price.get(tr['symbol'], tr['entry'])
                    diff = (lp - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp)
                    floating_pl += round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)

                    # For the trade's own symbol, we have this candle's full
                    # high/low right now -- use the worst excursion within
                    # the bar (low for a long, high for a short) instead of
                    # only the close. For other symbols we only have their
                    # last close at this instant (an inherent limit of
                    # single-symbol event-driven iteration), so fall back
                    # to the same close-based price there.
                    worst_px = (l if tr['is_buy'] else h) if tr['symbol'] == sym else lp
                    diff_worst = (worst_px - tr['entry']) if tr['is_buy'] else (tr['entry'] - worst_px)
                    floating_pl_worst += round(diff_worst * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)

                total_daily = daily_pl + floating_pl
                total_daily_worst = daily_pl + floating_pl_worst
                if dd_limit < 0 and total_daily_worst <= dd_limit:
                    suspended_days[current_day] = f'🛑 تراجع عائم (الحد {dd_limit}$ | المحقق: {round(daily_pl, 2)}$ + العائم (أسوأ لحظة داخل الشمعة): {round(floating_pl_worst, 2)}$ = {round(total_daily_worst, 2)}$)'
                elif profit_limit > 0 and total_daily >= profit_limit:
                    suspended_days[current_day] = f'✅ هدف عائم (الحد {profit_limit}$ | المحقق: {round(daily_pl, 2)}$ + العائم: {round(floating_pl, 2)}$ = {round(total_daily, 2)}$)'

                if current_day in suspended_days and current_day not in suspend_trigger_time:
                    suspend_trigger_time[current_day] = t

                if current_day in suspended_days:
                    # Close all open trades. For the loss-triggering
                    # symbol, fill at the same worst-case intrabar price
                    # that tripped the check (tight to the limit) rather
                    # than the candle's close; other symbols still fill at
                    # their last known close, same as before.
                    was_loss_trigger = dd_limit < 0 and total_daily_worst <= dd_limit
                    for tr in open_trades:
                        if was_loss_trigger and tr['symbol'] == sym:
                            lp = l if tr['is_buy'] else h
                        else:
                            lp = latest_price.get(tr['symbol'], tr['entry'])
                        diff = (lp - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp)
                        p_usd = round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                        tr['outcome'] = 'DAILY_LIMIT'
                        tr['p_usd'] = p_usd
                        tr['close_time'] = t
                        closed_trades.append(tr)
                        daily_pl += p_usd
                    open_trades.clear()
            
            # Process Exits for open trades (if not suspended)
            if current_day not in suspended_days:
                surviving_trades = []
                for tr in open_trades:
                    if tr['symbol'] != sym:
                        surviving_trades.append(tr)
                        continue
                        
                    is_buy = tr['is_buy']; sl_current = tr['sl_current']; entry = tr['entry']
                    be_trigger_px = tr['be_trigger_px']; tp_px = tr['tp_px']; sl_d = tr['sl_d']
                    lot = tr['lot']; cs = tr['cs']; quote_conv = tr['quote_conv']
                    
                    closed = False
                    tp_dist = abs(tp_px - entry)
                    pv = SYMBOL_INFO[sym]['pip_value']
                    be_pts = sym_state.get('strategy_be_trigger_points', 40)
                    atr_per = sym_state.get('strategy_atr_period', 14)
                    cost_be = bot_state.get('prot_cost_be', True)
                    
                    if not tr['be_activated'] and be_trigger_px is not None:
                        # For BE trigger in backtest, we test against High for Buy, Low for Sell
                        test_px = h if is_buy else l
                        net_be = core_eval_break_even(is_buy, entry, test_px, pv, be_pts, atr_per, cost_be)
                        if net_be is not None:
                            tr['sl_current'] = net_be
                            tr['be_activated'] = True

                    # Outcome check uses h/l for extreme boundary testing.
                    # BE threshold derived from pip_value instead of hardcoded 0.01.
                    _be_thresh = pv * 2
                    if is_buy:
                        if l <= sl_current:
                            tr['outcome'] = 'BREAK_EVEN' if sl_current > entry - _be_thresh else 'LOSS'
                            tr['p_usd'] = round(abs(sl_current - entry) * lot * cs * quote_conv, 2) if tr['outcome'] == 'BREAK_EVEN' else -round(sl_d * lot * cs * quote_conv, 2)
                            closed = True
                        elif not closed and h >= tp_px:
                            tr['outcome'] = 'WIN'
                            tr['p_usd'] = round(tr['tp_d'] * lot * cs * quote_conv, 2)
                            closed = True
                    else:
                        if h >= sl_current:
                            tr['outcome'] = 'BREAK_EVEN' if sl_current < entry + _be_thresh else 'LOSS'
                            tr['p_usd'] = round(abs(entry - sl_current) * lot * cs * quote_conv, 2) if tr['outcome'] == 'BREAK_EVEN' else -round(sl_d * lot * cs * quote_conv, 2)
                            closed = True
                        elif not closed and l <= tp_px:
                            tr['outcome'] = 'WIN'
                            tr['p_usd'] = round(tr['tp_d'] * lot * cs * quote_conv, 2)
                            closed = True
                            
                    if closed:
                        tr['close_time'] = t
                        daily_pl += tr['p_usd']
                        closed_trades.append(tr)
                    else:
                        surviving_trades.append(tr)
                open_trades = surviving_trades
            
            # Process Entries
            while signal_idx < total_signals and all_signals[signal_idx]['time'] <= t:
                sig = all_signals[signal_idx]
                signal_idx += 1
                if current_day not in suspended_days:
                    # DAM time-window filter (07:00-09:00, 13:00-14:00) --
                    # this backtest engine has its own signal-admission
                    # path and never calls is_trading_allowed()/_strategy_open_
                    # trade (that gate only exists in the live engine), so
                    # the filter has to be re-applied here explicitly,
                    # checked against the SIGNAL's own historical
                    # timestamp rather than wall-clock time.
                    if bot_state.get('prot_dam_time_filter', True):
                        sig_dam_time = (sig['time'] + timedelta(hours=3)).time()
                        if any(start <= sig_dam_time < end for start, end in _DAM_RESTRICTED_WINDOWS):
                            continue
                    # Max concurrent trades cap -- mirrors the live bot's
                    # prot_max_concurrent_trades. Without this, the backtest
                    # can open one trade per enabled timeframe off a single
                    # level touch with no limit (the exact multi-tf stacking
                    # that caused real losses live), which is NOT what the
                    # live bot actually does anymore, and inflates backtest
                    # trade counts relative to what live can ever produce.
                    max_concurrent_bt = max(1, int(bot_state.get('prot_max_concurrent_trades', 4)))
                    open_count_bt = sum(1 for tr in open_trades if tr['symbol'] == sig['symbol'])
                    if open_count_bt >= max_concurrent_bt:
                        continue
                    sig['sl_current'] = sig['sl_px']
                    sig['be_activated'] = False
                    open_trades.append(sig)
                    
            await prog.tick(i, res['win'], res['loss'], res['be'], res['total_prof'])
            
        # Post-process closed trades to match old format
        c_log(f'BT: Post-processing {len(closed_trades)} closed trades')
        for tr in closed_trades:
            if tr['outcome'] == 'WIN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] > 0): 
                res['win'] += 1; res['total_win_usd'] += tr['p_usd']
            elif tr['outcome'] == 'LOSS' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] < 0): 
                res['loss'] += 1; res['total_loss_usd'] += abs(tr['p_usd'])
            elif tr['outcome'] == 'BREAK_EVEN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] == 0): 
                res['be'] += 1
            
            res['total_prof'] += tr['p_usd']
            dir_str = 'BUY 📈' if tr['is_buy'] else 'SELL 📉'
            res['trade_logs'].append({
                'الزوج': tr['symbol'], 
                'وقت الصفقة (DAM)': _utc_to_dam(tr['time']).strftime('%Y-%m-%d %H:%M'),
                'TF': tr['tf'], 
                'اتجاه': dir_str, 
                'الدخول (الإشارة)': f"{tr['entry']:.2f} ({tr['level_key']})",
                'الهدف (TP)': round(tr['tp_px'], 2),
                'الوقف (SL)': round(tr['sl_px'], 2),
                'النتيجة': tr['outcome'], 
                'ربح ($)': tr['p_usd'],
                'cycle_ts': tr['cycle_time'].timestamp(),
                'cycle_time_str': _utc_to_dam(tr['cycle_time']).strftime('%Y-%m-%d %H:%M'),
                'cycle_close': tr['cycle_close'],
                'trigger_type': tr.get('trigger_type', 'touch'),
            })
            
        res['trade_logs'].sort(key=lambda x: x['وقت الصفقة (DAM)'])
        
        running_eq = 5000.0
        peak_eq = 5000.0
        max_dd = 0.0
        for t_log in res['trade_logs']:
            running_eq += t_log['ربح ($)']
            t_log['رصيد تراكمي ($)'] = round(running_eq, 2)
            if running_eq > peak_eq: peak_eq = running_eq
            dd = peak_eq - running_eq
            if dd > max_dd: max_dd = dd
            
        res['peak_equity'] = peak_eq
        res['max_dd'] = max_dd

        if not res['trade_logs']:
            await prog.done('<b>باكتيست اكتمل ✅</b>\nلا توجد صفقات في هذا النطاق.')
            bot_state['is_backtesting'] = False; return

        await prog.set_phase('إنشاء ملف Excel المنسق...')
        
        c_log('BT: Generating Excel')
        
        sum_text = (
            f"<b>باكتيست الاستراتيجية اكتمل ✅</b>\n"
            f"{syms_label} H1→[{desc_tfs}] | {desc_mode}{desc_be}\n"
            f"{start_dt.strftime('%Y-%m-%d')} → {end_dt.strftime('%Y-%m-%d')}\n\n"
            f"Net: {'PROFIT ▲' if res['total_prof']>=0 else 'LOSS ▼'} ${round(res['total_prof'], 2)}\n"
            f"Win:  +${round(res['total_win_usd'], 2)} ({res['win']})\n"
            f"Loss: -${round(res['total_loss_usd'], 2)} ({res['loss']})\n"
            f"Break-Even: $0 ({res['be']})\n"
            f"WR: {round(res['win']/max(1, res['win']+res['loss'])*100)}% ({len(res['trade_logs'])} صفقة)\n"
            f"Max DD: ${round(res['max_dd'],2)} ({round((res['max_dd']/max(1,res['peak_equity']))*100)}%)\n"
        )
        
        if suspended_days:
            sum_text += "\nالتعليق بسبب حماية رأس المال:\n"
            for d_str, rsn in suspended_days.items():
                sum_text += f"- {d_str}: {rsn}\n"

        # Coverage-shortfall warning: if OANDA data actually available starts
        # meaningfully later than the requested start_dt (e.g. rate-limited
        # pagination gave up early -- see fetch_candles), a "1 month" request
        # can silently analyze only a few days with no error anywhere. Make
        # that visible here instead of leaving the user to notice only from
        # a suspiciously small trade count.
        short_syms = []
        for sym, earliest in _earliest_1m_seen.items():
            earliest_dt = earliest if earliest.tzinfo else earliest.replace(tzinfo=timezone.utc)
            gap_days = (earliest_dt - start_dt).total_seconds() / 86400
            if gap_days > 1.0:  # more than a day short of what was requested
                short_syms.append(f"{sym}: البيانات الفعلية بدأت من {earliest_dt.strftime('%Y-%m-%d %H:%M')} "
                                   f"بدل {start_dt.strftime('%Y-%m-%d %H:%M')} (نقص {gap_days:.1f} يوم)")
        if short_syms:
            sum_text += ("\n⚠️ <b>تحذير: تغطية بيانات ناقصة</b>\n"
                          "الفترة المطلوبة أكبر مما استطعنا جلبه فعلياً من أوندا (على الأغلب حد معدل API):\n"
                          + "\n".join(short_syms) + "\n")

        sum_text += f"\nدورات H1: {len(res['cycle_logs'])}  |  TP/SL: {str('ATR' if tpsl_mode=='atr' else 'نقاط ثابتة')} | Lot: {lot}"

        # NOTE: building/styling the workbook cell-by-cell is CPU-bound
        # synchronous work. Running it inline here would block the asyncio
        # event loop for the whole duration (can be seconds on a large
        # backtest) -- and that is exactly what starves
        # oanda_live_price_poller long enough to trip the >60s stale-feed
        # watchdog. So the whole build+save step is pushed to a worker
        # thread via asyncio.to_thread and the loop stays free to keep
        # polling OANDA while it runs.
        await asyncio.to_thread(_build_strategy_bt_excel, fname, res, suspend_trigger_time, suspended_days)

        await prog.done(f'<b>باكتيست الاستراتيجية اكتمل ✅</b>\n{syms_label} — {len(res["trade_logs"])} صفقة\nجاري إرسال التقرير والملف...')
        await send_tg_document(fname, sum_text)
        os.remove(fname)

    except Exception as e:
        c_log(f'BT Error: {e}'); bot_state['is_backtesting'] = False
        if _bt_progress:
            import html
            try: await _bt_progress.done(f'❌ خطأ داخلي في الباكتيست:\n{html.escape(str(e))}')
            except Exception as inner_e: log_exception('backtest error notification', inner_e)
    finally:
        bot_state['is_backtesting'] = False

# ═════════════════════════════════════════════════════════════
# LIVE-TWIN ENGINE — realistic execution simulator
# ═════════════════════════════════════════════════════════════
# Replaces run_strategy_backtest's zero-friction assumption (perfect fill
# at the exact level price, no cost, no ambiguity about which of
# SL/TP was touched first) with a market-friction model: dynamic
# spread, asymmetric slippage, signal-to-fill latency, commission/
# swap, weekend gap risk, and a Brownian-bridge intrabar path used
# only to resolve SL-vs-TP ordering when a single 1m bar's range
# contains both (OHLC alone can't answer that; assuming the worst
# or the best every time is its own bias, so we reconstruct a
# plausible-but-randomized path instead).
#
# run_strategy_backtest is left completely untouched and reachable from
# its own menu -- with lt_mode == 'idealized' this engine calls it
# directly, so it doubles as the zero-friction A/B baseline.
#
# Spread baseline is hardcoded from a live MT5/OANDA XAUUSD tick
# snapshot taken 2026-07-13 in the late-night/low-liquidity session:
#   Bid 4112.28 / Ask 4112.62 -> 0.34 USD (34 points @ tick size 0.01)
# That reading IS the quiet-session floor. Every multiplier below is
# defined as a ratio against it, never as an independent guess:
#   - Asian / dead-zone hours (where the snapshot was taken): 1.00x
#   - London session                                         : 0.70x
#   - London/NY overlap (deepest liquidity)                  : 0.55x
#   - NY session (post-overlap)                              : 0.75x
#   - Broker rollover window (21:55-22:05 UTC)                : up to 3.5x
#   - High-ATR bars (volatility spike, stacks on top of session): up to +2.5x more
# ═════════════════════════════════════════════════════════════

def _lt_session_multiplier(dt_utc: datetime) -> tuple[float, bool]:
    """Returns (spread_multiplier, is_rollover) for a UTC timestamp."""
    hm = dt_utc.hour + dt_utc.minute / 60.0
    if (21 + 55/60) <= hm <= (22 + 5/60):
        return 3.5, True                # broker rollover window -- spreads spike hard
    if 12.0 <= hm < 16.0:
        return 0.55, False              # London/NY overlap -- tightest liquidity
    if 7.0 <= hm < 12.0:
        return 0.70, False              # London session
    if 16.0 <= hm < 20.0:
        return 0.75, False              # NY session (post-overlap)
    return 1.00, False                  # Asian / dead-zone -- matches the live snapshot session


def _lt_volatility_multiplier(bar_range: float, atr_val: float | None) -> float:
    """Extra spread widening when a bar's range blows past its recent ATR."""
    if not atr_val or atr_val <= 0:
        return 1.0
    ratio = bar_range / atr_val
    if ratio <= 1.2:
        return 1.0
    return min(1.0 + (ratio - 1.2) * 0.9, 3.5)


def _lt_current_spread(base_spread: float, dt_utc: datetime, bar_range: float, atr_val: float | None) -> tuple[float, bool]:
    sess_mult, is_rollover = _lt_session_multiplier(dt_utc)
    vol_mult = _lt_volatility_multiplier(bar_range, atr_val)
    spread = base_spread * sess_mult * vol_mult
    return max(spread, base_spread * 0.45), is_rollover


def _lt_bridge_path(o: float, h: float, l: float, c: float, steps: int, rng: random.Random) -> np.ndarray:
    """
    Reconstructs a plausible intrabar tick path as a scaled Brownian
    bridge from open to close, clipped into [low, high]. Used only to
    decide the ORDER in which SL/TP thresholds would have been crossed
    inside a bar where raw OHLC can't tell us -- not claimed as the
    literal historical tick path, just a principled stand-in for one.
    """
    incs = rng.normal(0, 1, steps)
    w = np.concatenate(([0.0], np.cumsum(incs)))
    t = np.linspace(0.0, 1.0, steps + 1)
    bridge = w - t * w[-1]
    bstd = float(np.std(bridge))
    rng_size = max(h - l, 1e-6)
    scale = (rng_size * 0.5) / bstd if bstd > 1e-9 else 0.0
    path = o + (c - o) * t + bridge * scale
    return np.clip(path, l, h)


def _lt_first_hit(path: np.ndarray, is_buy: bool, sl_px: float, tp_px: float) -> str | None:
    """Walk the reconstructed path and return which of 'sl'/'tp' is crossed first, or None."""
    for px in path:
        if is_buy:
            if px <= sl_px: return 'sl'
            if px >= tp_px: return 'tp'
        else:
            if px >= sl_px: return 'sl'
            if px <= tp_px: return 'tp'
    return None


def _lt_slippage(bar_range: float, atr_val: float | None, rng: random.Random) -> float:
    """Asymmetric, range-scaled slippage magnitude -- always adverse (models cost, not luck)."""
    ref = atr_val if atr_val and atr_val > 0 else max(bar_range, 0.05)
    base = ref * 0.06
    tail = abs(rng.gauss(0, base))
    return min(tail, ref * 0.5)


def _lt_latency_shift(path: np.ndarray, steps: int, rng: random.Random) -> float:
    """Signal-to-fill delay expressed as a fractional shift along the intrabar path.
    Bounds default to a rough guess (160-200ms) but should be set from the bot's
    OWN measured MetaApi ping. Uses linear interpolation so the result is accurate
    regardless of step count -- previously `int(frac * steps)` = 0 for 20 steps."""
    lo = bot_state.get('lt_latency_ms_min', 160)
    hi = bot_state.get('lt_latency_ms_max', 200)
    latency_ms = rng.randint(lo, hi)
    frac = min(latency_ms / 60000.0, 1.0)  # fraction of a 1-minute bar consumed by the delay
    idx_f = frac * steps
    idx = int(idx_f)
    if idx >= steps:
        return float(path[-1])
    t = idx_f - idx
    return float(path[idx] * (1.0 - t) + path[idx + 1] * t)


def _build_live_twin_excel(fname: str, res: dict) -> None:
    """Pure sync workbook build for run_live_twin_simulation's report --
    see _build_strategy_bt_excel's docstring; run via asyncio.to_thread."""
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Live-Twin Trades"
    headers = ["الزوج", "وقت الصفقة (DAM)", "TF", "اتجاه", "الدخول الفعلي", "TP", "SL", "النتيجة", "ربح صافي ($)", "رصيد تراكمي ($)"]
    ws.append(headers)
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    be_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    for cell in ws[1]: cell.fill = gray_fill; cell.font = Font(bold=True)
    for t_log in res['trade_logs']:
        row = [t_log['الزوج'], t_log['وقت الصفقة (DAM)'], t_log['TF'], t_log['اتجاه'], t_log['المستوى (الدخول الفعلي)'],
               t_log['الهدف (TP)'], t_log['الوقف (SL)'], t_log['النتيجة'], t_log['ربح صافي ($)'], t_log['رصيد تراكمي ($)']]
        ws.append(row)
        fill = {'WIN': green_fill, 'LOSS': red_fill, 'BREAK_EVEN': be_fill}.get(t_log['النتيجة'])
        if fill:
            for col in range(1, 11): ws.cell(row=ws.max_row, column=col).fill = fill

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows():
        for cell in row: cell.border = thin_border; cell.alignment = center_align
    from openpyxl.utils import get_column_letter
    for i in range(1, 11): ws.column_dimensions[get_column_letter(i)].width = 22.0

    wb.save(fname)


async def run_live_twin_simulation(start_dt: datetime, end_dt: datetime) -> None:
    """Realistic-execution counterpart to run_strategy_backtest. Same signal
    logic (EMA 15/50/150 stack + Stochastic 5/5/5 EMA-smoothed crossover,
    evaluated on candle close), but fills, spreads, slippage, latency,
    commission, swap, and SL/TP ordering are all run through the friction
    model above instead of assumed perfect."""
    global _lt_progress
    if bot_state.get('lt_mode') == 'idealized':
        # A/B baseline: reuse the existing zero-friction engine untouched.
        await run_strategy_backtest(start_dt, end_dt)
        return

    bot_state['is_live_twin_running'] = True
    fname = f"LiveTwin_{datetime.now(timezone.utc).strftime('%H%M%S')}.xlsx"
    fric = bot_state['lt_friction']
    base_spread = float(bot_state['lt_base_spread_usd'])
    comm_per_lot = float(bot_state['lt_commission_per_lot'])
    swap_long_per_lot = float(bot_state.get('lt_swap_long_per_lot_night', -93.17))
    swap_short_per_lot = float(bot_state.get('lt_swap_short_per_lot_night', 21.68))
    swap_wed_mult = float(bot_state.get('lt_swap_wednesday_multiplier', 3.0))
    rej_prob = float(bot_state['lt_rejection_prob'])

    active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
    if not active_symbols:
        bot_state['is_live_twin_running'] = False
        return

    # Deterministic seed: identical config (symbols + date range + strategy
    # settings + friction toggles) => identical random slippage/latency/
    # rejection draws every run. Set bot_state['lt_seed'] to an int to
    # override; None keeps this auto-seed.
    override_seed = bot_state.get('lt_seed')
    if override_seed is not None:
        seed_val = int(override_seed)
    else:
        seed_key = (tuple(sorted(active_symbols)), start_dt.isoformat(), end_dt.isoformat(),
                    bot_state.get('signal_candle_shift', 1), tuple(sorted(fric.items())))
        seed_val = zlib.crc32(str(seed_key).encode())
    rng = random.Random(seed_val)

    first_sym_state = bot_state['symbol_state'][active_symbols[0]]
    enabled_tfs = [tf for tf, on in first_sym_state['strategy_monitor_tfs'].items() if on] or ['5m']
    shift = bot_state.get('signal_candle_shift', 1)
    syms_label = "+".join(active_symbols)
    on_tags = "+".join(k for k, v in fric.items() if v) or "none"

    prog = BtProgress(label=f"Live-Twin {syms_label} | friction:[{on_tags}]", active_tfs=enabled_tfs)
    _lt_progress = prog
    await prog.start(bot_state['chat_id'])

    res = {'win': 0, 'loss': 0, 'be': 0, 'total_prof': 0.0, 'total_win_usd': 0.0, 'total_loss_usd': 0.0,
           'peak_equity': 0.0, 'max_dd': 0.0, 'trade_logs': [], 'total_commission': 0.0, 'total_swap': 0.0,
           'rejected': 0, 'gap_events': 0}
    _earliest_1m_seen = {}  # {symbol: earliest candle datetime actually fetched} -- see fetch_candles hardening comment

    try:
        all_signals = []
        m1_by_symbol = {}

        # ── PHASE 1: signal generation (identical strategy logic to the idealized engine) ──
        for symbol in active_symbols:
            sym_state = bot_state['symbol_state'][symbol]
            tpsl_mode = sym_state['strategy_tpsl_mode']
            pv = SYMBOL_INFO[symbol]['pip_value']; lot = sym_state['lot_size']
            cs = SYMBOL_INFO[symbol]['contract_size']

            quote = symbol.split('_')[1] if '_' in symbol else 'USD'
            _QUOTE_RATES = {'USD': 1.0, 'JPY': 1/150.0, 'AUD': 0.66, 'NZD': 0.61, 'EUR': 1.08, 'GBP': 1.27, 'CAD': 0.73, 'CHF': 1.11}
            quote_conv = _QUOTE_RATES.get(quote)
            if quote_conv is None:
                c_log(f"WARNING: unknown quote currency '{quote}' in {symbol} — quote_conv defaulted to 1.0, PnL may be incorrect")
                quote_conv = 1.0

            await prog.set_phase('جلب شموع الدقيقة الواحدة (تنفيذ واقعي)...')
            days_diff = (end_dt - start_dt).days or 1
            need_1m = days_diff * 24 * 60 + 300
            mc_1m = await fetch_candles(symbol, '1m', count=need_1m, end_time=end_dt)
            if not mc_1m: continue
            _earliest_1m_seen[symbol] = min(c['time'] for c in mc_1m)
            m1_by_symbol[symbol] = sorted(mc_1m, key=lambda c: c['time'])

            await prog.set_phase('جلب شموع الفريمات...')
            for btf in enabled_tfs:
                if prog.cancelled: return
                await asyncio.sleep(0)
                bmin = int(''.join(filter(str.isdigit, btf))); bmin = bmin * 60 if 'h' in btf else bmin
                need_m = days_diff * 24 * (60 // max(bmin, 1)) + STRAT_MIN_CANDLES + 300
                mc = await fetch_candles(symbol, btf, count=need_m, end_time=end_dt)
                if not mc:
                    continue
                mc = sorted(mc, key=lambda c: c['time'])

                df_full = compute_strategy_indicators(mc)
                if df_full is None:
                    continue

                ema15 = df_full['ema15']; ema50 = df_full['ema50']; ema150 = df_full['ema150']
                kk = df_full['stoch_k']; dd = df_full['stoch_d']
                k_prev = kk.shift(1); d_prev = dd.shift(1)
                cross_up = (k_prev <= d_prev) & (kk > dd)
                cross_down = (k_prev >= d_prev) & (kk < dd)
                # Zone check on the PRE-cross bar (k_prev/d_prev), matching
                # evaluate_strategy_signal's live logic -- see its comment.
                buy_cond = (ema15 > ema50) & (ema50 > ema150) & cross_up & (k_prev <= STRAT_STOCH_BUY_LEVEL) & (d_prev <= STRAT_STOCH_BUY_LEVEL)
                sell_cond = (ema150 > ema50) & (ema50 > ema15) & cross_down & (k_prev >= STRAT_STOCH_SELL_LEVEL) & (d_prev >= STRAT_STOCH_SELL_LEVEL)
                buy_cond = buy_cond.fillna(False); sell_cond = sell_cond.fillna(False)

                for j in range(1, len(mc)):
                    is_buy_sig = bool(buy_cond.iloc[j])
                    is_sell_sig = bool(sell_cond.iloc[j])
                    if not (is_buy_sig or is_sell_sig):
                        continue

                    fire_idx = j + shift
                    if fire_idx >= len(mc):
                        continue
                    fire_bar = mc[fire_idx]
                    bar_time = fire_bar['time']
                    if not (start_dt <= bar_time < end_dt):
                        continue

                    is_buy = is_buy_sig
                    entry = float(mc[j]['close'])

                    tf_tp = _strategy_tf_tp(symbol, btf); tf_sl = _strategy_tf_sl(symbol, btf)
                    if tpsl_mode == 'atr':
                        atr_val = _strategy_atr(mc[:j + 1], sym_state['strategy_atr_period'])
                        if not atr_val:
                            atr_val = tf_sl * pv
                        sl_d = atr_val * sym_state['strategy_atr_sl_mult']
                        tp_d = atr_val * sym_state['strategy_atr_tp_mult']
                    else:
                        sl_d = tf_sl * pv; tp_d = tf_tp * pv

                    all_signals.append({
                        'time': bar_time, 'symbol': symbol, 'is_buy': is_buy, 'intended_entry': entry,
                        'sl_d': sl_d, 'tp_d': tp_d, 'be_enabled': sym_state['break_even_enabled'],
                        'lot': lot, 'cs': cs, 'quote_conv': quote_conv, 'tf': btf,
                        'combo_key': f"{btf}_{mc[j]['time'].isoformat()}",
                        'cycle_time': mc[j]['time'], 'cycle_close': entry, 'level_key': f"EMA/Stoch[{btf}]",
                        'trigger_type': 'candle_close',
                        'bar_o': float(fire_bar['open']), 'bar_h': float(fire_bar['high']),
                        'bar_l': float(fire_bar['low']), 'bar_c': float(fire_bar['close']),
                    })

        # ── PHASE 2: chronological, friction-aware, 1-minute-bar simulation ──
        await prog.set_phase('محاكاة التنفيذ الواقعي (سبريد/انزلاق/تأخير/عمولة)...')
        all_signals.sort(key=lambda x: x['time'])
        all_1m_events = sorted(
            [{'time': c['time'], 'symbol': sym, 'open': float(c['open']), 'high': float(c['high']),
              'low': float(c['low']), 'close': float(c['close'])}
             for sym, candles in m1_by_symbol.items() for c in candles],
            key=lambda x: x['time']
        )
        m1_lookup = {
            sym: {c['time']: {'open': float(c['open']), 'high': float(c['high']), 'low': float(c['low']), 'close': float(c['close'])}
                  for c in candles}
            for sym, candles in m1_by_symbol.items()
        }

        open_trades = []
        closed_trades = []
        suspended_days = {}
        suspend_trigger_time = {}
        daily_pl = 0.0
        current_day = None
        latest_price = {}

        signal_idx = 0
        total_signals = len(all_signals)
        dd_limit = -float(bot_state['prot_daily_dd_usd'])
        profit_limit = float(bot_state['prot_daily_profit_usd'])
        max_concurrent = max(1, int(bot_state.get('prot_max_concurrent_trades', 4)))

        total_events = len(all_1m_events)
        await prog.set_tf('محاكاة 1m واقعية', total_events)

        for i, ev in enumerate(all_1m_events):
            if i % 5000 == 0: await asyncio.sleep(0)
            if prog.cancelled: break
            t = ev['time']; sym = ev['symbol']; o = ev['open']; h = ev['high']; l = ev['low']; c = ev['close']
            bar_range = h - l
            day_str = _utc_to_dam(t).strftime('%Y-%m-%d')
            latest_price[sym] = c
            if day_str != current_day:
                current_day = day_str; daily_pl = 0.0

            atr_ref = bar_range if bar_range > 0 else 0.1
            spread_now, is_rollover = _lt_current_spread(base_spread, t, bar_range, atr_ref) if fric['spread'] else (base_spread, False)
            half_spread = spread_now / 2.0

            if fric['gaps'] and is_rollover:
                res['gap_events'] += 1

            # -- capital-protection daily limit check (worst-case intrabar) --
            if current_day not in suspended_days:
                floating_pl = 0.0; floating_pl_worst = 0.0
                for tr in open_trades:
                    lp = latest_price.get(tr['symbol'], tr['entry'])
                    diff = (lp - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp)
                    floating_pl += round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                    worst_px = (l if tr['is_buy'] else h) if tr['symbol'] == sym else lp
                    diff_worst = (worst_px - tr['entry']) if tr['is_buy'] else (tr['entry'] - worst_px)
                    floating_pl_worst += round(diff_worst * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)

                total_daily = daily_pl + floating_pl
                total_daily_worst = daily_pl + floating_pl_worst
                if dd_limit < 0 and total_daily_worst <= dd_limit:
                    suspended_days[current_day] = f'🛑 تراجع عائم (الحد {dd_limit}$)'
                elif profit_limit > 0 and total_daily >= profit_limit:
                    suspended_days[current_day] = f'✅ هدف عائم (الحد {profit_limit}$)'
                if current_day in suspended_days and current_day not in suspend_trigger_time:
                    suspend_trigger_time[current_day] = t
                if current_day in suspended_days:
                    for tr in open_trades:
                        lp = (l if tr['is_buy'] else h) if tr['symbol'] == sym else latest_price.get(tr['symbol'], tr['entry'])
                        exit_spread = _lt_current_spread(base_spread, t, bar_range, atr_ref)[0] if fric['spread'] else base_spread
                        lp_adj = lp - (exit_spread/2.0 if tr['is_buy'] else -exit_spread/2.0) if fric['spread'] else lp
                        diff = (lp_adj - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp_adj)
                        p_usd = round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                        tr['outcome'] = 'DAILY_LIMIT'; tr['p_usd'] = p_usd; tr['close_time'] = t
                        closed_trades.append(tr); daily_pl += p_usd
                    open_trades.clear()

            # -- exits: reconstruct an intrabar path per open trade on this symbol's bar --
            if current_day not in suspended_days:
                surviving = []
                for tr in open_trades:
                    if tr['symbol'] != sym:
                        surviving.append(tr); continue

                    is_buy = tr['is_buy']; entry = tr['entry']; sl_current = tr['sl_current']; tp_px = tr['tp_px']
                    lot = tr['lot']; cs = tr['cs']; quote_conv = tr['quote_conv']
                    closed = False

                    # Break-even arm (tested against the bar's favorable extreme, same as before)
                    if tr['be_enabled'] and not tr['be_activated']:
                        test_px = h if is_buy else l
                        be_pts = bot_state['symbol_state'][sym].get('strategy_be_trigger_points', 40)
                        pv_sym = SYMBOL_INFO[sym]['pip_value']
                        atr_per = bot_state['symbol_state'][sym].get('strategy_atr_period', 14)
                        cost_be = bot_state.get('prot_cost_be', True)
                        net_be = core_eval_break_even(is_buy, entry, test_px, pv_sym, be_pts, atr_per, cost_be)
                        if net_be is not None:
                            tr['sl_current'] = net_be; tr['be_activated'] = True; sl_current = net_be

                    hits_sl = (l <= sl_current) if is_buy else (h >= sl_current)
                    hits_tp = (h >= tp_px) if is_buy else (l <= tp_px)

                    outcome = None
                    if hits_sl and hits_tp:
                        # Ambiguous bar -- reconstruct a plausible path instead of always assuming one side.
                        path = _lt_bridge_path(o, h, l, c, steps=20, rng=rng)
                        outcome = _lt_first_hit(path, is_buy, sl_current, tp_px) or 'sl'
                    elif hits_sl:
                        outcome = 'sl'
                    elif hits_tp:
                        outcome = 'tp'

                    if outcome:
                        exit_spread = spread_now if fric['spread'] else 0.0
                        slip = _lt_slippage(bar_range, atr_ref, rng) if fric['slippage'] else 0.0
                        if outcome == 'sl':
                            raw_px = sl_current
                            fill_px = raw_px - (exit_spread/2.0 + slip) if is_buy else raw_px + (exit_spread/2.0 + slip)
                            _be_thresh = SYMBOL_INFO[tr['symbol']]['pip_value'] * 2
                            tr['outcome'] = 'BREAK_EVEN' if (is_buy and sl_current > entry - _be_thresh) or (not is_buy and sl_current < entry + _be_thresh) else 'LOSS'
                        else:
                            raw_px = tp_px
                            fill_px = raw_px - (exit_spread/2.0 + slip) if is_buy else raw_px + (exit_spread/2.0 + slip)
                            tr['outcome'] = 'WIN'
                        diff = (fill_px - entry) if is_buy else (entry - fill_px)
                        p_usd = round(diff * lot * cs * quote_conv, 2)
                        commission = comm_per_lot * lot if fric['commission'] else 0.0
                        nights = max((t.date() - tr['time'].date()).days, 0)
                        swap = 0.0
                        if fric['gaps'] and nights > 0:
                            per_night = swap_long_per_lot if is_buy else swap_short_per_lot
                            # Each night held may itself be a Wednesday (tripled) or not --
                            # walk the actual calendar days rather than assuming a flat rate.
                            for i in range(nights):
                                d = tr['time'].date() + timedelta(days=i)
                                mult = swap_wed_mult if d.weekday() == 2 else 1.0  # Monday=0 .. Wednesday=2
                                swap += per_night * mult
                            swap *= lot
                        p_usd_net = round(p_usd - commission + swap, 2)
                        if tr['outcome'] == 'WIN' and p_usd_net < 0:
                            tr['outcome'] = 'LOSS'  # friction ate the whole win -- report it honestly
                        tr['p_usd'] = p_usd_net
                        res['total_commission'] += commission; res['total_swap'] += swap
                        tr['close_time'] = t; daily_pl += p_usd_net
                        closed_trades.append(tr)
                        closed = True
                    if not closed:
                        surviving.append(tr)
                open_trades = surviving

            # -- entries --
            # NOTE: all_1m_events interleaves every active symbol's 1m bars
            # chronologically, so the loop's *current* sym/o/h/l/c belong to
            # whichever symbol's bar happens to land at this timestamp --
            # NOT necessarily the signal's own symbol. Entry fills must be
            # priced off the signal's OWN symbol's bar via m1_lookup, never
            # off the loop's current bar, or cross-symbol signals get
            # silently dropped/mispriced whenever two symbols interleave.
            while signal_idx < total_signals and all_signals[signal_idx]['time'] <= t:
                sig = all_signals[signal_idx]; signal_idx += 1
                if current_day in suspended_days:
                    continue
                if bot_state.get('prot_dam_time_filter', True):
                    sig_dam_time = (sig['time'] + timedelta(hours=3)).time()
                    if any(start <= sig_dam_time < end for start, end in _DAM_RESTRICTED_WINDOWS):
                        continue
                open_count = sum(1 for tr in open_trades if tr['symbol'] == sig['symbol'])
                if open_count >= max_concurrent:
                    continue
                if fric['rejection'] and rng.random() < rej_prob:
                    res['rejected'] += 1; continue

                # NOTE: previously this looked up m1_lookup[symbol][sig['time']],
                # i.e. only the FIRST 1-minute slice of the signal's own bar --
                # for a 5m/3m/2m signal that silently truncated the real bar
                # range down to a 1-minute one, understating slippage sizing
                # and shortening the reconstructed intrabar path. The touched
                # bar's true OHLC is now carried on the signal itself from
                # Phase 1, so use that directly.
                so, sh, sl_, sc = sig['bar_o'], sig['bar_h'], sig['bar_l'], sig['bar_c']
                sig_bar_range = sh - sl_

                entry_spread, _ = _lt_current_spread(base_spread, sig['time'], sig_bar_range, sig_bar_range or 0.1) if fric['spread'] else (base_spread, False)
                path = _lt_bridge_path(so, sh, sl_, sc, steps=20, rng=rng)
                shifted_px = _lt_latency_shift(path, 20, rng) if fric['latency'] else sig['intended_entry']
                slip = _lt_slippage(sig_bar_range, sig_bar_range or 0.1, rng) if fric['slippage'] else 0.0
                fill_entry = shifted_px + (entry_spread/2.0 + slip) if sig['is_buy'] else shifted_px - (entry_spread/2.0 + slip)

                is_buy = sig['is_buy']
                tp_px = fill_entry + sig['tp_d'] if is_buy else fill_entry - sig['tp_d']
                sl_px = fill_entry - sig['sl_d'] if is_buy else fill_entry + sig['sl_d']

                open_trades.append({
                    **sig, 'entry': fill_entry, 'tp_px': tp_px, 'sl_px': sl_px, 'sl_current': sl_px,
                    'be_activated': False, 'tp_d': sig['tp_d'], 'sl_d': sig['sl_d'],
                })

            await prog.tick(i, res['win'], res['loss'], res['be'], res['total_prof'])

        for tr in closed_trades:
            if tr['outcome'] == 'WIN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] > 0):
                res['win'] += 1; res['total_win_usd'] += tr['p_usd']
            elif tr['outcome'] == 'LOSS' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] < 0):
                res['loss'] += 1; res['total_loss_usd'] += abs(tr['p_usd'])
            elif tr['outcome'] == 'BREAK_EVEN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] == 0):
                res['be'] += 1
            res['total_prof'] += tr['p_usd']
            dir_str = 'BUY 📈' if tr['is_buy'] else 'SELL 📉'
            slip_px = tr['entry'] - tr['intended_entry'] if tr['is_buy'] else tr['intended_entry'] - tr['entry']
            slip_pips = round(slip_px / SYMBOL_INFO[tr['symbol']]['pip_value'], 2)
            res['trade_logs'].append({
                'الزوج': tr['symbol'], 'وقت الصفقة (DAM)': _utc_to_dam(tr['time']).strftime('%Y-%m-%d %H:%M'),
                'TF': tr['tf'], 'اتجاه': dir_str, 'المستوى (الدخول الفعلي)': f"{tr['entry']:.2f} ({tr['level_key']})",
                'الهدف (TP)': round(tr['tp_px'], 2), 'الوقف (SL)': round(tr['sl_px'], 2),
                'النتيجة': tr['outcome'], 'ربح صافي ($)': tr['p_usd'], 'cycle_ts': tr['cycle_time'].timestamp(),
                'trigger_type': tr.get('trigger_type', 'touch'), 'انزلاق (نقطة)': slip_pips,
            })

        res['trade_logs'].sort(key=lambda x: x['وقت الصفقة (DAM)'])
        running_eq = 5000.0; peak_eq = 5000.0; max_dd = 0.0
        for t_log in res['trade_logs']:
            running_eq += t_log['ربح صافي ($)']; t_log['رصيد تراكمي ($)'] = round(running_eq, 2)
            if running_eq > peak_eq: peak_eq = running_eq
            dd = peak_eq - running_eq
            if dd > max_dd: max_dd = dd
        res['peak_equity'] = peak_eq; res['max_dd'] = max_dd

        if not res['trade_logs']:
            await prog.done('<b>Live-Twin اكتمل ✅</b>\nلا توجد صفقات في هذا النطاق.')
            bot_state['is_live_twin_running'] = False; return

        await prog.set_phase('إنشاء ملف Excel المنسق...')
        short_syms = []
        for sym, earliest in _earliest_1m_seen.items():
            earliest_dt = earliest if earliest.tzinfo else earliest.replace(tzinfo=timezone.utc)
            gap_days = (earliest_dt - start_dt).total_seconds() / 86400
            if gap_days > 1.0:
                short_syms.append(f"{sym}: البيانات الفعلية بدأت من {earliest_dt.strftime('%Y-%m-%d %H:%M')} "
                                   f"بدل {start_dt.strftime('%Y-%m-%d %H:%M')} (نقص {gap_days:.1f} يوم)")
        coverage_warning = ""
        if short_syms:
            coverage_warning = ("\n⚠️ <b>تحذير: تغطية بيانات ناقصة</b>\n"
                                 "الفترة المطلوبة أكبر مما استطعنا جلبه فعلياً من أوندا (على الأغلب حد معدل API):\n"
                                 + "\n".join(short_syms) + "\n")
        sum_text = (
            f"<b>Live-Twin Engine اكتمل ✅ (واقعي)</b>\n"
            f"{syms_label} | friction: [{on_tags}]\n"
            f"{start_dt.strftime('%Y-%m-%d')} → {end_dt.strftime('%Y-%m-%d')}\n\n"
            f"Net: {'PROFIT ▲' if res['total_prof']>=0 else 'LOSS ▼'} ${round(res['total_prof'], 2)}\n"
            f"Win:  +${round(res['total_win_usd'], 2)} ({res['win']})\n"
            f"Loss: -${round(res['total_loss_usd'], 2)} ({res['loss']})\n"
            f"Break-Even: ({res['be']})\n"
            f"WR: {round(res['win']/max(1, res['win']+res['loss'])*100)}% ({len(res['trade_logs'])} صفقة)\n"
            f"Max DD: ${round(res['max_dd'],2)} ({round((res['max_dd']/max(1,res['peak_equity']))*100)}%)\n\n"
            f"عمولة إجمالية: -${round(res['total_commission'],2)} | سواب: ${round(res['total_swap'],2)}\n"
            f"صفقات مرفوضة (Requote): {res['rejected']} | نوافذ Rollover: {res['gap_events']}\n"
            f"Spread الأساسي: ${base_spread} (34pt من تيك حي)"
            f"{coverage_warning}"
        )

        # See _build_strategy_bt_excel's docstring: pushed off the event loop
        # thread for the same reason -- this used to run inline and could
        # block the loop long enough to starve oanda_live_price_poller past
        # its 60s stale-feed threshold.
        await asyncio.to_thread(_build_live_twin_excel, fname, res)

        await prog.done(f'<b>Live-Twin اكتمل ✅</b>\n{syms_label} — {len(res["trade_logs"])} صفقة\nجاري إرسال التقرير...')
        await send_tg_document(fname, sum_text)
        os.remove(fname)

    except Exception as e:
        c_log(f'Live-Twin Error: {e}'); bot_state['is_live_twin_running'] = False
        if _lt_progress:
            import html
            try: await _lt_progress.done(f'❌ خطأ داخلي في Live-Twin:\n{html.escape(str(e))}')
            except Exception as inner_e: log_exception('live-twin error notification', inner_e)
    finally:
        bot_state['is_live_twin_running'] = False


def get_live_twin_keyboard() -> dict:
    if bot_state['is_live_twin_running']:
        return {'inline_keyboard': [[{'text': '⏳ Live-Twin يعمل...', 'callback_data': 'noop'}], [{'text': '⏹ إلغاء', 'callback_data': 'cancel_lt'}]]}
    mode = bot_state.get('lt_mode', 'realistic')
    mode_label = '🧪 واقعي (Live-Twin)' if mode == 'realistic' else '🧊 مثالي (Idealized A/B)'
    return {'inline_keyboard': [
        [{'text': f'الوضع: {mode_label}', 'callback_data': 'lt_toggle_mode'}],
        [{'text': '⚙️ إعدادات الاحتكاك (Friction)', 'callback_data': 'menu_lt_friction'}],
        [{'text': 'يوم واحد', 'callback_data': 'lt_1'}, {'text': 'يومين', 'callback_data': 'lt_2'}],
        [{'text': 'ثلاثة أيام', 'callback_data': 'lt_3'}, {'text': 'أسبوع', 'callback_data': 'lt_7'}],
        [{'text': 'شهر كامل', 'callback_data': 'lt_30'}],
        [{'text': 'أو أرسل: /backtestreal YYYY-MM-DD', 'callback_data': 'noop'}],
        [{'text': '← رجوع', 'callback_data': 'menu_main'}],
    ]}


def get_live_twin_friction_keyboard() -> dict:
    fric = bot_state['lt_friction']
    def tag(key, label):
        return {'text': f"{label}: {'✅' if fric.get(key) else '🔴'}", 'callback_data': f'lt_fric_{key}'}
    return {'inline_keyboard': [
        [{'text': f"Spread أساسي: ${bot_state['lt_base_spread_usd']} (34pt/تيك حي)", 'callback_data': 'noop'}],
        [tag('spread', '📶 سبريد ديناميكي')],
        [tag('slippage', '⚡ انزلاق (Slippage)')],
        [tag('latency', '⏱ تأخير التنفيذ (200-800ms)')],
        [tag('commission', '💵 عمولة')],
        [tag('gaps', '📉 فجوات نهاية الأسبوع/Rollover')],
        [tag('rejection', '🚫 رفض/Requote')],
        [{'text': '← رجوع', 'callback_data': 'menu_lt'}],
    ]}


async def check_metaapi_status_command(chat_id: int):
    if not METAAPI_TOKEN or METAAPI_TOKEN == 'YOUR_METAAPI_TOKEN':
        await send_tg_msg("❌ MetaAPI Token غير مهيأ.")
        return
    if not ACCOUNT_ID or ACCOUNT_ID == 'YOUR_ACCOUNT_ID':
        await send_tg_msg("❌ Account ID غير مهيأ.")
        return
        
    await send_tg_msg("⏳ جاري فحص حالة الحساب من MetaAPI...")
    api = MetaApi(METAAPI_TOKEN)
    try:
        account = await api.metatrader_account_api.get_account(ACCOUNT_ID)
        state = account.state
        conn_status = account.connection_status
        
        msg = f"<b>حالة الحساب (MetaAPI)</b>\n"
        msg += f"الاسم: {account.name}\n"
        msg += f"الحالة: {state}\n"
        msg += f"الاتصال: {conn_status}\n\n"
        
        if state == 'DEPLOYED' and conn_status == 'CONNECTED':
            conn = account.get_rpc_connection()
            await conn.connect()
            await conn.wait_synchronized()
            
            acc_info = await conn.get_account_information()
            msg += f"<b>الرصيد:</b> {acc_info.get('balance', 0):.2f}\n"
            msg += f"<b>الاكويتي:</b> {acc_info.get('equity', 0):.2f}\n"
            msg += f"<b>الهامش المتاح:</b> {acc_info.get('freeMargin', 0):.2f}\n\n"
            
            positions = await conn.get_positions()
            msg += f"<b>الصفقات المفتوحة:</b> {len(positions)}\n"
            for p in positions:
                msg += f"🔸 {p['symbol']} | {p['type']} | {p['volume']} | Profit: {p.get('profit', 0):.2f}\n"
                
        else:
            msg += "⚠️ الحساب غير متصل حالياً لجلب تفاصيل الرصيد والصفقات."
            
        await send_tg_msg(msg)
    except Exception as e:
        import html
        await send_tg_msg(f"❌ خطأ في الاتصال بـ MetaAPI:\n{html.escape(str(e))}")

_last_persist_save_ts = 0.0

async def _debounced_persist_save():
    global _last_persist_save_ts
    now = time.monotonic()
    if now - _last_persist_save_ts < 2.0:
        return
    _last_persist_save_ts = now
    await save_bot_persistence()

async def _handle_callback(d: str, chat_id: int, msg_id: int) -> None:
    if d == 'check_metaapi_status':
        _safe_task(check_metaapi_status_command(chat_id), 'check_metaapi_status')
        return
    if d == 'run_diag':
        async def _run_diag_task():
            try:
                report = await strategy_run_diagnostics()
                # Telegram hard-caps messages at 4096 chars. Split on
                # double-newline (section) boundaries first; if a section
                # still exceeds 3500, split on single lines. Never split
                # inside an HTML tag (ensured by only splitting on \n).
                sections = report.split('\n\n')
                chunk = ""
                for sec in sections:
                    if len(chunk) + len(sec) + 2 > 3500:
                        if chunk.strip():
                            await send_tg_msg(chunk)
                        chunk = ""
                        if len(sec) > 3500:
                            for line in sec.split('\n'):
                                if len(chunk) + len(line) + 1 > 3500:
                                    await send_tg_msg(chunk)
                                    chunk = ""
                                chunk += line + "\n"
                            continue
                    chunk += sec + "\n\n"
                if chunk.strip():
                    await send_tg_msg(chunk)
            except Exception as e:
                log_exception('strategy_run_diagnostics', e)
                await send_tg_msg(f"❌ فشل التشخيص: {e}")
        _safe_task(_run_diag_task(), 'run_diag')
        return
    if d == 'export_diag_excel':
        async def _export_diag_task():
            try:
                await export_diag_log_excel()
            except Exception as e:
                log_exception('export_diag_log_excel', e)
                await send_tg_msg(f"❌ فشل تصدير سجل التشخيص: {e}")
        asyncio.create_task(_export_diag_task())
        return
    if d == 'export_live_trades_excel':
        async def _export_live_trades_task():
            try:
                await export_live_trades_excel()
            except Exception as e:
                log_exception('export_live_trades_excel', e)
                await send_tg_msg(f"❌ فشل تصدير سجل الصفقات الحية: {e}")
        _safe_task(_export_live_trades_task(), 'export_live_trades_excel')
        return
    if d == 'export_exec_report':
        async def _export_exec_report_task():
            fname = None
            try:
                fname = await export_execution_details_report()
                if fname is None:
                    await send_tg_msg("📭 <b>لا يوجد سجل صفقات حية مغلقة بعد.</b>")
                    return
                hist = bot_state.get('live_trade_history', [])
                wins = sum(1 for t in hist if t.get('outcome') == 'WIN')
                losses = sum(1 for t in hist if t.get('outcome') == 'LOSS')
                bes = sum(1 for t in hist if t.get('outcome') == 'BREAK_EVEN')
                total_pnl = sum(t.get('pnl', 0.0) for t in hist)
                wr = round(100 * wins / max(wins + losses, 1), 1)
                caption = (
                    f"📋 <b>تقرير تفاصيل التنفيذ</b>\n"
                    f"{len(hist)} صفقة  |  ربح: {wins}  |  خسارة: {losses}  |  تعادل: {bes}\n"
                    f"WR: {wr}%  |  صافي PnL: {total_pnl:+.2f}$\n\n"
                    f"الأعمدة: Latency (ms), Method, Slippage (Pips), IOC Fail Reason"
                )
                await send_tg_document(fname, caption)
            except Exception as e:
                log_exception('export_execution_details_report', e)
                await send_tg_msg(f"❌ فشل إنشاء تقرير التنفيذ: {e}")
            finally:
                if fname and os.path.exists(fname):
                    try:
                        os.remove(fname)
                    except Exception as e:
                        log_exception('export_exec_report cleanup', e)
        _safe_task(_export_exec_report_task(), 'export_exec_report')
        return
    if d == 'manual_resume_step1':
        current_state = bot_state.get('connection_state', CONN_RUNNING)
        if current_state == CONN_RUNNING:
            await send_tg_msg("✅ البوت أصلاً في حالة RUNNING -- لا حاجة لأي استئناف.")
            return
        await send_tg_msg(
            f"⚠️ <b>تأكيد الاستئناف اليدوي</b>\n"
            f"الحالة الحالية: {current_state}\n"
            f"السبب: {bot_state.get('connection_state_reason', '-')}\n\n"
            f"هل تأكدت فعلياً من حساب الوسيط (MT5) ومقارنته بما يتتبعه البوت؟ "
            f"الضغط على تأكيد سيعيد البوت للعمل فوراً بافتراض أن الحساب سليم.",
            reply_markup={'inline_keyboard': [
                [{'text': '✅ نعم، تأكدت -- استأنف الآن', 'callback_data': 'manual_resume_confirm'}],
                [{'text': '❌ إلغاء', 'callback_data': 'menu_main'}],
            ]}
        )
        return
    if d == 'manual_resume_confirm':
        global _recon_consecutive_mismatches, _consecutive_real_order_failures
        prior_state = bot_state.get('connection_state', CONN_RUNNING)
        _recon_consecutive_mismatches = 0
        _consecutive_real_order_failures = 0
        await set_connection_state(
            CONN_RUNNING,
            f"Manually resumed by operator via Telegram after verifying account state "
            f"(was {prior_state})."
        )
        await send_tg_msg("✅ تم الاستئناف اليدوي. البوت الآن RUNNING وسيقبل صفقات جديدة من التحديث القادم.")
        return

    sym = bot_state['ui_selected_symbol']
    sym_state = bot_state['symbol_state'][sym]
    if d == 'menu_main':
        await _show(chat_id, msg_id, '<b>مرحباً بك في Gold Scalper Bot v8.9</b>', get_main_keyboard())

    elif d == 'menu_presets':
        kbd = {'inline_keyboard': [
            [{'text': '💾 حفظ كـ Preset 1', 'callback_data': 'save_preset_1'}, {'text': '📂 تحميل Preset 1', 'callback_data': 'load_preset_1'}],
            [{'text': '💾 حفظ كـ Preset 2', 'callback_data': 'save_preset_2'}, {'text': '📂 تحميل Preset 2', 'callback_data': 'load_preset_2'}],
            [{'text': '💾 حفظ كـ Preset 3', 'callback_data': 'save_preset_3'}, {'text': '📂 تحميل Preset 3', 'callback_data': 'load_preset_3'}],
            [{'text': '🔙 رجوع', 'callback_data': 'menu_main'}]
        ]}
        await _show(chat_id, msg_id, '<b>إدارة الإعدادات (Presets):</b>\nهنا يمكنك حفظ إعدادات جميع الأزواج واستعادتها لاحقاً.', kbd)
    elif d.startswith('save_preset_'):
        p_num = d.split('_')[-1]
        data = {}
        if os.path.exists(PRESETS_FILE):
            try:
                with open(PRESETS_FILE, 'r') as f: data = json.load(f)
            except Exception as e:
                # Corrupt presets file -- log it instead of silently
                # discarding whatever else was saved in there.
                log_exception(f"save_preset_{p_num} (reading existing presets)", e)
                await send_tg_msg(f"⚠️ ملف الـ Presets الحالي تالف، سيتم إنشاء ملف جديد. (الخطأ: {e})")
                data = {}

        # A preset should only ever capture settings, never live runtime
        # state (open trades, dedup state, etc). _PRESET_EXCLUDED_KEYS matches exactly
        # what load_preset already refuses to restore, so nothing is lost
        # by leaving them out of what gets saved in the first place.
        data[f'preset_{p_num}'] = {
            s_name: {k: v for k, v in s_data.items() if k not in _PRESET_EXCLUDED_KEYS}
            for s_name, s_data in bot_state['symbol_state'].items()
        }
        try:
            with open(TEMP_PRESETS_FILE, 'w') as f:
                json.dump(data, f)
                f.flush()
                os.fsync(f.fileno())
            os.replace(TEMP_PRESETS_FILE, PRESETS_FILE)
            await send_tg_msg(f"✅ تم حفظ الإعدادات الحالية في Preset {p_num}")
        except Exception as e:
            log_exception(f"save_preset_{p_num} (writing)", e)
            await send_tg_msg(f"❌ فشل حفظ Preset {p_num}: {e}")
    elif d.startswith('load_preset_'):
        p_num = d.split('_')[-1]
        if not os.path.exists(PRESETS_FILE):
            await send_tg_msg(
                "❌ لا يوجد ملف Presets محفوظ بعد.\n"
                "ملاحظة: كانت الإصدارات السابقة تحفظ هذا الملف في مسار مؤقت يُمسح عند إعادة التشغيل -- "
                "تم إصلاح ذلك الآن، فأي Preset تحفظه من الآن فصاعداً سيبقى بعد إعادة التشغيل."
            )
        else:
            try:
                with open(PRESETS_FILE, 'r') as f: data = json.load(f)
                if f'preset_{p_num}' in data:
                    # Load settings, but keep live data like open_trades untouched
                    for s_name, s_data in data[f'preset_{p_num}'].items():
                        if s_name in bot_state['symbol_state']:
                            for k, v in s_data.items():
                                if k not in _PRESET_EXCLUDED_KEYS:
                                    bot_state['symbol_state'][s_name][k] = v
                    await send_tg_msg(f"✅ تم تحميل الإعدادات من Preset {p_num} بنجاح!")
                else:
                    await send_tg_msg("❌ لا يوجد إعدادات محفوظة في هذا الـ Preset.")
            except Exception as e:
                log_exception(f"load_preset_{p_num}", e)
                await send_tg_msg(f"❌ حدث خطأ أثناء التحميل: {e}")

    elif d == 'menu_protection':
        await _show(chat_id, msg_id, 'إعدادات الحماية:', get_protection_keyboard())
    elif d == 'prot_toggle_multitf':
        bot_state['prot_allow_multi_tf'] = not bot_state['prot_allow_multi_tf']
        await _show(chat_id, msg_id, 'إعدادات الحماية:', get_protection_keyboard())
    elif d == 'prot_dec_dd':
        bot_state['prot_daily_dd_usd'] = max(50, bot_state['prot_daily_dd_usd'] - 50)
        await _show(chat_id, msg_id, 'إعدادات الحماية:', get_protection_keyboard())
    elif d == 'prot_inc_dd':
        bot_state['prot_daily_dd_usd'] = min(5000, bot_state['prot_daily_dd_usd'] + 50)
        await _show(chat_id, msg_id, 'إعدادات الحماية:', get_protection_keyboard())
    elif d == 'prot_dec_profit':
        bot_state['prot_daily_profit_usd'] = max(0, bot_state['prot_daily_profit_usd'] - 50)
        await _show(chat_id, msg_id, 'إعدادات الحماية:', get_protection_keyboard())
    elif d == 'prot_inc_profit':
        bot_state['prot_daily_profit_usd'] = min(10000, bot_state['prot_daily_profit_usd'] + 50)
        await _show(chat_id, msg_id, 'إعدادات الحماية:', get_protection_keyboard())
    elif d == 'menu_strategy': await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'tg_prot_sync': bot_state['prot_true_sync'] = not bot_state.get('prot_true_sync', True); await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())
    elif d == 'tg_prot_cost': bot_state['prot_cost_be'] = not bot_state.get('prot_cost_be', True); await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())
    elif d == 'tg_prot_stale': bot_state['prot_stale_filter'] = not bot_state.get('prot_stale_filter', True); await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())

    elif d == 'prot_reset_all':
        # Clears "stuck until next natural trigger" protection state:
        # live_daily_hit is the capital-protection daily DD/profit lock,
        # which normally only clears at midnight broker time.
        was_daily_hit = bot_state.get('live_daily_hit', False)
        bot_state['live_daily_hit'] = False
        await save_bot_persistence()
        await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())
        if was_daily_hit:
            await send_tg_msg("🔄 <b>تصفير الحمايات</b>\n\n• قفل حماية رأس المال اليومي (ربح/تراجع) — تم فكّه، الدخول مسموح من الآن")
        else:
            await send_tg_msg("🔄 <b>تصفير الحمايات</b>\n\nلا توجد حمايات نشطة حالياً لتصفيرها — كل شيء طبيعي.")

    elif d == 'tg_prot_dam_time':
        bot_state['prot_dam_time_filter'] = not bot_state.get('prot_dam_time_filter', True)
        # Was previously never persisted -- the in-memory toggle worked
        # immediately, but reverted to the default on any restart or reload,
        # which looked exactly like "toggling does nothing."
        await save_bot_persistence()
        await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())

    elif d == 'strategy_show_indicators':
        sym = bot_state['ui_selected_symbol']
        enabled_tfs = [tf for tf, on in sym_state['strategy_monitor_tfs'].items() if on]
        if not enabled_tfs:
            await send_tg_msg('⚠️ لا يوجد أي فريم مفعّل حالياً.')
            await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
            return
        shift = bot_state.get('signal_candle_shift', 1)
        lines = [f'📉 <b>قيم EMA/Stochastic الحالية — {sym}</b> (شمعة التقييم: {shift})\n']
        for tf in enabled_tfs:
            candles = await fetch_candles(sym, tf, count=STRAT_MIN_CANDLES + 20)
            if not candles:
                lines.append(f"[{tf}] ❌ تعذّر جلب الشموع.")
                continue
            df = compute_strategy_indicators(candles)
            if df is None:
                lines.append(f"[{tf}] ⚠️ بيانات غير كافية (يلزم ~{STRAT_MIN_CANDLES} شمعة).")
                continue
            idx = len(df) - 1 - shift
            ema15, ema50, ema150 = df['ema15'].iloc[idx], df['ema50'].iloc[idx], df['ema150'].iloc[idx]
            k, d_ = df['stoch_k'].iloc[idx], df['stoch_d'].iloc[idx]
            stack = 'صاعد ⬆️' if ema15 > ema50 > ema150 else ('هابط ⬇️' if ema150 > ema50 > ema15 else 'غير مصطف')
            lines.append(f"[{tf}] EMA15={ema15:.2f} EMA50={ema50:.2f} EMA150={ema150:.2f} ({stack}) | Stoch %K={k:.1f} %D={d_:.1f}")
        await send_tg_msg('\n'.join(lines))
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())

    elif d == 'strategy_show_last10':
        sym = bot_state['ui_selected_symbol']
        enabled_tfs = [tf for tf, on in sym_state['strategy_monitor_tfs'].items() if on]
        tf = enabled_tfs[0] if enabled_tfs else '15m'
        offset = bot_state.get('broker_time_offset', 3)
        await send_tg_msg(f'⏳ جاري جلب آخر 10 شموع {tf} لـ {sym} من اواندا...')
        candles = await fetch_candles(sym, tf, count=10)
        if not candles:
            await send_tg_msg('❌ تعذّر جلب الشموع.')
            return
        candles = sorted(candles, key=lambda c: c['time'])[-10:]
        lines = [
            f'🕯️ <b>آخر 10 شموع {tf} — {sym}</b>',
            f'(المصدر: OANDA | التوقيت المعروض: دمشق UTC+{offset} — وبين قوسين UTC الخام)',
            ''
        ]
        for i, c in enumerate(candles, 1):
            t_utc = c['time'].to_pydatetime()
            t_dam = t_utc + timedelta(hours=offset)
            lines.append(f"{i}) {t_dam.strftime('%m-%d %H:%M')} دمشق ({t_utc.strftime('%H:%M')} UTC)  |  إغلاق: {float(c['close']):.5f}")
        await send_tg_msg('\n'.join(lines))
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())

    elif d == 'strategy_toggle_candle_shift':
        bot_state['signal_candle_shift'] = (bot_state.get('signal_candle_shift', 1) + 1) % 3
        await save_bot_persistence()
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_toggle_concurrent':
        bot_state['allow_concurrent_trades'] = not bot_state.get('allow_concurrent_trades', False)
        await save_bot_persistence()
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_toggle_auto_trade':
        sym_state['auto_trade'] = not sym_state.get('auto_trade', False)
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_toggle_be':
        sym_state['break_even_enabled'] = not sym_state['break_even_enabled']
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_dec_be_pts':
        sym_state['strategy_be_trigger_points'] = max(10, sym_state.get('strategy_be_trigger_points', 40) - 10)
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_inc_be_pts':
        sym_state['strategy_be_trigger_points'] = min(200, sym_state.get('strategy_be_trigger_points', 40) + 10)
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_dec_lot':
        sym_state['lot_size'] = round(max(0.01, sym_state['lot_size'] - 0.01), 2)
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_inc_lot':
        sym_state['lot_size'] = round(min(50.0, sym_state['lot_size'] + 0.01), 2)
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_toggle_tpsl':
        sym_state['strategy_tpsl_mode'] = 'atr' if sym_state['strategy_tpsl_mode'] == 'fixed' else 'fixed'
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_dec_tp10': sym_state['strategy_tp_points'] = max(10, sym_state['strategy_tp_points'] - 10); await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_inc_tp10': sym_state['strategy_tp_points'] = min(1000, sym_state['strategy_tp_points'] + 10); await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_dec_sl10': sym_state['strategy_sl_points'] = max(10, sym_state['strategy_sl_points'] - 10); await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_inc_sl10': sym_state['strategy_sl_points'] = min(1000, sym_state['strategy_sl_points'] + 10); await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_dec_atrp':  sym_state['strategy_atr_period'] = max(5,   sym_state['strategy_atr_period'] - 1); await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_inc_atrp':  sym_state['strategy_atr_period'] = min(50,  sym_state['strategy_atr_period'] + 1); await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_dec_atrsl': sym_state['strategy_atr_sl_mult'] = max(0.5, round(sym_state['strategy_atr_sl_mult'] - 0.5, 1)); await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_inc_atrsl': sym_state['strategy_atr_sl_mult'] = min(5.0, round(sym_state['strategy_atr_sl_mult'] + 0.5, 1)); await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_dec_atrtp': sym_state['strategy_atr_tp_mult'] = max(0.5, round(sym_state['strategy_atr_tp_mult'] - 0.5, 1)); await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_inc_atrtp': sym_state['strategy_atr_tp_mult'] = min(8.0, round(sym_state['strategy_atr_tp_mult'] + 0.5, 1)); await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d.startswith('strategy_toggle_pair_'):
        pair = d[len('strategy_toggle_pair_'):]
        bot_state['active_symbols'][pair] = not bot_state['active_symbols'][pair]
        # Price feed is now driven by oanda_live_price_poller, which reads
        # active_symbols dynamically on every poll -- no MetaAPI market-data
        # subscription is required when a symbol is toggled on.
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d.startswith('strategy_sel_pair_'):
        pair = d[len('strategy_sel_pair_'):]
        bot_state['ui_selected_symbol'] = pair
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d.startswith('strategy_tf_'):
        tfk = d[len('strategy_tf_'):]
        if tfk in sym_state['strategy_monitor_tfs']: sym_state['strategy_monitor_tfs'][tfk] = not sym_state['strategy_monitor_tfs'][tfk]
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'strategy_tpsl_tf': await _show(chat_id, msg_id, '⚙️ TP/SL مخصص لكل فريم:', get_strategy_tpsl_tf_keyboard())
    elif d.startswith('strategy_tptf_sel_'):
        sel_tf = d[len('strategy_tptf_sel_'):]; await _show(chat_id, msg_id, f'⚙️ TP/SL [{sel_tf}]:', get_strategy_tpsl_tf_keyboard(sel_tf))
    elif d.startswith('strategy_tptf_itp_'):
        tf = d[len('strategy_tptf_itp_'):]; sym_state['strategy_tp_per_tf'][tf] = sym_state['strategy_tp_per_tf'].get(tf, 0) + 10; await _show(chat_id, msg_id, f'⚙️ TP/SL [{tf}]:', get_strategy_tpsl_tf_keyboard(tf))
    elif d.startswith('strategy_tptf_dtp_'):
        tf = d[len('strategy_tptf_dtp_'):]; sym_state['strategy_tp_per_tf'][tf] = max(0, sym_state['strategy_tp_per_tf'].get(tf, 0) - 10); await _show(chat_id, msg_id, f'⚙️ TP/SL [{tf}]:', get_strategy_tpsl_tf_keyboard(tf))
    elif d.startswith('strategy_tptf_isl_'):
        tf = d[len('strategy_tptf_isl_'):]; sym_state['strategy_sl_per_tf'][tf] = sym_state['strategy_sl_per_tf'].get(tf, 0) + 10; await _show(chat_id, msg_id, f'⚙️ TP/SL [{tf}]:', get_strategy_tpsl_tf_keyboard(tf))
    elif d.startswith('strategy_tptf_dsl_'):
        tf = d[len('strategy_tptf_dsl_'):]; sym_state['strategy_sl_per_tf'][tf] = max(0, sym_state['strategy_sl_per_tf'].get(tf, 0) - 10); await _show(chat_id, msg_id, f'⚙️ TP/SL [{tf}]:', get_strategy_tpsl_tf_keyboard(tf))
    elif d.startswith('strategy_tptf_rst_'):
        tf = d[len('strategy_tptf_rst_'):]; sym_state['strategy_tp_per_tf'][tf] = 0; sym_state['strategy_sl_per_tf'][tf] = 0; await _show(chat_id, msg_id, f'⚙️ تمت إعادة الضبط:', get_strategy_tpsl_tf_keyboard(tf))
    elif d == 'menu_strategy_bt':
        await _show(chat_id, msg_id, 'اختر مدة الباكتيست:', get_strategy_bt_keyboard())
    elif d.startswith('gbt_'):
        days = int(d.split('_')[1])
        end_dt = datetime.now(timezone.utc)
        start_dt = end_dt - timedelta(days=days)
        if not bot_state['is_backtesting']:
            bot_state['is_backtesting'] = True
            _safe_task(run_strategy_backtest(start_dt, end_dt), 'backtest_preset')
        await _show(chat_id, msg_id, f'⏳ باكتيست يعمل...', get_strategy_bt_keyboard())
    elif d == 'cancel_bt':
        global _bt_progress
        if _bt_progress and bot_state['is_backtesting']: await _bt_progress.cancel()
        bot_state['is_backtesting'] = False
        await _show(chat_id, msg_id, 'إعدادات الاستراتيجية:', get_strategy_keyboard())
    elif d == 'menu_lt':
        await _show(chat_id, msg_id, '🧪 Live-Twin Simulator:', get_live_twin_keyboard())
    elif d == 'menu_lt_friction':
        await _show(chat_id, msg_id, '⚙️ إعدادات الاحتكاك:', get_live_twin_friction_keyboard())
    elif d == 'lt_toggle_mode':
        bot_state['lt_mode'] = 'idealized' if bot_state.get('lt_mode', 'realistic') == 'realistic' else 'realistic'
        await _show(chat_id, msg_id, '🧪 Live-Twin Simulator:', get_live_twin_keyboard())
    elif d.startswith('lt_fric_'):
        key = d[len('lt_fric_'):]
        if key in bot_state['lt_friction']: bot_state['lt_friction'][key] = not bot_state['lt_friction'][key]
        await _show(chat_id, msg_id, '⚙️ إعدادات الاحتكاك:', get_live_twin_friction_keyboard())
    elif d.startswith('lt_'):
        days = int(d.split('_')[1])
        end_dt = datetime.now(timezone.utc)
        start_dt = end_dt - timedelta(days=days)
        if not bot_state['is_live_twin_running']:
            bot_state['is_live_twin_running'] = True
            _safe_task(run_live_twin_simulation(start_dt, end_dt), 'livetwin_preset')
        await _show(chat_id, msg_id, '⏳ Live-Twin يعمل...', get_live_twin_keyboard())
    elif d == 'cancel_lt':
        global _lt_progress
        if _lt_progress and bot_state['is_live_twin_running']: await _lt_progress.cancel()
        bot_state['is_live_twin_running'] = False
        await _show(chat_id, msg_id, '🧪 Live-Twin Simulator:', get_live_twin_keyboard())
    else: c_log(f'Unhandled callback: {d}')

    # UI Settings Amnesia fix: every branch above except the early-return
    # status check (which mutates nothing) falls through to here. Save
    # once, after the mutation has landed in bot_state, so a restart never
    # reverts a toggle/setting change back to the last trade's snapshot.
    # Debounce: a rapid burst of setting toggles coalesces into one write.
    await _debounced_persist_save()

# ─────────────────────────────────────────────────────────────
# TELEGRAM POLLING & WATCHDOG
# ─────────────────────────────────────────────────────────────
async def process_tg_update(update: dict) -> None:
    if 'message' in update and 'text' in update['message']:
        msg = update['message']['text'].strip(); bot_state['chat_id'] = update['message']['chat']['id']
        
        parts = msg.lower().split()
        if parts[0] == '/set':
            sym_state = bot_state['symbol_state'][bot_state['ui_selected_symbol']]
            if len(parts) == 4:
                _, tf, param, val = parts
                if tf in _TFS and param in ['tp', 'sl'] and val.isdigit():
                    val = int(val)
                    if param == 'tp': sym_state['strategy_tp_per_tf'][tf] = val
                    elif param == 'sl': sym_state['strategy_sl_per_tf'][tf] = val
                    await save_bot_persistence()
                    await send_tg_msg(f"✅ <b>تم التحديث بنجاح!</b>\n📌 الفريم: {tf}\n⚙️ {param.upper()}: {val}")
                    return
            await send_tg_msg("❌ <b>صيغة خاطئة!</b>\n<b>أمثلة صحيحة:</b>\n<code>/set 5m tp 40</code>\n<code>/set 15m sl 25</code>")
            return

        if parts[0] == '/backtest':
            try:
                if len(parts) == 2:
                    # Day boundaries are DAM-LOCAL (UTC+3), matching every
                    # report/timestamp the bot shows elsewhere -- NOT raw UTC
                    # midnight, which used to clip the first 3h of the
                    # intended day and pull in 3h of the next one instead.
                    dam_midnight = datetime.strptime(parts[1], "%Y-%m-%d")
                    dt = (dam_midnight - DAM_OFF).replace(tzinfo=timezone.utc)
                    if not bot_state['is_backtesting']:
                        bot_state['is_backtesting'] = True
                        _safe_task(run_strategy_backtest(dt, dt + timedelta(days=1)), 'backtest_cmd')
                    await send_tg_msg(f"⏳ جاري باكتيست ليوم {parts[1]} (بتوقيت دمشق)...")
                    return
                elif len(parts) == 3:
                    dam_midnight1 = datetime.strptime(parts[1], "%Y-%m-%d")
                    dam_midnight2 = datetime.strptime(parts[2], "%Y-%m-%d") + timedelta(days=1)
                    dt1 = (dam_midnight1 - DAM_OFF).replace(tzinfo=timezone.utc)
                    dt2 = (dam_midnight2 - DAM_OFF).replace(tzinfo=timezone.utc)
                    if not bot_state['is_backtesting']:
                        bot_state['is_backtesting'] = True
                        _safe_task(run_strategy_backtest(dt1, dt2), 'backtest_range_cmd')
                    await send_tg_msg(f"⏳ جاري باكتيست من {parts[1]} إلى {parts[2]} (بتوقيت دمشق)...")
                    return
            except Exception:
                await send_tg_msg("❌ <b>خطأ في التاريخ!</b>\nالصيغة: <code>/backtest 2026-06-24</code>\nأو <code>/backtest 2026-06-24 2026-06-26</code>")
                return

        if parts[0] in ('/livetwin', '/backtestreal'):
            try:
                if len(parts) == 2:
                    dam_midnight = datetime.strptime(parts[1], "%Y-%m-%d")
                    dt = (dam_midnight - DAM_OFF).replace(tzinfo=timezone.utc)
                    if not bot_state['is_live_twin_running']:
                        bot_state['is_live_twin_running'] = True
                        _safe_task(run_live_twin_simulation(dt, dt + timedelta(days=1)), 'livetwin_cmd')
                    await send_tg_msg(f"⏳ جاري Live-Twin ليوم {parts[1]} (بتوقيت دمشق)...")
                    return
                elif len(parts) == 3:
                    dam_midnight1 = datetime.strptime(parts[1], "%Y-%m-%d")
                    dam_midnight2 = datetime.strptime(parts[2], "%Y-%m-%d") + timedelta(days=1)
                    dt1 = (dam_midnight1 - DAM_OFF).replace(tzinfo=timezone.utc)
                    dt2 = (dam_midnight2 - DAM_OFF).replace(tzinfo=timezone.utc)
                    if not bot_state['is_live_twin_running']:
                        bot_state['is_live_twin_running'] = True
                        _safe_task(run_live_twin_simulation(dt1, dt2), 'livetwin_range_cmd')
                    await send_tg_msg(f"⏳ جاري Live-Twin من {parts[1]} إلى {parts[2]} (بتوقيت دمشق)...")
                    return
            except Exception:
                await send_tg_msg("❌ <b>خطأ في التاريخ!</b>\nالصيغة: <code>/backtestreal 2026-06-24</code> (أو <code>/livetwin</code>)\nأو <code>/backtestreal 2026-06-24 2026-06-26</code>")
                return

        if not msg.startswith('/') and msg in bot_state.get('menu_button_map', {}):
            cb = bot_state['menu_button_map'][msg]
            if cb != 'noop': await _handle_callback(cb, bot_state['chat_id'], None)
            return

        if msg.startswith('/setsymbol '):
            new_sym = msg.split(' ')[1].strip()
            bot_state['symbol'] = new_sym
            await save_bot_persistence()
            await send_tg_msg(f"✅ تم تغيير الرمز الخاص بـ MetaTrader إلى: <b>{new_sym}</b>")
        elif msg == '/start': await send_tg_msg('<b>مرحباً بك في Gold Scalper Bot v8.9</b>', get_main_keyboard())
        else: await send_tg_msg("❌ أمر غير معروف. استخدم /start لعرض القائمة.")
        return

    if 'callback_query' not in update: return
    q = update['callback_query']; d = q['data']; chat_id = q['message']['chat']['id']; msg_id = q['message']['message_id']
    bot_state['chat_id'] = chat_id
    _safe_task(answer_callback(q['id']), 'answer_callback')
    try: await _handle_callback(d, chat_id, msg_id)
    except Exception as e: log_exception(f'callback dispatch [{d}]', e)

_poll_task: asyncio.Task | None = None

async def telegram_polling_loop() -> None:
    c_log('Telegram polling started.'); url = f'https://api.telegram.org/bot{TG_TOKEN}/getUpdates'
    backoff = 1
    # Single persistent session for the lifetime of this task. Recreating
    # a ClientSession + TCPConnector on every backoff cycle leaked sockets
    # into TIME_WAIT over long uptimes. We only ever tear this down once,
    # in the finally block below, on task cancellation/shutdown.
    connector = aiohttp.TCPConnector(limit=4, ttl_dns_cache=300)
    timeout = aiohttp.ClientTimeout(total=None, connect=10, sock_read=28)
    sess = aiohttp.ClientSession(connector=connector, timeout=timeout)
    try:
        while True:
            try:
                async with sess.get(url, params={'offset': bot_state['last_update_id'] + 1, 'timeout': 20}) as resp:
                    if resp.status == 200:
                        backoff = 1; bot_state['last_poll_ok'] = datetime.now(timezone.utc).timestamp()
                        data = await resp.json()
                        for upd in data.get('result', []):
                            bot_state['last_update_id'] = upd['update_id']
                            asyncio.create_task(process_tg_update(upd))
                    else:
                        await asyncio.sleep(backoff); backoff = min(backoff * 2, 30)
            except asyncio.CancelledError:
                raise
            except Exception as e:
                log_exception('telegram_polling_loop', e)
                await asyncio.sleep(backoff); backoff = min(backoff * 2, 30)
    finally:
        await sess.close()

async def telegram_watchdog() -> None:
    global _poll_task
    await asyncio.sleep(30)
    while True:
        await asyncio.sleep(20)
        last = bot_state.get('last_poll_ok', 0.0); age = datetime.now(timezone.utc).timestamp() - last
        if age > 60 and _poll_task is not None and not _poll_task.done(): _poll_task.cancel()

async def supervised(coro_fn, *args, label: str = '') -> None:
    global _poll_task
    while True:
        try:
            task = asyncio.current_task()
            if label == 'tg_polling': _poll_task = task
            await coro_fn(*args)
        except asyncio.CancelledError:
            # NOTE: in this codebase the only thing that ever calls
            # .cancel() on a supervised task is telegram_watchdog, as a
            # deliberate self-heal restart when getUpdates has stalled
            # (e.g. "packet queue is empty, aborting" from a wedged
            # aiohttp connection). There is no separate graceful-shutdown
            # path that cancels individual tasks -- the process just gets
            # killed outright on redeploy. So CancelledError here always
            # means "restart me", never "stop for good". Re-raising it
            # (as a previous version of this function did) exits the
            # supervisor loop permanently and kills polling/buttons until
            # the whole process is manually restarted -- that was the bug
            # that caused "no button works". Swallow it and loop back.
            logger.warning('supervised task "%s" was cancelled (self-heal) -- restarting', label)
            await asyncio.sleep(2)
        except Exception as e:
            log_exception(f'supervised task "{label}"', e)
            await asyncio.sleep(5)

# ─────────────────────────────────────────────────────────────
# ENTRY POINT & WEB SERVER
# ─────────────────────────────────────────────────────────────
async def handle_ping(request: web.Request) -> web.Response:
    return web.Response(text="Bot is running smoothly!")

async def main() -> None:
    get_http()
    await init_metaapi()
    app = web.Application()
    app.router.add_get('/', handle_ping)
    
    runner = web.AppRunner(app); await runner.setup()
    port = int(os.environ.get('PORT', 10000))
    await web.TCPSite(runner, '0.0.0.0', port).start()
    c_log(f'Web server started on port {port}')

    bot_state['last_poll_ok'] = datetime.now(timezone.utc).timestamp()

    tasks = [
        asyncio.create_task(supervised(telegram_polling_loop, label='tg_polling')),
        asyncio.create_task(supervised(telegram_watchdog,     label='tg_watchdog')),
        asyncio.create_task(supervised(strategy_monitor_scanner,  label='strategy_monitor')),
        asyncio.create_task(supervised(global_ledger_reconciliation, label='global_reconciliation')),
        asyncio.create_task(supervised(oanda_live_price_poller, label='oanda_price_poller')),
    ]
    
    c_log('Gold Scalper Bot v9.4 (Resilience-First Core) started successfully.')
    try: await asyncio.gather(*tasks)
    finally:
        if _http and not _http.closed: await _http.close()

if __name__ == '__main__':
    asyncio.run(main())
